import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import os
import sys
from datetime import datetime, timedelta
import hashlib
import secrets
from contextlib import contextmanager
from functools import partial
from PIL import Image, ImageTk, ImageDraw
import json
import csv
import calendar
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 获取exe所在目录
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 数据目录
DATA_DIR = os.path.join(BASE_DIR, 'data')
BACKUP_DIR = os.path.join(DATA_DIR, 'bak')

# 确保目录存在
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

# 数据库文件路径（新路径在data目录）
DB_FILE = os.path.join(DATA_DIR, 'inventory.db')
# 旧数据库路径（用于迁移）
OLD_DB_FILE = os.path.join(BASE_DIR, 'inventory.db')

# 如果旧数据库存在但新数据库不存在，则迁移旧数据库
if os.path.exists(OLD_DB_FILE) and not os.path.exists(DB_FILE):
    try:
        import shutil
        shutil.copy2(OLD_DB_FILE, DB_FILE)
        # 可选：备份旧数据库后删除
        # os.remove(OLD_DB_FILE)
    except Exception as e:
        print(f"迁移旧数据库时出错: {e}")

SETTINGS_FILE = os.path.join(BASE_DIR, 'settings.json')

# ==================== 密码哈希函数 ====================
def hash_password(password: str) -> str:
    """使用PBKDF2哈希密码（比MD5安全，无需额外依赖）"""
    salt = secrets.token_hex(16)  # 生成随机盐
    key = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100000)
    return f"{salt}:{key.hex()}"

def verify_password(password: str, hashed: str) -> bool:
    """验证密码（兼容旧MD5格式）"""
    try:
        if ':' not in hashed:
            # 旧格式（MD5），兼容处理
            md5_hash = hashlib.md5(password.encode()).hexdigest()
            return md5_hash == hashed
        
        salt, key_hex = hashed.split(':', 1)
        key = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100000)
        return key.hex() == key_hex
    except:
        return False

# ==================== 数据库连接上下文管理器 ====================
@contextmanager
def get_db_connection():
    """数据库连接上下文管理器，自动处理事务提交/回滚和连接关闭"""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row  # 返回字典式结果
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

# ==================== 数据库索引创建函数 ====================
def add_database_indexes(cursor):
    """添加常用索引以提高查询速度"""
    indexes = [
        # 商品相关索引
        "CREATE INDEX IF NOT EXISTS idx_products_code ON products(code)",
        "CREATE INDEX IF NOT EXISTS idx_products_name ON products(name)",
        "CREATE INDEX IF NOT EXISTS idx_products_status ON products(status)",
        
        # 库存相关索引
        "CREATE INDEX IF NOT EXISTS idx_inventory_product ON inventory(product_id)",
        "CREATE INDEX IF NOT EXISTS idx_inventory_warehouse ON inventory(warehouse_id)",
        "CREATE INDEX IF NOT EXISTS idx_inventory_product_warehouse ON inventory(product_id, warehouse_id)",
        
        # 入库单索引
        "CREATE INDEX IF NOT EXISTS idx_inbound_order_date ON inbound_orders(order_date)",
        "CREATE INDEX IF NOT EXISTS idx_inbound_order_no ON inbound_orders(order_no)",
        "CREATE INDEX IF NOT EXISTS idx_inbound_detail_order ON inbound_details(order_id)",
        "CREATE INDEX IF NOT EXISTS idx_inbound_detail_product ON inbound_details(product_id)",
        
        # 出库单索引
        "CREATE INDEX IF NOT EXISTS idx_outbound_order_date ON outbound_orders(order_date)",
        "CREATE INDEX IF NOT EXISTS idx_outbound_order_no ON outbound_orders(order_no)",
        "CREATE INDEX IF NOT EXISTS idx_outbound_detail_order ON outbound_details(order_id)",
        "CREATE INDEX IF NOT EXISTS idx_outbound_detail_product ON outbound_details(product_id)",
    ]
    
    for index_sql in indexes:
        try:
            cursor.execute(index_sql)
        except sqlite3.OperationalError:
            # 索引可能已存在或其他原因，忽略
            pass
        except Exception as e:
            print(f"创建索引失败: {e}")

class LoginWindow:
    """登录窗口"""
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("进销存管理系统 - 登录")
        
        # 先隐藏窗口，避免在左上角闪烁
        self.root.withdraw()
        
        # 设置窗口大小（不设置位置）
        self.root.resizable(False, False)
        
        # 设置背景
        self.setup_background()

        # 创建登录界面
        self.create_login_form()

        # 初始化数据库
        self.init_database()
        
        # 居中窗口（在内容加载完成后）
        self.center_window()
        
        # 显示窗口
        self.root.deiconify()

    def center_window(self):
        """居中窗口"""
        self.root.update_idletasks()  # 确保窗口内容已渲染
        width = 400
        height = 500
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        # 一次性设置大小和位置
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def setup_background(self):
        """设置渐变背景"""
        canvas = tk.Canvas(self.root, width=400, height=500, highlightthickness=0)
        canvas.pack(fill='both', expand=True)

        # 创建渐变
        for i in range(500):
            gray = int(30 + (50 * i / 500))
            color = f'#{gray:02x}{gray+20:02x}{gray+40:02x}'
            canvas.create_line(0, i, 400, i, fill=color)

        # 保存canvas引用
        self.canvas = canvas

    def create_login_form(self):
        """创建登录表单"""
        # Logo区域
        logo_frame = tk.Frame(self.root, bg='#2C3E50')
        logo_frame.place(relx=0.5, rely=0.15, anchor='center')

        # 创建图标
        self.create_logo(logo_frame)

        # 登录框
        login_frame = tk.Frame(self.root, bg='white', relief='raised', bd=1)
        login_frame.place(relx=0.5, rely=0.5, anchor='center', width=300, height=250)

        # 标题
        tk.Label(login_frame, text="进销存管理系统",
                font=('微软雅黑', 18, 'bold'),
                bg='white', fg='#2C3E50').pack(pady=(20, 30))

        # 用户名
        username_frame = tk.Frame(login_frame, bg='white')
        username_frame.pack(pady=5)
        tk.Label(username_frame, text="用户名：", font=('微软雅黑', 10),
                bg='white', width=8, anchor='e').pack(side='left')
        self.username_entry = tk.Entry(username_frame, font=('微软雅黑', 10), width=20)
        self.username_entry.pack(side='left')
        self.username_entry.insert(0, 'admin')

        # 密码
        password_frame = tk.Frame(login_frame, bg='white')
        password_frame.pack(pady=5)
        tk.Label(password_frame, text="密码：", font=('微软雅黑', 10),
                bg='white', width=8, anchor='e').pack(side='left')
        self.password_entry = tk.Entry(password_frame, font=('微软雅黑', 10), width=20, show='*')
        self.password_entry.pack(side='left')
        self.password_entry.insert(0, '123456')

        # 记住密码
        self.remember_var = tk.BooleanVar(value=True)
        tk.Checkbutton(login_frame, text="记住密码", variable=self.remember_var,
                      font=('微软雅黑', 9), bg='white').pack(pady=10)

        # 登录按钮
        login_btn = tk.Button(login_frame, text="登  录", font=('微软雅黑', 12, 'bold'),
                            bg='#3498DB', fg='white', width=20, height=1,
                            command=self.login, cursor='hand2')
        login_btn.pack(pady=10)

        # 提示
        tk.Label(login_frame, text="默认账号：admin  密码：123456",
                font=('微软雅黑', 9), bg='white', fg='#7F8C8D').pack()

    def create_logo(self, parent):
        """创建Logo"""
        # 创建一个简单的图标
        img = Image.new('RGBA', (80, 80), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)

        # 绘制一个仓库图标
        draw.rectangle([10, 30, 70, 60], fill='#3498DB', outline='#2980B9', width=2)
        draw.polygon([(10, 30), (40, 10), (70, 30)], fill='#E74C3C', outline='#C0392B', width=2)
        draw.rectangle([30, 40, 50, 60], fill='#F39C12', outline='#D68910', width=2)

        logo = ImageTk.PhotoImage(img)
        label = tk.Label(parent, image=logo, bg='#2C3E50')  # 修复背景色
        label.image = logo  # 保持引用
        label.pack()

    def init_database(self):
        """初始化数据库"""
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        # 用户表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                role TEXT DEFAULT 'user',
                permissions TEXT,
                created_time DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 如果permissions列不存在，添加该列
        try:
            cursor.execute("ALTER TABLE users ADD COLUMN permissions TEXT")
        except sqlite3.OperationalError:
            # 列已存在，忽略错误
            pass

        # 检查是否有默认用户
        cursor.execute("SELECT * FROM users WHERE username='admin'")
        if not cursor.fetchone():
            # 创建默认管理员账号（使用新的PBKDF2哈希）
            password = hash_password('123456')
            cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                          ('admin', password, 'admin'))

        # 商品分类表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                parent_id INTEGER,
                sort_order INTEGER DEFAULT 0,
                FOREIGN KEY (parent_id) REFERENCES categories (id)
            )
        ''')

        # 商品表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                category_id INTEGER,
                specification TEXT,
                unit TEXT DEFAULT '个',
                purchase_price DECIMAL(10, 2),
                sale_price DECIMAL(10, 2),
                min_stock INTEGER DEFAULT 0,
                max_stock INTEGER DEFAULT 999999,
                status INTEGER DEFAULT 1,
                created_time DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_time DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (category_id) REFERENCES categories (id)
            )
        ''')

        # 仓库表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS warehouses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                address TEXT,
                manager TEXT,
                contact TEXT,
                status INTEGER DEFAULT 1
            )
        ''')

        # 供应商表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                contact_person TEXT,
                phone TEXT,
                address TEXT,
                email TEXT,
                status INTEGER DEFAULT 1
            )
        ''')

        # 客户表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                contact_person TEXT,
                phone TEXT,
                address TEXT,
                email TEXT,
                status INTEGER DEFAULT 1
            )
        ''')

        # 入库单表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS inbound_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no TEXT UNIQUE NOT NULL,
                warehouse_id INTEGER,
                supplier TEXT,
                order_date DATE NOT NULL,
                total_amount DECIMAL(10, 2) DEFAULT 0,
                operator TEXT,
                status INTEGER DEFAULT 0,
                created_time DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (warehouse_id) REFERENCES warehouses (id)
            )
        ''')

        # 入库单明细表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS inbound_details (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER NOT NULL,
                product_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL,
                price DECIMAL(10, 2) NOT NULL,
                amount DECIMAL(10, 2) NOT NULL,
                batch_no TEXT,
                production_date DATE,
                expire_date DATE,
                FOREIGN KEY (order_id) REFERENCES inbound_orders (id),
                FOREIGN KEY (product_id) REFERENCES products (id)
            )
        ''')

        # 出库单表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS outbound_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no TEXT UNIQUE NOT NULL,
                warehouse_id INTEGER,
                customer TEXT,
                order_date DATE NOT NULL,
                total_amount DECIMAL(10, 2) DEFAULT 0,
                operator TEXT,
                status INTEGER DEFAULT 0,
                created_time DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (warehouse_id) REFERENCES warehouses (id)
            )
        ''')

        # 出库单明细表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS outbound_details (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER NOT NULL,
                product_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL,
                price DECIMAL(10, 2) NOT NULL,
                amount DECIMAL(10, 2) NOT NULL,
                batch_no TEXT,
                FOREIGN KEY (order_id) REFERENCES outbound_orders (id),
                FOREIGN KEY (product_id) REFERENCES products (id)
            )
        ''')

        # 库存表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS inventory (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                warehouse_id INTEGER NOT NULL,
                quantity INTEGER DEFAULT 0,
                available_quantity INTEGER DEFAULT 0,
                batch_no TEXT,
                updated_time DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (product_id) REFERENCES products (id),
                FOREIGN KEY (warehouse_id) REFERENCES warehouses (id),
                UNIQUE(product_id, warehouse_id, batch_no)
            )
        ''')

        # 不再自动插入示例数据，用户需要手动添加数据

        # 添加数据库索引以提高查询性能
        add_database_indexes(cursor)

        conn.commit()
        conn.close()

    def login(self):
        """登录验证"""
        username = self.username_entry.get()
        password = self.password_entry.get()

        if not username or not password:
            messagebox.showerror("错误", "请输入用户名和密码")
            return

        # 验证用户（使用新的密码验证函数，兼容旧MD5格式）
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        cursor.execute("SELECT id, username, role, password FROM users WHERE username=?",
                      (username,))
        user_data = cursor.fetchone()
        conn.close()

        if user_data and verify_password(password, user_data[3]):
            # 如果用户使用旧MD5密码，登录成功后可以升级为新格式（可选）
            user = (user_data[0], user_data[1], user_data[2])
            # 保存登录信息
            if self.remember_var.get():
                self.save_login_info(username)

            # 打开主窗口
            self.root.destroy()
            main_window = MainWindow(user)
            main_window.run()
        else:
            messagebox.showerror("错误", "用户名或密码错误")

    def save_login_info(self, username):
        """保存登录信息"""
        try:
            with open(os.path.join(BASE_DIR, 'login.json'), 'w') as f:
                json.dump({'username': username}, f)
        except:
            pass

    def run(self):
        """运行登录窗口"""
        self.root.mainloop()

class MainWindow:
    """主窗口"""
    def __init__(self, user):
        self.user = user
        self.root = tk.Tk()
        self.root.title(f"进销存管理系统 - {user[1]} ({user[2]})")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 600)

        # 设置主题
        style = ttk.Style()
        style.theme_use('clam')

        # 自定义样式
        self.setup_styles(style)

        # 创建界面
        self.create_ui()

        # 绑定窗口关闭事件（点击X按钮时调用logout方法）
        self.root.protocol("WM_DELETE_WINDOW", self.logout)

        # 创建菜单并设置初始选中状态
        current_page = "dashboard"
        self.current_page = current_page
        self.show_page(current_page)

    def setup_styles(self, style):
        """设置界面样式"""
        # 配置按钮样式
        style.configure('Menu.TButton',
                       font=('微软雅黑', 10),
                       padding=(10, 5))

        # 配置框架样式
        style.configure('Card.TFrame',
                       relief='raised',
                       borderwidth=1)

    def show_date_picker(self, entry_widget):
        """显示日期选择器（现代化设计，类似Ant Design风格）"""
        # 获取当前日期值
        current_value = entry_widget.get()
        try:
            if current_value:
                current_date = datetime.strptime(current_value, '%Y-%m-%d')
            else:
                current_date = datetime.now()
        except:
            current_date = datetime.now()

        # 创建日期选择窗口（更大的尺寸，更美观）
        date_win = tk.Toplevel(self.root)
        date_win.title("")
        date_win.geometry("330x400")
        date_win.transient(self.root)
        date_win.grab_set()
        date_win.resizable(False, False)
        date_win.configure(bg='#FFFFFF')

        # 居中（相对于输入框位置）
        try:
            entry_widget.update_idletasks()
            x = entry_widget.winfo_rootx()
            y = entry_widget.winfo_rooty() + entry_widget.winfo_height() + 5
            # 确保窗口不超出屏幕
            screen_width = date_win.winfo_screenwidth()
            screen_height = date_win.winfo_screenheight()
            if x + 330 > screen_width:
                x = screen_width - 330 - 10
            if y + 400 > screen_height:
                y = entry_widget.winfo_rooty() - 400 - 5
            if x < 0:
                x = 10
            if y < 0:
                y = 10
            date_win.geometry(f'330x400+{x}+{y}')
        except:
            # 如果获取位置失败，使用屏幕居中
            date_win.update_idletasks()
            x = (date_win.winfo_screenwidth() // 2) - 165
            y = (date_win.winfo_screenheight() // 2) - 200
            date_win.geometry(f'330x400+{x}+{y}')

        # 主容器（带边框阴影效果）
        main_frame = tk.Frame(date_win, bg='#FFFFFF', relief='solid', bd=1)
        main_frame.pack(fill='both', expand=True, padx=0, pady=0)

        # 顶部工具栏（年份月份导航）
        toolbar_frame = tk.Frame(main_frame, bg='#FAFAFA', height=55)
        toolbar_frame.pack(fill='x')
        toolbar_frame.pack_propagate(False)

        year_var = tk.IntVar(value=current_date.year)
        month_var = tk.IntVar(value=current_date.month)

        # 左箭头按钮（上一月）
        def prev_month():
            m = month_var.get() - 1
            if m < 1:
                month_var.set(12)
                year_var.set(year_var.get() - 1)
            else:
                month_var.set(m)
            update_calendar()

        prev_btn = tk.Button(toolbar_frame, text="◀", font=('Arial', 14, 'bold'),
                            bg='#FAFAFA', fg='#595959', relief='flat', cursor='hand2',
                            width=3, command=prev_month, activebackground='#E6F7FF')
        prev_btn.pack(side='left', padx=8, pady=12)

        # 年月显示
        nav_frame = tk.Frame(toolbar_frame, bg='#FAFAFA')
        nav_frame.pack(side='left', expand=True)

        year_label = tk.Label(nav_frame, textvariable=year_var, font=('微软雅黑', 14, 'bold'),
                             bg='#FAFAFA', fg='#262626')
        year_label.pack(side='left', padx=3)

        month_label_text = tk.StringVar()
        month_label_text.set(f"{month_var.get()}月")
        month_label = tk.Label(nav_frame, textvariable=month_label_text, font=('微软雅黑', 14, 'bold'),
                              bg='#FAFAFA', fg='#262626')
        month_label.pack(side='left', padx=3)

        def update_month_label():
            month_label_text.set(f"{month_var.get()}月")

        # 右箭头按钮（下一月）
        def next_month():
            m = month_var.get() + 1
            if m > 12:
                month_var.set(1)
                year_var.set(year_var.get() + 1)
            else:
                month_var.set(m)
            update_calendar()

        next_btn = tk.Button(toolbar_frame, text="▶", font=('Arial', 14, 'bold'),
                            bg='#FAFAFA', fg='#595959', relief='flat', cursor='hand2',
                            width=3, command=next_month, activebackground='#E6F7FF')
        next_btn.pack(side='right', padx=8, pady=12)

        # 今天按钮
        today_btn = tk.Button(toolbar_frame, text="今天", font=('微软雅黑', 10),
                             bg='#FAFAFA', fg='#1890FF', relief='flat', cursor='hand2',
                             command=lambda: select_date(datetime.now().strftime('%Y-%m-%d')))
        today_btn.pack(side='right', padx=8, pady=12)

        # 日期选择区域
        date_container = tk.Frame(main_frame, bg='#FFFFFF')
        date_container.pack(pady=15, padx=15, fill='both', expand=True)

        # 星期标题行
        weekdays = ['日', '一', '二', '三', '四', '五', '六']
        weekday_frame = tk.Frame(date_container, bg='#FFFFFF')
        weekday_frame.pack(fill='x', pady=(0, 8))
        for i, day in enumerate(weekdays):
            color = '#FF4D4F' if i == 0 or i == 6 else '#8C8C8C'  # 周末红色
            label = tk.Label(weekday_frame, text=day, font=('微软雅黑', 11),
                           bg='#FFFFFF', fg=color, width=6)
            label.pack(side='left', expand=True)

        # 日期按钮容器
        date_buttons = []
        calendar_frame = tk.Frame(date_container, bg='#FFFFFF')
        calendar_frame.pack(fill='both', expand=True)

        def select_date(date_value):
            """选择日期并关闭窗口"""
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, date_value)
            date_win.destroy()

        def update_calendar():
            # 清空现有按钮
            for widget in calendar_frame.winfo_children():
                widget.destroy()
            date_buttons.clear()

            year = year_var.get()
            month = month_var.get()
            update_month_label()
            
            # 获取当月第一天是星期几和总天数
            first_day, num_days = calendar.monthrange(year, month)
            today = datetime.now()
            selected_date_str = current_value if current_value else None
            
            # 上个月的日期（灰色显示）
            prev_year = year if month > 1 else year - 1
            prev_month_num = month - 1 if month > 1 else 12
            prev_month_days = calendar.monthrange(prev_year, prev_month_num)[1]
            start_day = prev_month_days - first_day + 1
            
            row = 0
            col = 0
            
            # 填充上个月的日期（灰色）
            for day in range(start_day, prev_month_days + 1):
                btn = tk.Button(calendar_frame, text=str(day), font=('微软雅黑', 11),
                               bg='#FFFFFF', fg='#BFBFBF', relief='flat',
                               width=6, height=2, cursor='hand2',
                               activebackground='#F5F5F5', borderwidth=0,
                               command=lambda d=day, y=prev_year, m=prev_month_num: 
                               select_date(f"{y}-{m:02d}-{d:02d}"))
                btn.grid(row=row, column=col, padx=1, pady=1, sticky='nsew')
                date_buttons.append(btn)
                col += 1
                if col > 6:
                    col = 0
                    row += 1
            
            # 当前月的日期
            for day in range(1, num_days + 1):
                date_str = f"{year}-{month:02d}-{day:02d}"
                
                # 判断是否是今天、选中日期
                is_today = (year == today.year and month == today.month and day == today.day)
                is_selected = (date_str == selected_date_str)
                
                # 设置按钮样式（Ant Design风格）
                if is_selected:
                    bg_color = '#1890FF'
                    fg_color = '#FFFFFF'
                    active_bg = '#40A9FF'
                elif is_today:
                    bg_color = '#E6F7FF'
                    fg_color = '#1890FF'
                    active_bg = '#BAE7FF'
                else:
                    bg_color = '#FFFFFF'
                    fg_color = '#262626'
                    active_bg = '#F5F5F5'
                
                btn = tk.Button(calendar_frame, text=str(day), font=('微软雅黑', 11),
                               bg=bg_color, fg=fg_color, relief='flat',
                               width=6, height=2, cursor='hand2',
                               activebackground=active_bg, activeforeground=fg_color if not is_selected else '#FFFFFF',
                               borderwidth=0,
                               command=lambda d=date_str: select_date(d))
                
                # 如果是今天，添加边框
                if is_today and not is_selected:
                    btn.config(highlightbackground='#1890FF', highlightthickness=1)
                
                btn.grid(row=row, column=col, padx=1, pady=1, sticky='nsew')
                date_buttons.append(btn)
                
                col += 1
                if col > 6:
                    col = 0
                    row += 1
            
            # 填充下个月的日期（灰色）
            next_year = year if month < 12 else year + 1
            next_month_num = month + 1 if month < 12 else 1
            next_day = 1
            while row < 6:  # 最多显示6行
                btn = tk.Button(calendar_frame, text=str(next_day), font=('微软雅黑', 11),
                               bg='#FFFFFF', fg='#BFBFBF', relief='flat',
                               width=6, height=2, cursor='hand2',
                               activebackground='#F5F5F5', borderwidth=0,
                               command=lambda d=next_day, y=next_year, m=next_month_num: 
                               select_date(f"{y}-{m:02d}-{d:02d}"))
                btn.grid(row=row, column=col, padx=1, pady=1, sticky='nsew')
                date_buttons.append(btn)
                col += 1
                if col > 6:
                    col = 0
                    row += 1
                next_day += 1
            
            # 配置网格权重
            for i in range(7):
                calendar_frame.grid_columnconfigure(i, weight=1, uniform='col')
            for i in range(6):
                calendar_frame.grid_rowconfigure(i, weight=1, uniform='row')

        # 绑定年份和月份变化事件
        year_var.trace_add('write', lambda *args: update_calendar())
        month_var.trace_add('write', lambda *args: update_calendar())

        # 初始显示日历
        update_calendar()

        # 底部快速选择区域
        quick_frame = tk.Frame(main_frame, bg='#FAFAFA', height=45)
        quick_frame.pack(fill='x', side='bottom')
        quick_frame.pack_propagate(False)

        quick_btns = [
            ("今天", datetime.now().strftime('%Y-%m-%d')),
            ("昨天", (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')),
            ("一周前", (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')),
            ("一月前", (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        ]

        btn_container = tk.Frame(quick_frame, bg='#FAFAFA')
        btn_container.pack(expand=True)
        
        for text, date_val in quick_btns:
            btn = tk.Button(btn_container, text=text, font=('微软雅黑', 10),
                           bg='#FAFAFA', fg='#1890FF', relief='flat', cursor='hand2',
                           activebackground='#E6F7FF',
                           command=lambda d=date_val: select_date(d))
            btn.pack(side='left', padx=8, pady=8)

    def create_ui(self):
        """创建主界面"""
        # 创建顶部栏
        self.create_top_bar()

        # 创建主体容器
        main_container = tk.Frame(self.root, bg='#F0F2F5')
        main_container.pack(fill='both', expand=True)

        # 左侧菜单
        self.create_left_menu(main_container)

        # 右侧内容区
        self.create_content_area(main_container)

    def create_top_bar(self):
        """创建顶部栏"""
        top_bar = tk.Frame(self.root, bg='#2C3E50', height=60)
        top_bar.pack(fill='x')
        top_bar.pack_propagate(False)

        # Logo和标题
        left_frame = tk.Frame(top_bar, bg='#2C3E50')
        left_frame.pack(side='left', padx=20, pady=10)

        tk.Label(left_frame, text="📦", font=('Arial', 24),
                bg='#2C3E50', fg='white').pack(side='left')
        tk.Label(left_frame, text="进销存管理系统",
                font=('微软雅黑', 18, 'bold'),
                bg='#2C3E50', fg='white').pack(side='left', padx=10)

        # 用户信息和时间
        right_frame = tk.Frame(top_bar, bg='#2C3E50')
        right_frame.pack(side='right', padx=20, pady=10)

        # 时间标签
        self.time_label = tk.Label(right_frame, font=('微软雅黑', 10),
                                  bg='#2C3E50', fg='white')
        self.time_label.pack(side='right', padx=20)

        # 用户信息
        tk.Label(right_frame, text=f"用户：{self.user[1]}",
                font=('微软雅黑', 10),
                bg='#2C3E50', fg='white').pack(side='right')

        # 更新时间
        self.update_time()

    def get_user_permissions(self, username):
        """获取用户权限列表"""
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("SELECT permissions FROM users WHERE username=?", (username,))
            result = cursor.fetchone()
            conn.close()
            
            if result and result[0]:
                try:
                    return json.loads(result[0])
                except:
                    return []
            return []  # 默认无权限（管理员返回空列表表示全部权限）
        except:
            return []
    
    def has_permission(self, username, permission):
        """检查用户是否有某个权限"""
        # 管理员默认拥有所有权限
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("SELECT role FROM users WHERE username=?", (username,))
            result = cursor.fetchone()
            conn.close()
            if result and result[0] == 'admin':
                return True
        except:
            pass
        
        # 检查权限列表
        permissions = self.get_user_permissions(username)
        # 如果权限列表为空，默认无权限；如果包含该权限，返回True
        return len(permissions) > 0 and permission in permissions

    def create_left_menu(self, parent):
        """创建左侧菜单（根据用户权限显示）"""
        menu_frame = tk.Frame(parent, bg='#34495E', width=250)
        menu_frame.pack(side='left', fill='y')
        menu_frame.pack_propagate(False)

        # 菜单标题
        tk.Label(menu_frame, text="功能菜单",
                font=('微软雅黑', 14, 'bold'),
                bg='#34495E', fg='white').pack(pady=20)

        # 菜单按钮字典，用于保存按钮引用
        self.menu_buttons = {}

        # 菜单按钮（权限key）
        menus = [
            ('dashboard', '📊', '仪表盘', self.show_dashboard),
            ('inbound', '📥', '入库管理', self.show_inbound),
            ('outbound', '📤', '出库管理', self.show_outbound),
            ('inventory', '📦', '库存查询', self.show_inventory),
            ('basic_management', '📋', '基础管理', self.show_basic_management),
            ('reports', '📈', '统计报表', self.show_reports),
            ('settings', '⚙️', '系统设置', self.show_settings)
        ]

        # 获取当前用户的权限
        current_user = self.user[1]  # username
        user_permissions = self.get_user_permissions(current_user)
        is_admin = self.user[2] == 'admin'  # role
        
        for key, icon, text, command in menus:
            # 检查权限：管理员拥有所有权限，普通用户检查权限列表
            if not is_admin:
                if len(user_permissions) == 0:  # 权限列表为空，无权限
                    continue
                # 特殊处理：基础管理需要至少有一个子权限（products、suppliers、customers、warehouses）
                if key == 'basic_management':
                    has_basic_permission = any(p in user_permissions for p in ['products', 'suppliers', 'customers', 'warehouses'])
                    if not has_basic_permission:
                        continue
                elif key not in user_permissions:  # 不在权限列表中，跳过
                    continue
            
            btn = tk.Button(menu_frame,
                          text=f"  {icon}  {text}",
                          font=('微软雅黑', 11),
                          bg='#34495E', fg='white',
                          activebackground='#2980B9',
                          activeforeground='white',
                          relief='flat',
                          anchor='w',
                          pady=12,
                          command=lambda k=key: self.show_page(k))
            btn.pack(fill='x', padx=10, pady=2)
            
            # 保存按钮引用
            self.menu_buttons[key] = btn

            # 绑定鼠标悬停事件
            def make_hover(button, key_name):
                def on_enter(e):
                    # 只有非选中状态才改变颜色
                    if key_name != getattr(self, 'current_page', None):
                        button.config(bg='#2980B9')
                def on_leave(e):
                    # 只有非选中状态才恢复颜色
                    if key_name != getattr(self, 'current_page', None):
                        button.config(bg='#34495E')
                return on_enter, on_leave
            
            on_enter, on_leave = make_hover(btn, key)
            btn.bind('<Enter>', on_enter)
            btn.bind('<Leave>', on_leave)

        # 底部退出按钮
        tk.Frame(menu_frame, bg='#2C3E50', height=1).pack(fill='x', pady=20)

        logout_btn = tk.Button(menu_frame,
                             text="  🚪  退出系统",
                             font=('微软雅黑', 11),
                             bg='#E74C3C', fg='white',
                             activebackground='#C0392B',
                             relief='flat',
                             anchor='w',
                             pady=12,
                             command=self.logout)
        logout_btn.pack(fill='x', padx=10, pady=10)

        logout_btn.bind('<Enter>', lambda e: logout_btn.config(bg='#C0392B'))
        logout_btn.bind('<Leave>', lambda e: logout_btn.config(bg='#E74C3C'))

    def create_content_area(self, parent):
        """创建右侧内容区（使用Tab页）"""
        self.content_frame = tk.Frame(parent, bg='#F0F2F5')
        self.content_frame.pack(side='right', fill='both', expand=True)

        # 创建Tab页容器
        self.notebook = ttk.Notebook(self.content_frame)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # 绑定Tab页关闭事件
        # 方法1：右键菜单
        self.notebook.bind('<Button-3>', self.on_tab_right_click)
        # 方法2：中键点击关闭
        self.notebook.bind('<Button-2>', self.on_tab_middle_click)
        # 方法3：双击Tab标签区域关闭当前Tab
        def on_notebook_double_click(event):
            """双击Tab标签区域关闭当前Tab"""
            try:
                # Tab标签通常在Notebook的上方区域（约30-40像素）
                # 检测点击位置是否在Tab标签区域
                if event.y < 35:  # Tab标签区域大约在顶部35像素内
                    current_tab = self.notebook.index(self.notebook.select())
                    if current_tab >= 0:
                        self.close_tab(current_tab)
            except:
                pass
        
        self.notebook.bind('<Double-Button-1>', on_notebook_double_click)
        
        # 方法4：添加键盘快捷键 Ctrl+W 关闭当前Tab
        self.root.bind('<Control-w>', lambda e: self.close_current_tab())

        # 页面字典，保存每个Tab页的Frame和内容
        self.pages = {}
        
        # Tab页标题映射
        self.page_titles = {
            'dashboard': '📊 仪表盘',
            'inbound': '📥 入库管理',
            'outbound': '📤 出库管理',
            'inventory': '📦 库存查询',
            'basic_management': '📋 基础管理',
            'products': '📝 商品管理',
            'suppliers': '🏢 供应商管理',
            'customers': '👥 客户管理',
            'warehouses': '🏭 仓库管理',
            'reports': '📈 统计报表',
            'settings': '⚙️ 系统设置'
        }

    def show_page(self, page_name):
        """显示指定页面（Tab页方式）"""
        # 更新菜单按钮的选中状态
        if hasattr(self, 'menu_buttons'):
            # 恢复所有菜单按钮的默认颜色
            for key, btn in self.menu_buttons.items():
                btn.config(bg='#34495E')
            
            # 设置当前选中菜单项的颜色
            if page_name in self.menu_buttons:
                self.menu_buttons[page_name].config(bg='#2980B9')
        
        # 保存当前页面
        self.current_page = page_name
        
        # 检查Tab页是否已存在
        if page_name in self.pages:
            # Tab页已存在，切换到该Tab页
            self.notebook.select(self.pages[page_name]['frame'])
        else:
            # 创建新的Tab页
            tab_frame = tk.Frame(self.notebook, bg='#F0F2F5')
            tab_title = self.page_titles.get(page_name, page_name)
            
            # 添加Tab页，如果非仪表盘或已有其他Tab，在标题后添加关闭符号提示
            # 注意：ttk.Notebook的Tab标签不支持自定义控件，所以我们在标题中添加提示
            display_title = f"{tab_title} [×]" if (page_name != 'dashboard' or len(self.notebook.tabs()) > 0) else tab_title
            self.notebook.add(tab_frame, text=display_title)
            
            # 保存Tab页信息
            self.pages[page_name] = {
                'frame': tab_frame,
                'title': tab_title,
                'page_name': page_name
            }
            
            # 切换到新创建的Tab页
            self.notebook.select(tab_frame)
            
            # 根据页面名称显示对应内容（传入tab_frame作为容器）
            if page_name == 'dashboard':
                self.show_dashboard(tab_frame)
            elif page_name == 'inbound':
                self.show_inbound(tab_frame)
            elif page_name == 'outbound':
                self.show_outbound(tab_frame)
            elif page_name == 'inventory':
                self.show_inventory(tab_frame)
            elif page_name == 'basic_management':
                self.show_basic_management(tab_frame)
            elif page_name == 'products':
                self.show_products(tab_frame)
            elif page_name == 'suppliers':
                self.show_suppliers(tab_frame)
            elif page_name == 'customers':
                self.show_customers(tab_frame)
            elif page_name == 'reports':
                self.show_reports(tab_frame)
            elif page_name == 'settings':
                self.show_settings(tab_frame)

    def on_tab_right_click(self, event):
        """Tab页右键点击事件"""
        try:
            # 尝试获取点击的Tab页索引
            tab_index = self.notebook.index(f"@{event.x},{event.y}")
            if tab_index >= 0:
                # 获取当前选中的Tab
                current_tab = self.notebook.index(self.notebook.select())
                # 如果点击的是当前Tab，使用当前Tab的索引
                if tab_index == current_tab or tab_index < len(self.notebook.tabs()):
                    # 创建右键菜单
                    menu = tk.Menu(self.root, tearoff=0)
                    menu.add_command(label="关闭", command=lambda idx=tab_index: self.close_tab(idx))
                    menu.add_command(label="关闭其他", command=lambda idx=tab_index: self.close_other_tabs(idx))
                    menu.add_command(label="关闭所有", command=self.close_all_tabs)
                    try:
                        menu.tk_popup(event.x_root, event.y_root)
                    finally:
                        menu.grab_release()
        except:
            # 如果获取索引失败，尝试使用当前选中的Tab
            try:
                current_tab = self.notebook.index(self.notebook.select())
                menu = tk.Menu(self.root, tearoff=0)
                menu.add_command(label="关闭", command=lambda: self.close_tab(current_tab))
                menu.add_command(label="关闭其他", command=lambda: self.close_other_tabs(current_tab))
                menu.add_command(label="关闭所有", command=self.close_all_tabs)
                try:
                    menu.tk_popup(event.x_root, event.y_root)
                finally:
                    menu.grab_release()
            except:
                pass

    def on_tab_middle_click(self, event):
        """Tab页中键点击事件（直接关闭）"""
        try:
            tab_index = self.notebook.index(f"@{event.x},{event.y}")
            if tab_index >= 0:
                self.close_tab(tab_index)
        except:
            # 如果获取索引失败，使用当前选中的Tab
            try:
                current_tab = self.notebook.index(self.notebook.select())
                self.close_tab(current_tab)
            except:
                pass
    
    def close_current_tab(self):
        """关闭当前选中的Tab页"""
        try:
            current_tab = self.notebook.index(self.notebook.select())
            if current_tab >= 0:
                self.close_tab(current_tab)
        except:
            pass
    
    def close_tab_by_name(self, page_name):
        """通过页面名称关闭Tab页"""
        if page_name not in self.pages:
            return
        
        # 如果关闭的是仪表盘，不允许关闭（至少保留一个Tab页）
        if page_name == 'dashboard' and len(self.notebook.tabs()) <= 1:
            return
        
        # 找到Tab索引
        tab_frame = self.pages[page_name]['frame']
        try:
            tab_index = None
            for i, tab_id in enumerate(self.notebook.tabs()):
                if self.notebook.nametowidget(tab_id) == tab_frame:
                    tab_index = i
                    break
            
            if tab_index is not None:
                self.close_tab(tab_index)
        except:
            pass

    def close_tab(self, tab_index):
        """关闭指定的Tab页"""
        if tab_index < 0 or tab_index >= len(self.notebook.tabs()):
            return
        
        # 获取要关闭的Tab页Frame
        tab_id = self.notebook.tabs()[tab_index]
        tab_frame = self.notebook.nametowidget(tab_id)
        
        # 找到对应的page_name
        page_name = None
        for key, value in self.pages.items():
            if value['frame'] == tab_frame:
                page_name = key
                break
        
        # 如果关闭的是仪表盘，不允许关闭（至少保留一个Tab页）
        if page_name == 'dashboard' and len(self.notebook.tabs()) <= 1:
            return
        
        # 删除Tab页
        self.notebook.forget(tab_index)
        
        # 从pages字典中删除
        if page_name:
            del self.pages[page_name]
        
        # 如果关闭后没有Tab页了，重新打开仪表盘
        if len(self.notebook.tabs()) == 0:
            self.show_page('dashboard')

    def close_other_tabs(self, current_index):
        """关闭除当前Tab页外的其他Tab页"""
        tabs_to_close = []
        for i in range(len(self.notebook.tabs())):
            if i != current_index:
                tabs_to_close.append(i)
        
        # 从后往前删除，避免索引变化
        for i in reversed(tabs_to_close):
            self.close_tab(i)

    def close_all_tabs(self):
        """关闭所有Tab页（保留仪表盘）"""
        tabs_to_close = []
        for i in range(len(self.notebook.tabs())):
            tab_id = self.notebook.tabs()[i]
            tab_frame = self.notebook.nametowidget(tab_id)
            # 检查是否是仪表盘
            is_dashboard = False
            for key, value in self.pages.items():
                if value['frame'] == tab_frame and key == 'dashboard':
                    is_dashboard = True
                    break
            if not is_dashboard:
                tabs_to_close.append(i)
        
        # 从后往前删除
        for i in reversed(tabs_to_close):
            self.close_tab(i)

    def show_dashboard(self, container=None):
        """显示仪表盘"""
        if container is None:
            container = self.page_container if hasattr(self, 'page_container') else self.notebook.nametowidget(self.notebook.select())
        
        # 清空容器（如果之前有内容）
        for widget in container.winfo_children():
            widget.destroy()
        
        # 标题
        title_frame = tk.Frame(container, bg='#F0F2F5')
        title_frame.pack(fill='x', pady=(0, 20))
        tk.Label(title_frame, text="仪表盘",
                font=('微软雅黑', 24, 'bold'),
                bg='#F0F2F5', fg='#2C3E50').pack(anchor='w')

        # 统计卡片容器（保存为实例变量，以便后续刷新）
        self.dashboard_cards_frame = tk.Frame(container, bg='#F0F2F5')
        self.dashboard_cards_frame.pack(fill='x')

        # 获取统计数据
        stats = self.get_dashboard_stats()

        # 创建统计卡片（保存引用以便刷新）
        self.dashboard_cards = []
        cards = [
            ("商品总数", stats['products'], '#3498DB', '📦'),
            ("今日入库", stats['today_inbound'], '#2ECC71', '📥'),
            ("今日出库", stats['today_outbound'], '#E74C3C', '📤'),
            ("库存预警", stats['low_stock'], '#F39C12', '⚠️')
        ]

        for i, (title, value, color, icon) in enumerate(cards):
            card = self.create_stat_card(self.dashboard_cards_frame, title, value, color, icon)
            card.grid(row=0, column=i, padx=10, sticky='ew')
            self.dashboard_cards.append(card)

        self.dashboard_cards_frame.grid_columnconfigure(0, weight=1)
        self.dashboard_cards_frame.grid_columnconfigure(1, weight=1)
        self.dashboard_cards_frame.grid_columnconfigure(2, weight=1)
        self.dashboard_cards_frame.grid_columnconfigure(3, weight=1)
        
        # 保存容器引用
        self.dashboard_container = container

        # 快捷操作按钮区域
        quick_actions_frame = tk.Frame(container, bg='white', relief='raised', bd=1)
        quick_actions_frame.pack(fill='both', expand=True, pady=20)

        tk.Label(quick_actions_frame, text="快捷操作",
                font=('微软雅黑', 14, 'bold'),
                bg='white').pack(anchor='w', padx=20, pady=(20, 10))

        # 创建快捷操作按钮网格
        self.create_quick_action_buttons(quick_actions_frame)

    def create_stat_card(self, parent, title, value, color, icon):
        """创建统计卡片"""
        card = tk.Frame(parent, bg='white', relief='raised', bd=1)

        # 顶部色条
        top_bar = tk.Frame(card, bg=color, height=5)
        top_bar.pack(fill='x')
        top_bar.pack_propagate(False)

        # 内容
        content = tk.Frame(card, bg='white')
        content.pack(fill='both', expand=True, pady=20)

        # 图标和数值（保存标签引用以便更新）
        icon_label = tk.Label(content, text=icon, font=('Arial', 30), bg='white')
        icon_label.pack()

        value_label = tk.Label(content, text=value, font=('微软雅黑', 28, 'bold'),
                              bg='white', fg=color)
        value_label.pack(pady=5)

        title_label = tk.Label(content, text=title, font=('微软雅黑', 12),
                              bg='white', fg='#7F8C8D')
        title_label.pack()

        # 将标签保存到card的属性中，以便后续更新
        card.value_label = value_label
        card.card_title = title
        card.card_color = color

        return card
    
    def refresh_dashboard(self):
        """刷新仪表盘数据"""
        # 如果仪表盘卡片存在，更新它们
        if hasattr(self, 'dashboard_cards') and self.dashboard_cards:
            stats = self.get_dashboard_stats()
            card_values = [
                stats['products'],
                stats['today_inbound'],
                stats['today_outbound'],
                stats['low_stock']
            ]
            
            for i, card in enumerate(self.dashboard_cards):
                if hasattr(card, 'value_label') and i < len(card_values):
                    card.value_label.config(text=card_values[i])
        
    def create_quick_action_buttons(self, parent):
        """创建快捷操作按钮网格"""
        # 按钮容器
        buttons_container = tk.Frame(parent, bg='white')
        buttons_container.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        # 获取当前用户权限
        current_user = self.user[1]
        user_permissions = self.get_user_permissions(current_user)
        is_admin = self.user[2] == 'admin'

        # 快捷操作按钮配置
        # (图标, 文本, 命令类型, 命令参数, 权限key, 颜色)
        quick_actions = [
            ('📥', '采购入库', 'add_inbound', None, 'inbound', '#2ECC71'),
            ('📤', '销货出库', 'add_outbound', None, 'outbound', '#E74C3C'),
            ('📦', '库存查询', 'show_page', 'inventory', 'inventory', '#3498DB'),
            ('📝', '商品管理', 'show_page', 'basic_management', 'products', '#9B59B6'),
            ('🏢', '供应商管理', 'show_page', 'basic_management', 'suppliers', '#16A085'),
            ('👥', '客户管理', 'show_page', 'basic_management', 'customers', '#F39C12'),
            ('🏭', '仓库管理', 'show_page', 'basic_management', 'warehouses', '#34495E'),
            ('📈', '统计报表', 'show_page', 'reports', 'reports', '#E67E22'),
        ]

        # 创建4x2网格布局
        grid_frame = tk.Frame(buttons_container, bg='white')
        grid_frame.pack(fill='both', expand=True, pady=20)

        row = 0
        col = 0
        max_cols = 4

        for idx, (icon, text, cmd_type, cmd_param, permission_key, color) in enumerate(quick_actions):
            # 检查权限（管理员显示所有，普通用户根据权限显示）
            show_button = False
            if is_admin:
                show_button = True
            else:
                if permission_key in user_permissions:
                    show_button = True
                elif permission_key in ['products', 'suppliers', 'customers', 'warehouses']:
                    # 基础管理的子权限：如果有任一基础管理权限就显示
                    has_basic_permission = any(p in user_permissions for p in ['products', 'suppliers', 'customers', 'warehouses'])
                    if has_basic_permission:
                        show_button = True
            
            if not show_button:
                continue

            # 创建命令函数（使用functools.partial避免闭包问题）
            if cmd_type == 'add_inbound':
                action_cmd = partial(self.add_inbound_order)
            elif cmd_type == 'add_outbound':
                action_cmd = partial(self.add_outbound_order)
            elif cmd_type == 'show_page':
                action_cmd = partial(self.show_page, cmd_param)
            else:
                action_cmd = lambda: None

            # 创建按钮（使用Button组件，更简单直接）
            btn = tk.Button(grid_frame, 
                           text=f"{icon}\n{text}",
                           font=('微软雅黑', 11, 'bold'),
                           bg='white',
                           fg=color,
                           activebackground='#F0F7FF',
                           activeforeground=color,
                           relief='solid',
                           bd=1,
                           cursor='hand2',
                           command=action_cmd,
                           compound='top',  # 图标在上方
                           width=12,
                           height=6,
                           padx=10,
                           pady=15)
            btn.grid(row=row, column=col, padx=15, pady=15, sticky='nsew')

            # 更新位置
            col += 1
            if col >= max_cols:
                col = 0
                row += 1

        # 配置网格权重（确保按钮均匀分布）
        for i in range(max_cols):
            grid_frame.grid_columnconfigure(i, weight=1)
        if row >= 0:
            for i in range(row + 1):
                grid_frame.grid_rowconfigure(i, weight=1)

    def get_dashboard_stats(self):
        """获取仪表盘统计数据"""
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        # 商品总数
        cursor.execute("SELECT COUNT(*) FROM products")
        products = cursor.fetchone()[0]

        # 今日入库
        today = datetime.now().strftime('%Y-%m-%d')
        cursor.execute("SELECT COUNT(*) FROM inbound_orders WHERE order_date=?", (today,))
        today_inbound = cursor.fetchone()[0]

        # 今日出库
        cursor.execute("SELECT COUNT(*) FROM outbound_orders WHERE order_date=?", (today,))
        today_outbound = cursor.fetchone()[0]

        # 库存预警（低于最小库存）
        cursor.execute('''
            SELECT COUNT(*) FROM inventory i
            JOIN products p ON i.product_id = p.id
            WHERE i.quantity < p.min_stock
        ''')
        low_stock = cursor.fetchone()[0]

        conn.close()

        return {
            'products': products,
            'today_inbound': today_inbound,
            'today_outbound': today_outbound,
            'low_stock': low_stock
        }

    def load_recent_transactions(self, tree):
        """加载最近交易记录"""
        # 清空现有数据
        for item in tree.get_children():
            tree.delete(item)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        # 获取最近入库明细记录（JOIN商品表获取商品名称）
        cursor.execute('''
            SELECT 
                io.created_time,
                '入库' as type,
                io.order_no,
                p.name as product_name,
                id.quantity,
                io.operator
            FROM inbound_details id
            INNER JOIN inbound_orders io ON id.order_id = io.id
            INNER JOIN products p ON id.product_id = p.id
            ORDER BY io.created_time DESC
            LIMIT 20
        ''')
        inbound_records = cursor.fetchall()

        # 获取最近出库明细记录（JOIN商品表获取商品名称）
        cursor.execute('''
            SELECT 
                oo.created_time,
                '出库' as type,
                oo.order_no,
                p.name as product_name,
                od.quantity,
                oo.operator
            FROM outbound_details od
            INNER JOIN outbound_orders oo ON od.order_id = oo.id
            INNER JOIN products p ON od.product_id = p.id
            ORDER BY oo.created_time DESC
            LIMIT 20
        ''')
        outbound_records = cursor.fetchall()

        # 合并并排序（按创建时间倒序）
        all_records = list(inbound_records) + list(outbound_records)
        all_records.sort(key=lambda x: x[0] if x[0] else '', reverse=True)

        # 格式化时间并插入表格
        for record in all_records[:15]:  # 显示最近15条记录
            created_time = record[0]
            # 格式化时间：如果是datetime对象，转换为字符串；如果是字符串，直接使用
            if created_time:
                if isinstance(created_time, str):
                    time_str = created_time[:16] if len(created_time) > 16 else created_time  # 取前16个字符（YYYY-MM-DD HH:MM）
                else:
                    time_str = str(created_time)[:16]
            else:
                time_str = ''
            
            tree.insert('', 'end', values=(
                time_str,
                record[1],  # 类型（入库/出库）
                record[2],  # 单号
                record[3] if record[3] else '',  # 商品名称
                str(record[4]) if record[4] else '0',  # 数量
                record[5] if record[5] else ''  # 操作人
            ))

        conn.close()

    def show_inbound(self, container=None):
        """显示入库管理页面"""
        if container is None:
            container = self.page_container if hasattr(self, 'page_container') else self.notebook.nametowidget(self.notebook.select())
        
        # 标题栏
        title_bar = tk.Frame(container, bg='white', height=60, relief='raised', bd=1)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)

        tk.Label(title_bar, text="入库管理",
                font=('微软雅黑', 18, 'bold'),
                bg='white').pack(side='left', padx=20, pady=15)

        # 新增入库按钮
        add_btn = tk.Button(title_bar, text="+ 新增入库单",
                           font=('微软雅黑', 11),
                           bg='#2ECC71', fg='white',
                           pady=5, padx=20,
                           cursor='hand2',
                           command=self.add_inbound_order)
        add_btn.pack(side='right', padx=20, pady=10)

        # 搜索筛选区
        filter_frame = tk.Frame(container, bg='white', relief='raised', bd=1)
        filter_frame.pack(fill='x', pady=10)

        # 日期范围
        tk.Label(filter_frame, text="日期范围：",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=0, padx=20, pady=10)

        # 开始日期
        self.start_date = tk.StringVar(value=(datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        start_date_frame = tk.Frame(filter_frame, bg='white')
        start_date_frame.grid(row=0, column=1, padx=5)
        start_entry = tk.Entry(start_date_frame, textvariable=self.start_date,
                              font=('微软雅黑', 10), width=12)
        start_entry.pack(side='left')
        start_entry.bind('<Button-1>', lambda e: self.show_date_picker(start_entry))
        # 日期选择按钮
        start_date_btn = tk.Button(start_date_frame, text="📅", font=('微软雅黑', 9),
                                   width=2, height=1, bg='#ECF0F1', relief='flat',
                                   cursor='hand2', command=lambda: self.show_date_picker(start_entry))
        start_date_btn.pack(side='left', padx=(2, 0))

        tk.Label(filter_frame, text="至",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=2)

        # 结束日期
        self.end_date = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        end_date_frame = tk.Frame(filter_frame, bg='white')
        end_date_frame.grid(row=0, column=3, padx=5)
        end_entry = tk.Entry(end_date_frame, textvariable=self.end_date,
                            font=('微软雅黑', 10), width=12)
        end_entry.pack(side='left')
        end_entry.bind('<Button-1>', lambda e: self.show_date_picker(end_entry))
        # 日期选择按钮
        end_date_btn = tk.Button(end_date_frame, text="📅", font=('微软雅黑', 9),
                                 width=2, height=1, bg='#ECF0F1', relief='flat',
                                 cursor='hand2', command=lambda: self.show_date_picker(end_entry))
        end_date_btn.pack(side='left', padx=(2, 0))

        # 单号搜索
        tk.Label(filter_frame, text="单号：",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=4, padx=(20, 5))

        self.order_no = tk.StringVar()
        order_entry = tk.Entry(filter_frame, textvariable=self.order_no,
                              font=('微软雅黑', 10), width=15)
        order_entry.grid(row=0, column=5, padx=5)

        # 搜索按钮
        search_btn = tk.Button(filter_frame, text="搜索",
                              font=('微软雅黑', 10),
                              bg='#3498DB', fg='white',
                              pady=5, padx=20,
                              cursor='hand2',
                              command=self.search_inbound)
        search_btn.grid(row=0, column=6, padx=20)

        # 入库单列表
        list_frame = tk.Frame(container, bg='white', relief='raised', bd=1)
        list_frame.pack(fill='both', expand=True, pady=10)

        # 创建表格
        self.create_inbound_table(list_frame)

    def create_inbound_table(self, parent):
        """创建入库单表格"""
        # 表格工具栏
        toolbar = tk.Frame(parent, bg='white')
        toolbar.pack(fill='x', padx=20, pady=10)

        tk.Label(toolbar, text="入库单列表",
                font=('微软雅黑', 14, 'bold'),
                bg='white').pack(side='left')

        # 删除按钮
        delete_btn = tk.Button(toolbar, text="× 删除选中",
                             font=('微软雅黑', 10),
                             bg='#E74C3C', fg='white',
                             pady=5, padx=15,
                             cursor='hand2',
                             command=self.delete_inbound_order)
        delete_btn.pack(side='right', padx=5)

        # 表格框架
        table_frame = tk.Frame(parent, bg='white')
        table_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        # 创建Treeview（添加商品和规格列）
        columns = ('单号', '日期', '供应商', '仓库', '商品编码', '商品名称', '规格型号', '数量', '单价', '金额', '批次号', '状态', '操作人')
        self.inbound_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

        # 设置列标题和宽度
        column_widths = {
            '单号': 150,
            '日期': 100,
            '供应商': 120,
            '仓库': 100,
            '商品编码': 100,
            '商品名称': 150,
            '规格型号': 120,
            '数量': 80,
            '单价': 80,
            '金额': 100,
            '批次号': 100,
            '状态': 80,
            '操作人': 80
        }
        for col in columns:
            self.inbound_tree.heading(col, text=col)
            self.inbound_tree.column(col, width=column_widths.get(col, 100))

        # 添加滚动条
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.inbound_tree.yview)
        self.inbound_tree.configure(yscrollcommand=scrollbar.set)

        self.inbound_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # 加载数据
        self.load_inbound_data()

        # 绑定双击事件
        self.inbound_tree.bind('<Double-1>', self.on_inbound_double_click)

    def load_inbound_data(self):
        """加载入库单数据（展开明细，每个商品一行）"""
        # 清空现有数据
        for item in self.inbound_tree.get_children():
            self.inbound_tree.delete(item)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        # 查询入库单明细，每个商品一行
        cursor.execute('''
            SELECT io.order_no, io.order_date, io.supplier, w.name,
                   p.code, p.name, p.specification,
                   id.quantity, id.price, id.amount, id.batch_no,
                   io.status, io.operator
            FROM inbound_orders io
            LEFT JOIN warehouses w ON io.warehouse_id = w.id
            LEFT JOIN inbound_details id ON io.id = id.order_id
            LEFT JOIN products p ON id.product_id = p.id
            ORDER BY io.created_time DESC, io.id, id.id
        ''')

        records = cursor.fetchall()

        for record in records:
            status_text = '已入库' if record[11] == 1 else '待入库'
            self.inbound_tree.insert('', 'end', values=(
                record[0],  # 单号
                record[1],  # 日期
                record[2] or '',  # 供应商
                record[3] or '',  # 仓库
                record[4] or '',  # 商品编码
                record[5] or '',  # 商品名称
                record[6] or '',  # 规格型号
                record[7] or 0,  # 数量
                f"{float(record[8] or 0):.2f}",  # 单价
                f"¥{float(record[9] or 0):.2f}",  # 金额
                record[10] or '',  # 批次号
                status_text,  # 状态
                record[12] or ''  # 操作人
            ))

        conn.close()

    def delete_inbound_order(self):
        """删除入库单"""
        selection = self.inbound_tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先选择要删除的入库单记录")
            return

        item = self.inbound_tree.item(selection[0])
        order_no = item['values'][0]
        product_name = item['values'][5]  # 商品名称

        if not messagebox.askyesno("确认删除", f"确定要删除入库单 {order_no} 的商品 {product_name} 的记录吗？\n\n注意：删除后库存会相应调整！"):
            return

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        try:
            # 获取订单ID和明细ID
            cursor.execute("SELECT id FROM inbound_orders WHERE order_no=?", (order_no,))
            order_result = cursor.fetchone()
            if not order_result:
                messagebox.showerror("错误", "未找到对应的入库单")
                return

            order_id = order_result[0]

            # 从订单主表获取仓库ID
            cursor.execute("SELECT warehouse_id FROM inbound_orders WHERE id=?", (order_id,))
            warehouse_result = cursor.fetchone()
            if not warehouse_result or not warehouse_result[0]:
                messagebox.showerror("错误", "未找到对应的仓库信息")
                return
            warehouse_id = warehouse_result[0]

            # 获取商品ID
            product_code = item['values'][4]  # 商品编码
            cursor.execute("SELECT id FROM products WHERE code=?", (product_code,))
            product_result = cursor.fetchone()
            if not product_result:
                messagebox.showerror("错误", "未找到对应的商品")
                return

            product_id = product_result[0]

            # 获取明细信息（数量、批次号等）
            cursor.execute('''
                SELECT id, quantity, batch_no
                FROM inbound_details
                WHERE order_id=? AND product_id=?
            ''', (order_id, product_id))
            detail_result = cursor.fetchone()
            if not detail_result:
                messagebox.showerror("错误", "未找到对应的入库明细")
                return

            detail_id, quantity, batch_no = detail_result

            # 删除明细
            cursor.execute("DELETE FROM inbound_details WHERE id=?", (detail_id,))

            # 恢复库存（扣减）
            cursor.execute('''
                UPDATE inventory
                SET quantity = quantity - ?,
                    available_quantity = available_quantity - ?,
                    updated_time = CURRENT_TIMESTAMP
                WHERE product_id=? AND warehouse_id=? AND (batch_no=? OR (batch_no IS NULL AND ? IS NULL))
            ''', (quantity, quantity, product_id, warehouse_id, batch_no, batch_no))

            # 检查订单是否还有其他明细
            cursor.execute("SELECT COUNT(*) FROM inbound_details WHERE order_id=?", (order_id,))
            remaining_count = cursor.fetchone()[0]

            if remaining_count == 0:
                # 如果没有其他明细，删除订单
                cursor.execute("DELETE FROM inbound_orders WHERE id=?", (order_id,))
            else:
                # 重新计算订单总金额
                cursor.execute('''
                    SELECT SUM(amount) FROM inbound_details WHERE order_id=?
                ''', (order_id,))
                total_result = cursor.fetchone()
                new_total = total_result[0] if total_result[0] else 0
                cursor.execute('''
                    UPDATE inbound_orders SET total_amount=? WHERE id=?
                ''', (new_total, order_id))

            conn.commit()
            messagebox.showinfo("成功", "删除成功！")
            self.load_inbound_data()  # 刷新列表

        except Exception as e:
            conn.rollback()
            messagebox.showerror("错误", f"删除失败：{str(e)}")
        finally:
            conn.close()

    def on_inbound_double_click(self, event):
        """入库单双击事件"""
        selection = self.inbound_tree.selection()
        if selection:
            item = self.inbound_tree.item(selection[0])
            order_no = item['values'][0]
            # 可以在这里实现查看入库单详情的功能
            # messagebox.showinfo("提示", f"查看入库单：{order_no}")

    def show_outbound(self, container=None):
        """显示出库管理页面"""
        if container is None:
            container = self.page_container if hasattr(self, 'page_container') else self.notebook.nametowidget(self.notebook.select())
        
        # 标题栏
        title_bar = tk.Frame(container, bg='white', height=60, relief='raised', bd=1)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)

        tk.Label(title_bar, text="出库管理",
                font=('微软雅黑', 18, 'bold'),
                bg='white').pack(side='left', padx=20, pady=15)

        # 新增出库按钮
        add_btn = tk.Button(title_bar, text="+ 新增出库单",
                           font=('微软雅黑', 11),
                           bg='#E74C3C', fg='white',
                           pady=5, padx=20,
                           cursor='hand2',
                           command=self.add_outbound_order)
        add_btn.pack(side='right', padx=20, pady=10)

        # 搜索筛选区
        filter_frame = tk.Frame(container, bg='white', relief='raised', bd=1)
        filter_frame.pack(fill='x', pady=10)

        # 日期范围
        tk.Label(filter_frame, text="日期范围：",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=0, padx=20, pady=10)

        self.out_start_date = tk.StringVar(value=(datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        out_start_date_frame = tk.Frame(filter_frame, bg='white')
        out_start_date_frame.grid(row=0, column=1, padx=5)
        out_start_entry = tk.Entry(out_start_date_frame, textvariable=self.out_start_date,
                font=('微软雅黑', 10), width=12)
        out_start_entry.pack(side='left')
        out_start_entry.bind('<Button-1>', lambda e: self.show_date_picker(out_start_entry))
        # 日期选择按钮
        out_start_date_btn = tk.Button(out_start_date_frame, text="📅", font=('微软雅黑', 9),
                                       width=2, height=1, bg='#ECF0F1', relief='flat',
                                       cursor='hand2', command=lambda: self.show_date_picker(out_start_entry))
        out_start_date_btn.pack(side='left', padx=(2, 0))

        tk.Label(filter_frame, text="至",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=2)

        self.out_end_date = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        out_end_date_frame = tk.Frame(filter_frame, bg='white')
        out_end_date_frame.grid(row=0, column=3, padx=5)
        out_end_entry = tk.Entry(out_end_date_frame, textvariable=self.out_end_date,
                font=('微软雅黑', 10), width=12)
        out_end_entry.pack(side='left')
        out_end_entry.bind('<Button-1>', lambda e: self.show_date_picker(out_end_entry))
        # 日期选择按钮
        out_end_date_btn = tk.Button(out_end_date_frame, text="📅", font=('微软雅黑', 9),
                                     width=2, height=1, bg='#ECF0F1', relief='flat',
                                     cursor='hand2', command=lambda: self.show_date_picker(out_end_entry))
        out_end_date_btn.pack(side='left', padx=(2, 0))

        # 客户搜索
        tk.Label(filter_frame, text="客户：",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=4, padx=(20, 5))

        self.customer = tk.StringVar()
        tk.Entry(filter_frame, textvariable=self.customer,
                font=('微软雅黑', 10), width=15).grid(row=0, column=5, padx=5)

        # 搜索按钮
        search_btn = tk.Button(filter_frame, text="搜索",
                              font=('微软雅黑', 10),
                              bg='#3498DB', fg='white',
                              pady=5, padx=20,
                              cursor='hand2',
                              command=self.search_outbound)
        search_btn.grid(row=0, column=6, padx=20)

        # 出库单列表
        list_frame = tk.Frame(container, bg='white', relief='raised', bd=1)
        list_frame.pack(fill='both', expand=True, pady=10)

        # 创建出库单表格
        self.create_outbound_table(list_frame)

    def create_outbound_table(self, parent):
        """创建出库单表格"""
        # 表格工具栏
        toolbar = tk.Frame(parent, bg='white')
        toolbar.pack(fill='x', padx=20, pady=10)

        tk.Label(toolbar, text="出库单列表",
                font=('微软雅黑', 14, 'bold'),
                bg='white').pack(side='left')

        # 删除按钮
        delete_btn = tk.Button(toolbar, text="× 删除选中",
                             font=('微软雅黑', 10),
                             bg='#E74C3C', fg='white',
                             pady=5, padx=15,
                             cursor='hand2',
                             command=self.delete_outbound_order)
        delete_btn.pack(side='right', padx=5)

        # 表格框架
        table_frame = tk.Frame(parent, bg='white')
        table_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        # 创建Treeview（添加商品和规格列）
        columns = ('单号', '日期', '客户', '仓库', '商品编码', '商品名称', '规格型号', '出库数量', '单价', '金额', '状态', '操作人')
        self.outbound_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

        # 设置列标题和宽度
        column_widths = {
            '单号': 150,
            '日期': 100,
            '客户': 120,
            '仓库': 100,
            '商品编码': 100,
            '商品名称': 150,
            '规格型号': 120,
            '出库数量': 80,
            '单价': 80,
            '金额': 100,
            '状态': 80,
            '操作人': 80
        }
        for col in columns:
            self.outbound_tree.heading(col, text=col)
            self.outbound_tree.column(col, width=column_widths.get(col, 100))

        # 添加滚动条
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.outbound_tree.yview)
        self.outbound_tree.configure(yscrollcommand=scrollbar.set)

        self.outbound_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # 加载数据
        self.load_outbound_data()

        # 绑定双击事件
        self.outbound_tree.bind('<Double-1>', self.on_outbound_double_click)

    def load_outbound_data(self):
        """加载出库单数据（展开明细，每个商品一行）"""
        # 清空现有数据
        for item in self.outbound_tree.get_children():
            self.outbound_tree.delete(item)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        # 查询出库单明细，每个商品一行
        cursor.execute('''
            SELECT oo.order_no, oo.order_date, oo.customer, w.name,
                   p.code, p.name, p.specification,
                   od.quantity, od.price, od.amount,
                   oo.status, oo.operator
            FROM outbound_orders oo
            LEFT JOIN warehouses w ON oo.warehouse_id = w.id
            LEFT JOIN outbound_details od ON oo.id = od.order_id
            LEFT JOIN products p ON od.product_id = p.id
            ORDER BY oo.created_time DESC, oo.id, od.id
        ''')

        records = cursor.fetchall()

        for record in records:
            status_text = '已出库' if record[10] == 1 else '待出库'
            self.outbound_tree.insert('', 'end', values=(
                record[0],  # 单号
                record[1],  # 日期
                record[2] or '',  # 客户
                record[3] or '',  # 仓库
                record[4] or '',  # 商品编码
                record[5] or '',  # 商品名称
                record[6] or '',  # 规格型号
                record[7] or 0,  # 出库数量
                f"{float(record[8] or 0):.2f}",  # 单价
                f"¥{float(record[9] or 0):.2f}",  # 金额
                status_text,  # 状态
                record[11] or ''  # 操作人
            ))

        conn.close()

    def delete_outbound_order(self):
        """删除出库单"""
        selection = self.outbound_tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先选择要删除的出库单记录")
            return

        item = self.outbound_tree.item(selection[0])
        order_no = item['values'][0]
        product_name = item['values'][5]  # 商品名称

        if not messagebox.askyesno("确认删除", f"确定要删除出库单 {order_no} 的商品 {product_name} 的记录吗？\n\n注意：删除后库存会相应调整！"):
            return

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        try:
            # 获取订单ID和明细ID
            cursor.execute("SELECT id FROM outbound_orders WHERE order_no=?", (order_no,))
            order_result = cursor.fetchone()
            if not order_result:
                messagebox.showerror("错误", "未找到对应的出库单")
                return

            order_id = order_result[0]

            # 从订单主表获取仓库ID
            cursor.execute("SELECT warehouse_id FROM outbound_orders WHERE id=?", (order_id,))
            warehouse_result = cursor.fetchone()
            if not warehouse_result or not warehouse_result[0]:
                messagebox.showerror("错误", "未找到对应的仓库信息")
                return
            warehouse_id = warehouse_result[0]

            # 获取商品ID
            product_code = item['values'][4]  # 商品编码
            cursor.execute("SELECT id FROM products WHERE code=?", (product_code,))
            product_result = cursor.fetchone()
            if not product_result:
                messagebox.showerror("错误", "未找到对应的商品")
                return

            product_id = product_result[0]

            # 获取明细信息（数量、批次号等）
            cursor.execute('''
                SELECT id, quantity, batch_no
                FROM outbound_details
                WHERE order_id=? AND product_id=?
            ''', (order_id, product_id))
            detail_result = cursor.fetchone()
            if not detail_result:
                messagebox.showerror("错误", "未找到对应的出库明细")
                return

            detail_id, quantity, batch_no = detail_result

            # 删除明细
            cursor.execute("DELETE FROM outbound_details WHERE id=?", (detail_id,))

            # 恢复库存（增加）
            cursor.execute('''
                UPDATE inventory
                SET quantity = quantity + ?,
                    available_quantity = available_quantity + ?,
                    updated_time = CURRENT_TIMESTAMP
                WHERE product_id=? AND warehouse_id=? AND (batch_no=? OR (batch_no IS NULL AND ? IS NULL))
            ''', (quantity, quantity, product_id, warehouse_id, batch_no, batch_no))

            # 检查订单是否还有其他明细
            cursor.execute("SELECT COUNT(*) FROM outbound_details WHERE order_id=?", (order_id,))
            remaining_count = cursor.fetchone()[0]

            if remaining_count == 0:
                # 如果没有其他明细，删除订单
                cursor.execute("DELETE FROM outbound_orders WHERE id=?", (order_id,))
            else:
                # 重新计算订单总金额
                cursor.execute('''
                    SELECT SUM(amount) FROM outbound_details WHERE order_id=?
                ''', (order_id,))
                total_result = cursor.fetchone()
                new_total = total_result[0] if total_result[0] else 0
                cursor.execute('''
                    UPDATE outbound_orders SET total_amount=? WHERE id=?
                ''', (new_total, order_id))

            conn.commit()
            messagebox.showinfo("成功", "删除成功！")
            self.load_outbound_data()  # 刷新列表

        except Exception as e:
            conn.rollback()
            messagebox.showerror("错误", f"删除失败：{str(e)}")
        finally:
            conn.close()

    def on_outbound_double_click(self, event):
        """出库单双击事件"""
        selection = self.outbound_tree.selection()
        if selection:
            item = self.outbound_tree.item(selection[0])
            order_no = item['values'][0]
            messagebox.showinfo("提示", f"查看出库单：{order_no}")

    def show_inventory(self, container=None):
        """显示库存查询页面"""
        if container is None:
            container = self.page_container if hasattr(self, 'page_container') else self.notebook.nametowidget(self.notebook.select())
        
        # 确保容器有背景色和正确配置
        container.config(bg='#F0F2F5')
        
        # 清空容器（如果之前有内容）
        for widget in container.winfo_children():
            widget.destroy()
        
        # 主容器（用于填充）
        main_container = tk.Frame(container, bg='#F0F2F5')
        main_container.pack(fill='both', expand=True)
        
        # 标题栏
        title_bar = tk.Frame(main_container, bg='white', height=60, relief='raised', bd=1)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)

        tk.Label(title_bar, text="库存查询",
                font=('微软雅黑', 18, 'bold'),
                bg='white').pack(side='left', padx=20, pady=15)

        # 按钮组
        btn_frame = tk.Frame(title_bar, bg='white')
        btn_frame.pack(side='right', padx=20, pady=10)
        
        # 初始化按钮（危险操作，放在最左边）
        init_btn = tk.Button(btn_frame, text="⚠ 初始化数据",
                              font=('微软雅黑', 11),
                              bg='#E74C3C', fg='white',
                              pady=5, padx=20,
                              cursor='hand2',
                              command=self.init_database_data)
        init_btn.pack(side='right', padx=5)
        
        # 删除按钮
        delete_btn = tk.Button(btn_frame, text="× 删除选中",
                              font=('微软雅黑', 11),
                              bg='#E74C3C', fg='white',
                              pady=5, padx=20,
                              cursor='hand2',
                              command=self.delete_inventory)
        delete_btn.pack(side='right', padx=5)
        
        # 导出按钮
        export_btn = tk.Button(btn_frame, text="导出库存",
                              font=('微软雅黑', 11),
                              bg='#27AE60', fg='white',
                              pady=5, padx=20,
                              cursor='hand2',
                              command=self.export_inventory)
        export_btn.pack(side='right', padx=5)

        # 筛选区
        filter_frame = tk.Frame(main_container, bg='white', relief='raised', bd=1)
        filter_frame.pack(fill='x', pady=10, padx=10)

        # 商品搜索
        tk.Label(filter_frame, text="商品：",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=0, padx=20, pady=10)

        self.product_search = tk.StringVar()
        tk.Entry(filter_frame, textvariable=self.product_search,
                font=('微软雅黑', 10), width=15).grid(row=0, column=1, padx=5)

        # 规格
        tk.Label(filter_frame, text="规格型号：",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=2, padx=(20, 5))

        self.spec_search = tk.StringVar()
        tk.Entry(filter_frame, textvariable=self.spec_search,
                font=('微软雅黑', 10), width=15).grid(row=0, column=3, padx=5)

        # 仓库筛选
        tk.Label(filter_frame, text="仓库：",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=4, padx=(20, 5))

        self.warehouse_var = tk.StringVar()
        warehouse_combo = ttk.Combobox(filter_frame, textvariable=self.warehouse_var,
                                      font=('微软雅黑', 10), width=12, state='readonly')
        warehouse_combo['values'] = self.get_warehouses()
        warehouse_combo.grid(row=0, column=5, padx=5)
        warehouse_combo.set('全部仓库')

        # 库存状态
        tk.Label(filter_frame, text="库存状态：",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=6, padx=(20, 5))

        self.stock_status = tk.StringVar()
        status_combo = ttk.Combobox(filter_frame, textvariable=self.stock_status,
                                   font=('微软雅黑', 10), width=12, state='readonly')
        status_combo['values'] = ('全部', '正常', '库存不足', '积压')
        status_combo.grid(row=0, column=7, padx=5)
        status_combo.set('全部')

        # 搜索按钮
        search_btn = tk.Button(filter_frame, text="查询",
                              font=('微软雅黑', 10),
                              bg='#3498DB', fg='white',
                              pady=5, padx=20,
                              cursor='hand2',
                              command=self.search_inventory)
        search_btn.grid(row=0, column=8, padx=20)

        # 库存列表
        list_frame = tk.Frame(main_container, bg='white', relief='raised', bd=1)
        list_frame.pack(fill='both', expand=True, pady=(0, 10), padx=10)

        # 创建库存表格
        try:
            self.create_inventory_table(list_frame)
        except Exception as e:
            print(f"创建库存表格时出错: {e}")
            import traceback
            traceback.print_exc()
            # 即使出错也显示一个提示标签
            tk.Label(list_frame, text=f"加载库存数据时出错: {str(e)}", 
                    font=('微软雅黑', 12), bg='white', fg='red').pack(pady=50)

    def get_warehouses(self):
        """获取仓库列表"""
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM warehouses WHERE status=1")
        warehouses = ['全部仓库'] + [w[0] for w in cursor.fetchall()]
        conn.close()
        return warehouses

    def create_inventory_table(self, parent):
        """创建库存表格"""
        # 表格工具栏
        toolbar = tk.Frame(parent, bg='white')
        toolbar.pack(fill='x', padx=20, pady=10)

        tk.Label(toolbar, text="库存列表",
                font=('微软雅黑', 14, 'bold'),
                bg='white').pack(side='left')

        # 库存统计
        self.inventory_stats = tk.Label(toolbar, text="",
                                      font=('微软雅黑', 10),
                                      bg='white', fg='#7F8C8D')
        self.inventory_stats.pack(side='right')

        # 表格框架
        table_frame = tk.Frame(parent, bg='white')
        table_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        # 创建Treeview
        columns = ('商品编码', '商品名称', '规格型号', '仓库', '库存数量', '可用库存', '单位', '最小库存', '状态')
        self.inventory_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

        # 设置列标题和宽度
        column_widths = [100, 150, 120, 100, 100, 100, 60, 80, 80]
        for i, col in enumerate(columns):
            self.inventory_tree.heading(col, text=col)
            self.inventory_tree.column(col, width=column_widths[i])

        # 添加滚动条
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.inventory_tree.yview)
        self.inventory_tree.configure(yscrollcommand=scrollbar.set)

        self.inventory_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # 加载数据（延迟加载，确保表格已创建）
        try:
            self.load_inventory_data()
            if hasattr(self, 'inventory_stats'):
                self.update_inventory_stats()
        except Exception as e:
            print(f"初始化库存表格时出错: {e}")
            import traceback
            traceback.print_exc()

    def load_inventory_data(self):
        """加载库存数据"""
        try:
            # 检查inventory_tree是否存在
            if not hasattr(self, 'inventory_tree') or self.inventory_tree is None:
                return
            
            # 清空现有数据
            for item in self.inventory_tree.get_children():
                self.inventory_tree.delete(item)

            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()

            cursor.execute('''
                SELECT i.id, p.code, p.name, p.specification, w.name,
                       i.quantity, i.available_quantity, p.unit, p.min_stock
                FROM inventory i
                JOIN products p ON i.product_id = p.id
                JOIN warehouses w ON i.warehouse_id = w.id
                WHERE p.status = 1
                ORDER BY i.updated_time DESC
            ''')

            records = cursor.fetchall()

            for record in records:
                try:
                    # 判断库存状态（quantity在索引5，min_stock在索引8）
                    quantity = record[5] or 0
                    min_stock = record[8] or 0
                    
                    if quantity == 0:
                        status = '无库存'
                        status_color = '#95A5A6'
                    elif min_stock > 0 and quantity < min_stock:
                        status = '库存不足'
                        status_color = '#E74C3C'
                    elif min_stock > 0 and quantity > min_stock * 2:
                        status = '积压'
                        status_color = '#F39C12'
                    else:
                        status = '正常'
                        status_color = '#27AE60'

                    # 保存inventory_id到item的tags中，用于删除时识别
                    # record[0]是i.id，record[1-4]是code,name,specification,warehouse
                    # record[5-6]是quantity,available_quantity，record[7]是unit，record[8]是min_stock
                    self.inventory_tree.insert('', 'end', values=(
                        record[1] or '', record[2] or '', record[3] or '', record[4] or '',
                        record[5] or 0, record[6] or 0, record[7] or '', record[8] or 0, status
                    ), tags=(status_color, f"inv_id_{record[0]}"))
                except Exception as e:
                    print(f"加载库存数据行时出错: {e}")
                    continue

            # 设置行颜色
            self.inventory_tree.tag_configure('#E74C3C', background='#FADBD8')
            self.inventory_tree.tag_configure('#F39C12', background='#FCF3CF')
            self.inventory_tree.tag_configure('#27AE60', background='#D5F4E6')
            self.inventory_tree.tag_configure('#95A5A6', background='#ECF0F1')

            conn.close()
        except Exception as e:
            print(f"加载库存数据时出错: {e}")
            import traceback
            traceback.print_exc()
            try:
                conn.close()
            except:
                pass

    def update_inventory_stats(self):
        """更新库存统计"""
        if not hasattr(self, 'inventory_stats'):
            return
        total_items = len(self.inventory_tree.get_children())
        low_stock = sum(1 for item in self.inventory_tree.get_children()
                       if len(self.inventory_tree.item(item)['values']) > 8 and
                       self.inventory_tree.item(item)['values'][8] == '库存不足')

        self.inventory_stats.config(text=f"共 {total_items} 种商品，其中 {low_stock} 种库存不足")
    
    def delete_inventory(self):
        """删除库存记录"""
        if not hasattr(self, 'inventory_tree'):
            messagebox.showwarning("提示", "请先打开库存查询页面！")
            return
        
        selection = self.inventory_tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先选择要删除的库存记录！")
            return
        
        if not messagebox.askyesno("确认", "确定要删除选中的库存记录吗？\n\n此操作会清空该商品的库存！"):
            return
        
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            
            for item_id in selection:
                item = self.inventory_tree.item(item_id)
                tags = item['tags']
                
                # 从tags中提取inventory_id
                inventory_id = None
                for tag in tags:
                    if tag.startswith('inv_id_'):
                        inventory_id = int(tag.replace('inv_id_', ''))
                        break
                
                if inventory_id:
                    # 删除库存记录
                    cursor.execute("DELETE FROM inventory WHERE id=?", (inventory_id,))
            
            conn.commit()
            conn.close()
            
            messagebox.showinfo("成功", "库存记录已删除！")
            # 刷新列表
            self.load_inventory_data()
            self.update_inventory_stats()
            
        except Exception as e:
            messagebox.showerror("错误", f"删除库存记录失败：{str(e)}")

    def show_products(self, container=None):
        """显示商品管理页面"""
        if container is None:
            container = self.page_container if hasattr(self, 'page_container') else self.notebook.nametowidget(self.notebook.select())
        
        # 标题栏
        title_bar = tk.Frame(container, bg='white', height=60, relief='raised', bd=1)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)

        tk.Label(title_bar, text="商品管理",
                font=('微软雅黑', 18, 'bold'),
                bg='white').pack(side='left', padx=20, pady=15)

        # 按钮组
        btn_frame = tk.Frame(title_bar, bg='white')
        btn_frame.pack(side='right', padx=20, pady=10)

        tk.Button(btn_frame, text="+ 新增商品",
                 font=('微软雅黑', 11),
                 bg='#2ECC71', fg='white',
                 pady=5, padx=15,
                 cursor='hand2',
                 command=self.add_product).pack(side='left', padx=5)

        tk.Button(btn_frame, text="导入商品",
                 font=('微软雅黑', 11),
                 bg='#3498DB', fg='white',
                 pady=5, padx=15,
                 cursor='hand2',
                 command=self.import_products).pack(side='left', padx=5)

        tk.Button(btn_frame, text="导出商品",
                 font=('微软雅黑', 11),
                 bg='#27AE60', fg='white',
                 pady=5, padx=15,
                 cursor='hand2',
                 command=self.export_products).pack(side='left', padx=5)

        tk.Button(btn_frame, text="× 删除商品",
                 font=('微软雅黑', 11),
                 bg='#E74C3C', fg='white',
                 pady=5, padx=15,
                 cursor='hand2',
                 command=self.delete_product).pack(side='left', padx=5)

        # 商品列表
        list_frame = tk.Frame(container, bg='white', relief='raised', bd=1)
        list_frame.pack(fill='both', expand=True, pady=10)

        # 创建商品表格
        self.create_products_table(list_frame)

    def create_products_table(self, parent):
        """创建商品表格"""
        # 表格工具栏
        toolbar = tk.Frame(parent, bg='white')
        toolbar.pack(fill='x', padx=20, pady=10)

        tk.Label(toolbar, text="商品列表",
                font=('微软雅黑', 14, 'bold'),
                bg='white').pack(side='left')

        # 表格框架
        table_frame = tk.Frame(parent, bg='white')
        table_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        # 创建Treeview
        columns = ('编码', '名称', '分类', '规格型号', '单位', '采购价', '销售价', '库存', '状态')
        self.products_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

        # 设置列标题和宽度
        column_widths = [80, 150, 100, 150, 60, 80, 80, 80, 80]
        for i, col in enumerate(columns):
            self.products_tree.heading(col, text=col)
            self.products_tree.column(col, width=column_widths[i])

        # 添加滚动条
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.products_tree.yview)
        self.products_tree.configure(yscrollcommand=scrollbar.set)

        self.products_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # 加载数据
        self.load_products_data()

    def load_products_data(self):
        """加载商品数据"""
        # 清空现有数据
        for item in self.products_tree.get_children():
            self.products_tree.delete(item)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        cursor.execute('''
            SELECT p.code, p.name, c.name, p.specification, p.unit,
                   p.purchase_price, p.sale_price,
                   COALESCE(SUM(i.quantity), 0), p.status
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN inventory i ON p.id = i.product_id
            GROUP BY p.id
            ORDER BY p.created_time DESC
        ''')

        records = cursor.fetchall()

        for record in records:
            status_text = '启用' if record[8] == 1 else '停用'
            self.products_tree.insert('', 'end', values=(
                record[0], record[1], record[2] or '未分类', record[3] or '',
                record[4], f"¥{record[5] or 0:.2f}", f"¥{record[6] or 0:.2f}",
                record[7], status_text
            ))

        conn.close()

    def import_products(self):
        """导入商品（CSV格式）"""
        filename = filedialog.askopenfilename(
            title="选择CSV文件",
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        if not filename:
            return

        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()

            imported_count = 0
            skipped_count = 0
            error_list = []

            with open(filename, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        code = row.get('商品编码', '').strip()
                        name = row.get('商品名称', '').strip()

                        if not code or not name:
                            skipped_count += 1
                            continue

                        # 检查商品编码是否已存在
                        cursor.execute("SELECT id FROM products WHERE code=?", (code,))
                        if cursor.fetchone():
                            skipped_count += 1
                            error_list.append(f"{code} - {name}（编码已存在）")
                            continue

                        # 插入商品
                        cursor.execute('''
                            INSERT INTO products
                            (code, name, specification, unit, purchase_price, sale_price, min_stock, max_stock)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            code,
                            name,
                            row.get('规格型号', '').strip(),
                            row.get('单位', '个').strip(),
                            float(row.get('采购价', 0) or 0),
                            float(row.get('销售价', 0) or 0),
                            int(row.get('最小库存', 0) or 0),
                            int(row.get('最大库存', 999999) or 999999)
                        ))
                        imported_count += 1
                    except Exception as e:
                        skipped_count += 1
                        error_list.append(f"{row.get('商品编码', '未知')}: {str(e)}")

            conn.commit()
            conn.close()

            result_msg = f"导入完成！\n\n成功导入：{imported_count} 条\n跳过：{skipped_count} 条"
            if error_list:
                result_msg += f"\n\n错误详情：\n" + "\n".join(error_list[:10])
                if len(error_list) > 10:
                    result_msg += f"\n... 还有 {len(error_list) - 10} 条错误"

            messagebox.showinfo("导入结果", result_msg)
            self.load_products_data()  # 刷新列表

        except Exception as e:
            messagebox.showerror("错误", f"导入失败：{str(e)}")

    def export_products(self):
        """导出商品（CSV格式）"""
        filename = filedialog.asksaveasfilename(
            title="保存CSV文件",
            defaultextension=".csv",
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        if not filename:
            return

        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()

            cursor.execute('''
                SELECT code, name, specification, unit, purchase_price, sale_price, min_stock, max_stock
                FROM products
                WHERE status = 1
                ORDER BY code
            ''')

            with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['商品编码', '商品名称', '规格型号', '单位', '采购价', '销售价', '最小库存', '最大库存'])
                writer.writerows(cursor.fetchall())

            conn.close()
            messagebox.showinfo("成功", f"商品数据已导出到：\n{filename}")

        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")

    def delete_product(self):
        """删除商品"""
        selection = self.products_tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先选择要删除的商品")
            return

        item = self.products_tree.item(selection[0])
        product_code = item['values'][0]
        product_name = item['values'][1]

        # 检查是否有库存
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT SUM(quantity) FROM inventory
            WHERE product_id = (SELECT id FROM products WHERE code=?)
        ''', (product_code,))
        stock_result = cursor.fetchone()
        total_stock = stock_result[0] if stock_result[0] else 0

        if total_stock > 0:
            messagebox.showerror("错误", f"商品 {product_name} 还有库存 {total_stock}，无法删除！\n\n请先清空库存后再删除。")
            conn.close()
            return

        if not messagebox.askyesno("确认删除", f"确定要删除商品 {product_name} ({product_code}) 吗？\n\n此操作不可恢复！"):
            conn.close()
            return

        try:
            # 删除商品（软删除，设置为停用状态）
            cursor.execute("UPDATE products SET status=0 WHERE code=?", (product_code,))
            conn.commit()
            messagebox.showinfo("成功", "商品已删除！")
            self.load_products_data()  # 刷新列表
        except Exception as e:
            conn.rollback()
            messagebox.showerror("错误", f"删除失败：{str(e)}")
        finally:
            conn.close()

    def show_warehouses(self, container=None):
        """显示仓库管理页面"""
        if container is None:
            container = self.page_container if hasattr(self, 'page_container') else self.notebook.nametowidget(self.notebook.select())
        
        # 标题栏
        title_bar = tk.Frame(container, bg='white', height=60, relief='raised', bd=1)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)

        tk.Label(title_bar, text="仓库管理",
                font=('微软雅黑', 18, 'bold'),
                bg='white').pack(side='left', padx=20, pady=15)

        # 按钮组
        btn_frame = tk.Frame(title_bar, bg='white')
        btn_frame.pack(side='right', padx=20, pady=10)

        tk.Button(btn_frame, text="+ 新增仓库",
                 font=('微软雅黑', 11),
                 bg='#2ECC71', fg='white',
                 pady=5, padx=15,
                 cursor='hand2',
                 command=self.add_warehouse).pack(side='left', padx=5)

        tk.Button(btn_frame, text="× 删除仓库",
                 font=('微软雅黑', 11),
                 bg='#E74C3C', fg='white',
                 pady=5, padx=15,
                 cursor='hand2',
                 command=self.delete_warehouse).pack(side='left', padx=5)

        # 仓库列表
        list_frame = tk.Frame(container, bg='white', relief='raised', bd=1)
        list_frame.pack(fill='both', expand=True, pady=10)

        # 创建仓库表格
        self.create_warehouses_table(list_frame)

    def create_warehouses_table(self, parent):
        """创建仓库表格"""
        # 表格框架
        table_frame = tk.Frame(parent, bg='white')
        table_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # 创建Treeview
        columns = ('仓库名称', '地址', '负责人', '联系电话', '状态')
        self.warehouses_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

        # 设置列标题和宽度
        column_widths = [150, 200, 100, 120, 80]
        for i, col in enumerate(columns):
            self.warehouses_tree.heading(col, text=col)
            self.warehouses_tree.column(col, width=column_widths[i])

        # 添加滚动条
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.warehouses_tree.yview)
        self.warehouses_tree.configure(yscrollcommand=scrollbar.set)

        self.warehouses_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # 加载数据
        self.load_warehouses_data()

    def load_warehouses_data(self):
        """加载仓库数据"""
        if not hasattr(self, 'warehouses_tree'):
            return
        
        # 清空现有数据
        for item in self.warehouses_tree.get_children():
            self.warehouses_tree.delete(item)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT name, address, manager, contact, status FROM warehouses ORDER BY name")
        warehouses = cursor.fetchall()
        conn.close()

        for warehouse in warehouses:
            status_text = '启用' if warehouse[4] == 1 else '停用'
            self.warehouses_tree.insert('', 'end', values=(
                warehouse[0] or '',  # 仓库名称
                warehouse[1] or '',  # 地址
                warehouse[2] or '',  # 负责人
                warehouse[3] or '',  # 联系电话
                status_text  # 状态
            ))

    def add_warehouse(self):
        """新增仓库"""
        # #region agent log
        import json, os
        log_path = os.path.join(BASE_DIR, '.cursor', 'debug.log')
        try:
            os.makedirs(os.path.dirname(log_path), exist_ok=True)
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"A","location":"inventory_pro.py:2775","message":"add_warehouse called","data":{},"timestamp":int(__import__('time').time()*1000)})+'\n')
        except Exception as e:
            print(f"Log error: {e}")
        # #endregion
        add_win = tk.Toplevel(self.root)
        add_win.title("新增仓库")
        add_win.geometry("450x380")
        add_win.transient(self.root)
        add_win.grab_set()
        add_win.resizable(False, False)

        # 居中
        add_win.update_idletasks()
        x = (add_win.winfo_screenwidth() // 2) - 225
        y = (add_win.winfo_screenheight() // 2) - 190
        add_win.geometry(f'450x380+{x}+{y}')
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"C","location":"inventory_pro.py:2795","message":"Window geometry set","data":{"x":x,"y":y},"timestamp":int(__import__('time').time()*1000)})+'\n')
        except: pass
        # #endregion

        # 先创建按钮栏（固定在底部）
        btn_frame = tk.Frame(add_win, bg='#ECF0F1', height=70)
        btn_frame.pack(fill='x', side='bottom')
        btn_frame.pack_propagate(False)

        # 内容区域（在按钮栏上方）
        content_frame = tk.Frame(add_win, bg='white')
        content_frame.pack(fill='both', expand=True, padx=30, pady=20)

        # 仓库名称
        tk.Label(content_frame, text="仓库名称*：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        name_var = tk.StringVar()
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"D","location":"inventory_pro.py:2816","message":"name_var created","data":{"value":name_var.get()},"timestamp":int(__import__('time').time()*1000)})+'\n')
        except: pass
        # #endregion
        name_entry = tk.Entry(content_frame, textvariable=name_var,
                             font=('微软雅黑', 11), width=30)
        name_entry.pack(fill='x', pady=(0, 15))
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"D","location":"inventory_pro.py:2825","message":"name_entry created and packed","data":{"var_value":name_var.get(),"entry_value":name_entry.get()},"timestamp":int(__import__('time').time()*1000)})+'\n')
        except: pass
        # #endregion
        
        # 绑定事件监听StringVar变化
        def on_var_change(*args):
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"A","location":"inventory_pro.py:2834","message":"name_var changed","data":{"new_value":name_var.get()},"timestamp":int(__import__('time').time()*1000)})+'\n')
            except: pass
            # #endregion
        name_var.trace_add('write', on_var_change)
        
        # 绑定焦点事件
        def on_entry_focus_in(e):
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"B","location":"inventory_pro.py:2844","message":"Entry focus in","data":{"var_value":name_var.get(),"entry_value":name_entry.get()},"timestamp":int(__import__('time').time()*1000)})+'\n')
            except: pass
            # #endregion
        def on_entry_focus_out(e):
            # #region agent log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"B","location":"inventory_pro.py:2850","message":"Entry focus out","data":{"var_value":name_var.get(),"entry_value":name_entry.get()},"timestamp":int(__import__('time').time()*1000)})+'\n')
            except: pass
            # #endregion
        name_entry.bind('<FocusIn>', on_entry_focus_in)
        name_entry.bind('<FocusOut>', on_entry_focus_out)
        
        # 延迟检查值
        def check_value():
            # #region agent log
            try:
                var_val = name_var.get()
                entry_val = name_entry.get()
                entry_state = name_entry.cget('state')
                entry_has_focus = name_entry.focus_get() == name_entry
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"E","location":"inventory_pro.py:2860","message":"Delayed value check","data":{"var_value":var_val,"entry_value":entry_val,"entry_state":entry_state,"has_focus":entry_has_focus},"timestamp":int(__import__('time').time()*1000)})+'\n')
            except Exception as e:
                try:
                    with open(log_path, 'a', encoding='utf-8') as f:
                        f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"E","location":"inventory_pro.py:2860","message":"Check value error","data":{"error":str(e)},"timestamp":int(__import__('time').time()*1000)})+'\n')
                except: pass
            # #endregion
        add_win.after(200, check_value)
        add_win.after(500, check_value)
        add_win.after(1000, check_value)
        add_win.after(2000, check_value)
        
        # 监听窗口配置事件
        def on_configure(e):
            # #region agent log
            try:
                var_val = name_var.get()
                entry_val = name_entry.get()
                entry_text = name_entry.get()
                # 检查是否有内容但显示为空
                has_content_but_empty_display = (var_val != '' and entry_text == '')
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"C","location":"inventory_pro.py:2886","message":"Window configured","data":{"var_value":var_val,"entry_value":entry_val,"entry_text":entry_text,"has_content_but_empty":has_content_but_empty_display,"event_type":str(type(e))},"timestamp":int(__import__('time').time()*1000)})+'\n')
                # 如果StringVar有值但Entry显示为空，强制同步
                if has_content_but_empty_display:
                    name_entry.delete(0, tk.END)
                    name_entry.insert(0, var_val)
                    with open(log_path, 'a', encoding='utf-8') as f:
                        f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"C","location":"inventory_pro.py:2897","message":"Forced sync entry with var","data":{"var_value":var_val},"timestamp":int(__import__('time').time()*1000)})+'\n')
            except Exception as ex:
                try:
                    with open(log_path, 'a', encoding='utf-8') as f:
                        f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"C","location":"inventory_pro.py:2886","message":"Configure error","data":{"error":str(ex)},"timestamp":int(__import__('time').time()*1000)})+'\n')
                except: pass
            # #endregion
        add_win.bind('<Configure>', on_configure)
        
        # 监听Entry的<Key>事件
        def on_key(e):
            # #region agent log
            try:
                key_char = e.char if hasattr(e, 'char') else ''
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"F","location":"inventory_pro.py:2882","message":"Key pressed in entry","data":{"key":key_char,"var_value":name_var.get(),"entry_value":name_entry.get()},"timestamp":int(__import__('time').time()*1000)})+'\n')
            except: pass
            # #endregion
        name_entry.bind('<Key>', on_key)

        # 地址
        tk.Label(content_frame, text="地址：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        address_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=address_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 15))

        # 负责人
        tk.Label(content_frame, text="负责人：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        manager_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=manager_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 15))

        # 联系电话
        tk.Label(content_frame, text="联系电话：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        contact_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=contact_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 15))

        # 按钮容器
        btn_container = tk.Frame(btn_frame, bg='#ECF0F1')
        btn_container.pack(expand=True, pady=15)

        def save_warehouse():
            name = name_var.get().strip()
            if not name:
                messagebox.showerror("错误", "请输入仓库名称！")
                name_entry.focus()
                return

            try:
                conn = sqlite3.connect(DB_FILE)
                cursor = conn.cursor()

                # 检查仓库名称是否已存在
                cursor.execute("SELECT id FROM warehouses WHERE name=?", (name,))
                if cursor.fetchone():
                    messagebox.showerror("错误", "仓库名称已存在！")
                    conn.close()
                    return

                # 插入新仓库
                cursor.execute("INSERT INTO warehouses (name, address, manager, contact, status) VALUES (?, ?, ?, ?, 1)",
                             (name, address_var.get().strip(), manager_var.get().strip(), contact_var.get().strip()))
                conn.commit()
                conn.close()

                messagebox.showinfo("成功", f"仓库 {name} 添加成功！")
                add_win.destroy()
                if hasattr(self, 'warehouses_tree'):
                    self.load_warehouses_data()  # 刷新列表
            except sqlite3.IntegrityError:
                messagebox.showerror("错误", "仓库名称已存在！")
            except Exception as e:
                messagebox.showerror("错误", f"添加仓库失败：{str(e)}")
            finally:
                try:
                    conn.close()
                except:
                    pass

        tk.Button(btn_container, text="保存",
                 font=('微软雅黑', 11),
                 bg='#2ECC71', fg='white',
                 pady=5, padx=20,
                 command=save_warehouse).pack(side='left', padx=5)

        tk.Button(btn_container, text="取消",
                 font=('微软雅黑', 11),
                 bg='#95A5A6', fg='white',
                 pady=5, padx=20,
                 command=add_win.destroy).pack(side='left', padx=5)

    def delete_warehouse(self):
        """删除仓库"""
        if not hasattr(self, 'warehouses_tree'):
            messagebox.showwarning("提示", "请先打开仓库管理页面！")
            return

        selection = self.warehouses_tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先选择要删除的仓库！")
            return

        item = self.warehouses_tree.item(selection[0])
        warehouse_name = item['values'][0]

        # 检查仓库是否被使用
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()

            # 检查是否有库存
            cursor.execute("SELECT COUNT(*) FROM inventory i JOIN warehouses w ON i.warehouse_id = w.id WHERE w.name=?", (warehouse_name,))
            inventory_count = cursor.fetchone()[0]

            # 检查是否有入库单
            cursor.execute("SELECT COUNT(*) FROM inbound_orders io JOIN warehouses w ON io.warehouse_id = w.id WHERE w.name=?", (warehouse_name,))
            inbound_count = cursor.fetchone()[0]

            # 检查是否有出库单
            cursor.execute("SELECT COUNT(*) FROM outbound_orders oo JOIN warehouses w ON oo.warehouse_id = w.id WHERE w.name=?", (warehouse_name,))
            outbound_count = cursor.fetchone()[0]

            conn.close()

            if inventory_count > 0 or inbound_count > 0 or outbound_count > 0:
                messagebox.showerror("错误", 
                    f"仓库 {warehouse_name} 正在使用中，无法删除！\n\n"
                    f"当前库存记录：{inventory_count} 条\n"
                    f"入库单：{inbound_count} 条\n"
                    f"出库单：{outbound_count} 条\n\n"
                    "如需删除，请先清空相关数据。")
                return

            if not messagebox.askyesno("确认删除", f"确定要删除仓库 {warehouse_name} 吗？\n\n此操作不可恢复！"):
                return

            # 执行删除（软删除，设置为停用状态）
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("UPDATE warehouses SET status=0 WHERE name=?", (warehouse_name,))
            conn.commit()
            conn.close()

            messagebox.showinfo("成功", "仓库已删除！")
            self.load_warehouses_data()  # 刷新列表

        except Exception as e:
            messagebox.showerror("错误", f"删除仓库失败：{str(e)}")

    def show_basic_management(self, container=None):
        """显示基础管理页面（包含商品、供应商、客户管理）"""
        if container is None:
            container = self.page_container if hasattr(self, 'page_container') else self.notebook.nametowidget(self.notebook.select())
        
        # 标题栏
        title_bar = tk.Frame(container, bg='white', height=60, relief='raised', bd=1)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)

        tk.Label(title_bar, text="基础管理",
                font=('微软雅黑', 18, 'bold'),
                bg='white').pack(side='left', padx=20, pady=15)

        # 创建子功能选项卡
        notebook = ttk.Notebook(container)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)

        # 获取当前用户的权限
        current_user = self.user[1]
        user_permissions = self.get_user_permissions(current_user)
        is_admin = self.user[2] == 'admin'

        # 商品管理
        if is_admin or 'products' in user_permissions:
            products_frame = tk.Frame(notebook, bg='white')
            notebook.add(products_frame, text='📝 商品管理')
            self.show_products(products_frame)

        # 供应商管理
        if is_admin or 'suppliers' in user_permissions:
            suppliers_frame = tk.Frame(notebook, bg='white')
            notebook.add(suppliers_frame, text='🏢 供应商管理')
            self.show_suppliers(suppliers_frame)

        # 客户管理
        if is_admin or 'customers' in user_permissions:
            customers_frame = tk.Frame(notebook, bg='white')
            notebook.add(customers_frame, text='👥 客户管理')
            self.show_customers(customers_frame)

        # 仓库管理
        if is_admin or 'warehouses' in user_permissions:
            warehouses_frame = tk.Frame(notebook, bg='white')
            notebook.add(warehouses_frame, text='🏭 仓库管理')
            self.show_warehouses(warehouses_frame)

    def show_suppliers(self, container=None):
        """显示供应商管理页面"""
        if container is None:
            container = self.page_container if hasattr(self, 'page_container') else self.notebook.nametowidget(self.notebook.select())
        
        # 标题栏
        title_bar = tk.Frame(container, bg='white', height=60, relief='raised', bd=1)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)

        tk.Label(title_bar, text="供应商管理",
                font=('微软雅黑', 18, 'bold'),
                bg='white').pack(side='left', padx=20, pady=15)

        # 按钮组
        btn_frame = tk.Frame(title_bar, bg='white')
        btn_frame.pack(side='right', padx=20, pady=10)

        tk.Button(btn_frame, text="+ 新增供应商",
                 font=('微软雅黑', 11),
                 bg='#2ECC71', fg='white',
                 pady=5, padx=15,
                 cursor='hand2',
                 command=self.add_supplier).pack(side='left', padx=5)

        tk.Button(btn_frame, text="× 删除供应商",
                 font=('微软雅黑', 11),
                 bg='#E74C3C', fg='white',
                 pady=5, padx=15,
                 cursor='hand2',
                 command=self.delete_supplier).pack(side='left', padx=5)

        # 供应商列表
        list_frame = tk.Frame(container, bg='white', relief='raised', bd=1)
        list_frame.pack(fill='both', expand=True, pady=10)

        # 创建供应商表格
        self.create_suppliers_table(list_frame)

    def create_suppliers_table(self, parent):
        """创建供应商表格"""
        # 表格框架
        table_frame = tk.Frame(parent, bg='white')
        table_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # 创建Treeview
        columns = ('供应商名称', '联系人', '电话', '地址', '邮箱', '状态')
        self.suppliers_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

        # 设置列标题和宽度
        column_widths = [150, 100, 120, 200, 150, 80]
        for i, col in enumerate(columns):
            self.suppliers_tree.heading(col, text=col)
            self.suppliers_tree.column(col, width=column_widths[i])

        # 添加滚动条
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.suppliers_tree.yview)
        self.suppliers_tree.configure(yscrollcommand=scrollbar.set)

        self.suppliers_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # 加载数据
        self.load_suppliers_data()

    def load_suppliers_data(self):
        """加载供应商数据"""
        if not hasattr(self, 'suppliers_tree'):
            return
        
        # 清空现有数据
        for item in self.suppliers_tree.get_children():
            self.suppliers_tree.delete(item)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT name, contact_person, phone, address, email, status FROM suppliers ORDER BY name")
        suppliers = cursor.fetchall()
        conn.close()

        for supplier in suppliers:
            status_text = '启用' if supplier[5] == 1 else '停用'
            self.suppliers_tree.insert('', 'end', values=(
                supplier[0] or '',  # 供应商名称
                supplier[1] or '',  # 联系人
                supplier[2] or '',  # 电话
                supplier[3] or '',  # 地址
                supplier[4] or '',  # 邮箱
                status_text  # 状态
            ))

    def add_supplier(self):
        """新增供应商"""
        add_win = tk.Toplevel(self.root)
        add_win.title("新增供应商")
        add_win.geometry("450x480")
        add_win.transient(self.root)
        add_win.grab_set()
        add_win.resizable(False, False)

        # 居中
        add_win.update_idletasks()
        x = (add_win.winfo_screenwidth() // 2) - 225
        y = (add_win.winfo_screenheight() // 2) - 240
        add_win.geometry(f'450x480+{x}+{y}')

        # 先创建按钮栏（固定在底部）
        btn_frame = tk.Frame(add_win, bg='#ECF0F1', height=70)
        btn_frame.pack(fill='x', side='bottom')
        btn_frame.pack_propagate(False)

        # 内容区域（在按钮栏上方）
        content_frame = tk.Frame(add_win, bg='white')
        content_frame.pack(fill='both', expand=True, padx=30, pady=15)

        # 供应商名称
        tk.Label(content_frame, text="供应商名称*：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        name_var = tk.StringVar()
        name_entry = tk.Entry(content_frame, textvariable=name_var,
                             font=('微软雅黑', 11), width=30)
        name_entry.pack(fill='x', pady=(0, 12))
        name_entry.focus()

        # 联系人
        tk.Label(content_frame, text="联系人：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        contact_person_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=contact_person_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 电话
        tk.Label(content_frame, text="电话：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        phone_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=phone_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 地址
        tk.Label(content_frame, text="地址：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        address_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=address_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 邮箱
        tk.Label(content_frame, text="邮箱：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        email_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=email_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 按钮容器
        btn_container = tk.Frame(btn_frame, bg='#ECF0F1')
        btn_container.pack(expand=True, pady=15)

        def save_supplier():
            name = name_var.get().strip()
            if not name:
                messagebox.showerror("错误", "请输入供应商名称！")
                name_entry.focus()
                return

            try:
                conn = sqlite3.connect(DB_FILE)
                cursor = conn.cursor()

                # 检查供应商名称是否已存在
                cursor.execute("SELECT id FROM suppliers WHERE name=?", (name,))
                if cursor.fetchone():
                    messagebox.showerror("错误", "供应商名称已存在！")
                    conn.close()
                    return

                # 插入新供应商
                cursor.execute("INSERT INTO suppliers (name, contact_person, phone, address, email, status) VALUES (?, ?, ?, ?, ?, 1)",
                             (name, contact_person_var.get().strip(), phone_var.get().strip(), 
                              address_var.get().strip(), email_var.get().strip()))
                conn.commit()
                conn.close()

                messagebox.showinfo("成功", f"供应商 {name} 添加成功！")
                add_win.destroy()
                if hasattr(self, 'suppliers_tree'):
                    self.load_suppliers_data()  # 刷新列表
            except sqlite3.IntegrityError:
                messagebox.showerror("错误", "供应商名称已存在！")
            except Exception as e:
                messagebox.showerror("错误", f"添加供应商失败：{str(e)}")
            finally:
                # 确保连接关闭（如果之前没有关闭）
                try:
                    if 'conn' in locals() and conn:
                        conn.close()
                except:
                    pass

        tk.Button(btn_container, text="保存",
                 font=('微软雅黑', 11),
                 bg='#2ECC71', fg='white',
                 pady=5, padx=20,
                 command=save_supplier).pack(side='left', padx=5)

        tk.Button(btn_container, text="取消",
                 font=('微软雅黑', 11),
                 bg='#95A5A6', fg='white',
                 pady=5, padx=20,
                 command=add_win.destroy).pack(side='left', padx=5)

    def delete_supplier(self):
        """删除供应商"""
        if not hasattr(self, 'suppliers_tree'):
            messagebox.showwarning("提示", "请先打开供应商管理页面！")
            return

        selection = self.suppliers_tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先选择要删除的供应商！")
            return

        item = self.suppliers_tree.item(selection[0])
        supplier_name = item['values'][0]

        # 检查供应商是否被使用
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()

            # 检查是否有入库单
            cursor.execute("SELECT COUNT(*) FROM inbound_orders WHERE supplier=?", (supplier_name,))
            inbound_count = cursor.fetchone()[0]

            conn.close()

            if inbound_count > 0:
                messagebox.showerror("错误", 
                    f"供应商 {supplier_name} 正在使用中，无法删除！\n\n"
                    f"入库单：{inbound_count} 条\n\n"
                    "如需删除，请先删除相关入库单。")
                return

            if not messagebox.askyesno("确认删除", f"确定要删除供应商 {supplier_name} 吗？\n\n此操作不可恢复！"):
                return

            # 执行删除（软删除，设置为停用状态）
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("UPDATE suppliers SET status=0 WHERE name=?", (supplier_name,))
            conn.commit()
            conn.close()

            messagebox.showinfo("成功", "供应商已删除！")
            self.load_suppliers_data()  # 刷新列表

        except Exception as e:
            messagebox.showerror("错误", f"删除供应商失败：{str(e)}")

    def quick_add_supplier(self, supplier_combo, supplier_var, parent_win):
        """快捷新增供应商（从入库单界面调用）"""
        add_win = tk.Toplevel(parent_win)
        add_win.title("快捷新增供应商")
        add_win.geometry("450x480")
        add_win.transient(parent_win)
        add_win.grab_set()
        add_win.resizable(False, False)

        # 居中
        add_win.update_idletasks()
        x = (add_win.winfo_screenwidth() // 2) - 225
        y = (add_win.winfo_screenheight() // 2) - 240
        add_win.geometry(f'450x480+{x}+{y}')

        # 先创建按钮栏（固定在底部）
        btn_frame = tk.Frame(add_win, bg='#ECF0F1', height=70)
        btn_frame.pack(fill='x', side='bottom')
        btn_frame.pack_propagate(False)

        # 内容区域（在按钮栏上方）
        content_frame = tk.Frame(add_win, bg='white')
        content_frame.pack(fill='both', expand=True, padx=30, pady=15)

        # 供应商名称
        tk.Label(content_frame, text="供应商名称*：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        name_var = tk.StringVar()
        name_entry = tk.Entry(content_frame, textvariable=name_var,
                             font=('微软雅黑', 11), width=30)
        name_entry.pack(fill='x', pady=(0, 12))
        name_entry.focus()

        # 联系人
        tk.Label(content_frame, text="联系人：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        contact_person_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=contact_person_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 电话
        tk.Label(content_frame, text="电话：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        phone_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=phone_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 地址
        tk.Label(content_frame, text="地址：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        address_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=address_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 邮箱
        tk.Label(content_frame, text="邮箱：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        email_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=email_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 按钮容器
        btn_container = tk.Frame(btn_frame, bg='#ECF0F1')
        btn_container.pack(expand=True, pady=15)

        def save_and_refresh():
            name = name_var.get().strip()
            if not name:
                messagebox.showerror("错误", "请输入供应商名称！")
                name_entry.focus()
                return

            try:
                conn = sqlite3.connect(DB_FILE)
                cursor = conn.cursor()

                # 检查供应商名称是否已存在
                cursor.execute("SELECT id FROM suppliers WHERE name=?", (name,))
                if cursor.fetchone():
                    messagebox.showerror("错误", "供应商名称已存在！")
                    conn.close()
                    return

                # 插入新供应商
                cursor.execute("INSERT INTO suppliers (name, contact_person, phone, address, email, status) VALUES (?, ?, ?, ?, ?, 1)",
                             (name, contact_person_var.get().strip(), phone_var.get().strip(), 
                              address_var.get().strip(), email_var.get().strip()))
                conn.commit()
                conn.close()

                # 刷新供应商下拉框
                suppliers = self.get_suppliers_list()
                supplier_combo['values'] = suppliers
                # 选中新添加的供应商
                supplier_var.set(name)
                
                messagebox.showinfo("成功", f"供应商 {name} 添加成功！\n\n已自动选择该供应商。")
                add_win.destroy()
                if hasattr(self, 'suppliers_tree'):
                    self.load_suppliers_data()  # 刷新列表（如果供应商管理页面已打开）
            except sqlite3.IntegrityError:
                messagebox.showerror("错误", "供应商名称已存在！")
            except Exception as e:
                messagebox.showerror("错误", f"添加供应商失败：{str(e)}")
            finally:
                # 确保连接关闭（如果之前没有关闭）
                try:
                    if 'conn' in locals() and conn:
                        conn.close()
                except:
                    pass

        tk.Button(btn_container, text="保存",
                 font=('微软雅黑', 11),
                 bg='#2ECC71', fg='white',
                 pady=5, padx=20,
                 command=save_and_refresh).pack(side='left', padx=5)

        tk.Button(btn_container, text="取消",
                 font=('微软雅黑', 11),
                 bg='#95A5A6', fg='white',
                 pady=5, padx=20,
                 command=add_win.destroy).pack(side='left', padx=5)

    def show_customers(self, container=None):
        """显示客户管理页面"""
        if container is None:
            container = self.page_container if hasattr(self, 'page_container') else self.notebook.nametowidget(self.notebook.select())
        
        # 标题栏
        title_bar = tk.Frame(container, bg='white', height=60, relief='raised', bd=1)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)

        tk.Label(title_bar, text="客户管理",
                font=('微软雅黑', 18, 'bold'),
                bg='white').pack(side='left', padx=20, pady=15)

        # 按钮组
        btn_frame = tk.Frame(title_bar, bg='white')
        btn_frame.pack(side='right', padx=20, pady=10)

        tk.Button(btn_frame, text="+ 新增客户",
                 font=('微软雅黑', 11),
                 bg='#2ECC71', fg='white',
                 pady=5, padx=15,
                 cursor='hand2',
                 command=self.add_customer).pack(side='left', padx=5)

        tk.Button(btn_frame, text="× 删除客户",
                 font=('微软雅黑', 11),
                 bg='#E74C3C', fg='white',
                 pady=5, padx=15,
                 cursor='hand2',
                 command=self.delete_customer).pack(side='left', padx=5)

        # 客户列表
        list_frame = tk.Frame(container, bg='white', relief='raised', bd=1)
        list_frame.pack(fill='both', expand=True, pady=10)

        # 创建客户表格
        self.create_customers_table(list_frame)

    def create_customers_table(self, parent):
        """创建客户表格"""
        # 表格框架
        table_frame = tk.Frame(parent, bg='white')
        table_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # 创建Treeview
        columns = ('客户名称', '联系人', '电话', '地址', '邮箱', '状态')
        self.customers_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

        # 设置列标题和宽度
        column_widths = [150, 100, 120, 200, 150, 80]
        for i, col in enumerate(columns):
            self.customers_tree.heading(col, text=col)
            self.customers_tree.column(col, width=column_widths[i])

        # 添加滚动条
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.customers_tree.yview)
        self.customers_tree.configure(yscrollcommand=scrollbar.set)

        self.customers_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # 加载数据
        self.load_customers_data()

    def load_customers_data(self):
        """加载客户数据"""
        if not hasattr(self, 'customers_tree'):
            return
        
        # 清空现有数据
        for item in self.customers_tree.get_children():
            self.customers_tree.delete(item)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT name, contact_person, phone, address, email, status FROM customers ORDER BY name")
        customers = cursor.fetchall()
        conn.close()

        for customer in customers:
            status_text = '启用' if customer[5] == 1 else '停用'
            self.customers_tree.insert('', 'end', values=(
                customer[0] or '',  # 客户名称
                customer[1] or '',  # 联系人
                customer[2] or '',  # 电话
                customer[3] or '',  # 地址
                customer[4] or '',  # 邮箱
                status_text  # 状态
            ))

    def add_customer(self):
        """新增客户"""
        add_win = tk.Toplevel(self.root)
        add_win.title("新增客户")
        add_win.geometry("450x480")
        add_win.transient(self.root)
        add_win.grab_set()
        add_win.resizable(False, False)

        # 居中
        add_win.update_idletasks()
        x = (add_win.winfo_screenwidth() // 2) - 225
        y = (add_win.winfo_screenheight() // 2) - 240
        add_win.geometry(f'450x480+{x}+{y}')

        # 先创建按钮栏（固定在底部）
        btn_frame = tk.Frame(add_win, bg='#ECF0F1', height=70)
        btn_frame.pack(fill='x', side='bottom')
        btn_frame.pack_propagate(False)

        # 内容区域（在按钮栏上方）
        content_frame = tk.Frame(add_win, bg='white')
        content_frame.pack(fill='both', expand=True, padx=30, pady=15)

        # 客户名称
        tk.Label(content_frame, text="客户名称*：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        name_var = tk.StringVar()
        name_entry = tk.Entry(content_frame, textvariable=name_var,
                             font=('微软雅黑', 11), width=30)
        name_entry.pack(fill='x', pady=(0, 12))
        name_entry.focus()

        # 联系人
        tk.Label(content_frame, text="联系人：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        contact_person_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=contact_person_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 电话
        tk.Label(content_frame, text="电话：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        phone_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=phone_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 地址
        tk.Label(content_frame, text="地址：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        address_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=address_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 邮箱
        tk.Label(content_frame, text="邮箱：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        email_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=email_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 按钮容器
        btn_container = tk.Frame(btn_frame, bg='#ECF0F1')
        btn_container.pack(expand=True, pady=15)

        def save_customer():
            name = name_var.get().strip()
            if not name:
                messagebox.showerror("错误", "请输入客户名称！")
                name_entry.focus()
                return

            try:
                conn = sqlite3.connect(DB_FILE)
                cursor = conn.cursor()

                # 检查客户名称是否已存在
                cursor.execute("SELECT id FROM customers WHERE name=?", (name,))
                if cursor.fetchone():
                    messagebox.showerror("错误", "客户名称已存在！")
                    conn.close()
                    return

                # 插入新客户
                cursor.execute("INSERT INTO customers (name, contact_person, phone, address, email, status) VALUES (?, ?, ?, ?, ?, 1)",
                             (name, contact_person_var.get().strip(), phone_var.get().strip(), 
                              address_var.get().strip(), email_var.get().strip()))
                conn.commit()
                conn.close()

                messagebox.showinfo("成功", f"客户 {name} 添加成功！")
                add_win.destroy()
                if hasattr(self, 'customers_tree'):
                    self.load_customers_data()  # 刷新列表
            except sqlite3.IntegrityError:
                messagebox.showerror("错误", "客户名称已存在！")
            except Exception as e:
                messagebox.showerror("错误", f"添加客户失败：{str(e)}")
            finally:
                # 确保连接关闭（如果之前没有关闭）
                try:
                    if 'conn' in locals() and conn:
                        conn.close()
                except:
                    pass

        tk.Button(btn_container, text="保存",
                 font=('微软雅黑', 11),
                 bg='#2ECC71', fg='white',
                 pady=5, padx=20,
                 command=save_customer).pack(side='left', padx=5)

        tk.Button(btn_container, text="取消",
                 font=('微软雅黑', 11),
                 bg='#95A5A6', fg='white',
                 pady=5, padx=20,
                 command=add_win.destroy).pack(side='left', padx=5)

    def delete_customer(self):
        """删除客户"""
        if not hasattr(self, 'customers_tree'):
            messagebox.showwarning("提示", "请先打开客户管理页面！")
            return

        selection = self.customers_tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先选择要删除的客户！")
            return

        item = self.customers_tree.item(selection[0])
        customer_name = item['values'][0]

        # 检查客户是否被使用
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()

            # 检查是否有出库单
            cursor.execute("SELECT COUNT(*) FROM outbound_orders WHERE customer=?", (customer_name,))
            outbound_count = cursor.fetchone()[0]

            conn.close()

            if outbound_count > 0:
                messagebox.showerror("错误", 
                    f"客户 {customer_name} 正在使用中，无法删除！\n\n"
                    f"出库单：{outbound_count} 条\n\n"
                    "如需删除，请先删除相关出库单。")
                return

            if not messagebox.askyesno("确认删除", f"确定要删除客户 {customer_name} 吗？\n\n此操作不可恢复！"):
                return

            # 执行删除（软删除，设置为停用状态）
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("UPDATE customers SET status=0 WHERE name=?", (customer_name,))
            conn.commit()
            conn.close()

            messagebox.showinfo("成功", "客户已删除！")
            self.load_customers_data()  # 刷新列表

        except Exception as e:
            messagebox.showerror("错误", f"删除客户失败：{str(e)}")

    def quick_add_customer(self, customer_combo, customer_var, parent_win):
        """快捷新增客户（从出库单界面调用）"""
        add_win = tk.Toplevel(parent_win)
        add_win.title("快捷新增客户")
        add_win.geometry("450x480")
        add_win.transient(parent_win)
        add_win.grab_set()
        add_win.resizable(False, False)

        # 居中
        add_win.update_idletasks()
        x = (add_win.winfo_screenwidth() // 2) - 225
        y = (add_win.winfo_screenheight() // 2) - 240
        add_win.geometry(f'450x480+{x}+{y}')

        # 先创建按钮栏（固定在底部）
        btn_frame = tk.Frame(add_win, bg='#ECF0F1', height=70)
        btn_frame.pack(fill='x', side='bottom')
        btn_frame.pack_propagate(False)

        # 内容区域（在按钮栏上方）
        content_frame = tk.Frame(add_win, bg='white')
        content_frame.pack(fill='both', expand=True, padx=30, pady=15)

        # 客户名称
        tk.Label(content_frame, text="客户名称*：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        name_var = tk.StringVar()
        name_entry = tk.Entry(content_frame, textvariable=name_var,
                             font=('微软雅黑', 11), width=30)
        name_entry.pack(fill='x', pady=(0, 12))
        name_entry.focus()

        # 联系人
        tk.Label(content_frame, text="联系人：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        contact_person_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=contact_person_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 电话
        tk.Label(content_frame, text="电话：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        phone_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=phone_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 地址
        tk.Label(content_frame, text="地址：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        address_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=address_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 邮箱
        tk.Label(content_frame, text="邮箱：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        email_var = tk.StringVar()
        tk.Entry(content_frame, textvariable=email_var,
                font=('微软雅黑', 11), width=30).pack(fill='x', pady=(0, 12))

        # 按钮容器
        btn_container = tk.Frame(btn_frame, bg='#ECF0F1')
        btn_container.pack(expand=True, pady=15)

        def save_and_refresh():
            name = name_var.get().strip()
            if not name:
                messagebox.showerror("错误", "请输入客户名称！")
                name_entry.focus()
                return

            try:
                conn = sqlite3.connect(DB_FILE)
                cursor = conn.cursor()

                # 检查客户名称是否已存在
                cursor.execute("SELECT id FROM customers WHERE name=?", (name,))
                if cursor.fetchone():
                    messagebox.showerror("错误", "客户名称已存在！")
                    conn.close()
                    return

                # 插入新客户
                cursor.execute("INSERT INTO customers (name, contact_person, phone, address, email, status) VALUES (?, ?, ?, ?, ?, 1)",
                             (name, contact_person_var.get().strip(), phone_var.get().strip(), 
                              address_var.get().strip(), email_var.get().strip()))
                conn.commit()
                conn.close()

                # 刷新客户下拉框
                customers = self.get_customers_list()
                customer_combo['values'] = customers
                # 选中新添加的客户
                customer_var.set(name)
                
                messagebox.showinfo("成功", f"客户 {name} 添加成功！\n\n已自动选择该客户。")
                add_win.destroy()
                if hasattr(self, 'customers_tree'):
                    self.load_customers_data()  # 刷新列表（如果客户管理页面已打开）
            except sqlite3.IntegrityError:
                messagebox.showerror("错误", "客户名称已存在！")
            except Exception as e:
                messagebox.showerror("错误", f"添加客户失败：{str(e)}")
            finally:
                # 确保连接关闭（如果之前没有关闭）
                try:
                    if 'conn' in locals() and conn:
                        conn.close()
                except:
                    pass

        tk.Button(btn_container, text="保存",
                 font=('微软雅黑', 11),
                 bg='#2ECC71', fg='white',
                 pady=5, padx=20,
                 command=save_and_refresh).pack(side='left', padx=5)

        tk.Button(btn_container, text="取消",
                 font=('微软雅黑', 11),
                 bg='#95A5A6', fg='white',
                 pady=5, padx=20,
                 command=add_win.destroy).pack(side='left', padx=5)

    def show_reports(self, container=None):
        """显示统计报表页面"""
        if container is None:
            container = self.page_container if hasattr(self, 'page_container') else self.notebook.nametowidget(self.notebook.select())
        
        # 标题栏
        title_bar = tk.Frame(container, bg='white', height=60, relief='raised', bd=1)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)

        tk.Label(title_bar, text="统计报表",
                font=('微软雅黑', 18, 'bold'),
                bg='white').pack(side='left', padx=20, pady=15)

        # 报表类型选择
        report_frame = tk.Frame(container, bg='white', relief='raised', bd=1)
        report_frame.pack(fill='x', pady=10)

        # 报表类型
        tk.Label(report_frame, text="报表类型：",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=0, padx=20, pady=10)

        self.report_type = tk.StringVar(value='sales')
        report_combo = ttk.Combobox(report_frame, textvariable=self.report_type,
                                   font=('微软雅黑', 10), width=15, state='readonly')
        report_combo['values'] = ('销售报表', '采购报表', '库存报表', '利润报表')
        report_combo.grid(row=0, column=1, padx=5)

        # 时间范围
        tk.Label(report_frame, text="时间范围：",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=2, padx=(20, 5))

        self.report_start = tk.StringVar(value=(datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        report_start_date_frame = tk.Frame(report_frame, bg='white')
        report_start_date_frame.grid(row=0, column=3, padx=5)
        report_start_entry = tk.Entry(report_start_date_frame, textvariable=self.report_start,
                font=('微软雅黑', 10), width=12)
        report_start_entry.pack(side='left')
        report_start_entry.bind('<Button-1>', lambda e: self.show_date_picker(report_start_entry))
        # 日期选择按钮
        report_start_date_btn = tk.Button(report_start_date_frame, text="📅", font=('微软雅黑', 9),
                                           width=2, height=1, bg='#ECF0F1', relief='flat',
                                           cursor='hand2', command=lambda: self.show_date_picker(report_start_entry))
        report_start_date_btn.pack(side='left', padx=(2, 0))

        tk.Label(report_frame, text="至",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=4)

        self.report_end = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        report_end_date_frame = tk.Frame(report_frame, bg='white')
        report_end_date_frame.grid(row=0, column=5, padx=5)
        report_end_entry = tk.Entry(report_end_date_frame, textvariable=self.report_end,
                font=('微软雅黑', 10), width=12)
        report_end_entry.pack(side='left')
        report_end_entry.bind('<Button-1>', lambda e: self.show_date_picker(report_end_entry))
        # 日期选择按钮
        report_end_date_btn = tk.Button(report_end_date_frame, text="📅", font=('微软雅黑', 9),
                                         width=2, height=1, bg='#ECF0F1', relief='flat',
                                         cursor='hand2', command=lambda: self.show_date_picker(report_end_entry))
        report_end_date_btn.pack(side='left', padx=(2, 0))

        # 生成报表按钮
        generate_btn = tk.Button(report_frame, text="生成报表",
                               font=('微软雅黑', 10),
                               bg='#3498DB', fg='white',
                               pady=5, padx=20,
                               cursor='hand2',
                               command=self.generate_report)
        generate_btn.grid(row=0, column=6, padx=20)

        # 导出按钮
        export_btn = tk.Button(report_frame, text="导出Excel",
                             font=('微软雅黑', 10),
                             bg='#27AE60', fg='white',
                             pady=5, padx=20,
                             cursor='hand2',
                             command=self.export_report_to_excel)
        export_btn.grid(row=0, column=7)

        # 报表展示区
        report_display = tk.Frame(container, bg='white', relief='raised', bd=1)
        report_display.pack(fill='both', expand=True, pady=10)

        # 创建报表展示
        self.create_report_display(report_display)

    def create_report_display(self, parent):
        """创建报表展示区"""
        # 图表和表格容器
        container = tk.Frame(parent, bg='white')
        container.pack(fill='both', expand=True, padx=20, pady=20)

        # 创建报表表格框架
        table_frame = tk.Frame(container, bg='white')
        table_frame.pack(fill='both', expand=True)

        # 创建报表表格（列根据报表类型动态变化）
        columns = ('日期', '单号', '客户/供应商', '商品编码', '商品名称', '数量', '单价', '金额')
        self.report_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

        # 设置列标题和宽度
        column_widths = {
            '日期': 100,
            '单号': 150,
            '客户/供应商': 120,
            '商品编码': 100,
            '商品名称': 150,
            '数量': 80,
            '单价': 100,
            '金额': 100
        }
        for col in columns:
            self.report_tree.heading(col, text=col)
            self.report_tree.column(col, width=column_widths.get(col, 100))

        # 添加滚动条
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.report_tree.yview)
        self.report_tree.configure(yscrollcommand=scrollbar.set)

        self.report_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

    def show_settings(self, container=None):
        """显示系统设置页面"""
        if container is None:
            container = self.page_container if hasattr(self, 'page_container') else self.notebook.nametowidget(self.notebook.select())
        
        # 标题
        tk.Label(container, text="系统设置",
                font=('微软雅黑', 24, 'bold'),
                bg='#F0F2F5', fg='#2C3E50').pack(anchor='w', pady=(0, 20))

        # 设置选项卡
        notebook = ttk.Notebook(container)
        notebook.pack(fill='both', expand=True)

        # 基本设置
        basic_frame = tk.Frame(notebook, bg='white')
        notebook.add(basic_frame, text='基本设置')

        tk.Label(basic_frame, text="系统名称：",
                font=('微软雅黑', 10),
                bg='white').grid(row=0, column=0, padx=20, pady=10, sticky='e')

        # 加载保存的设置
        settings = self.load_settings()
        system_name = tk.StringVar(value=settings.get('system_name', '进销存管理系统'))
        tk.Entry(basic_frame, textvariable=system_name,
                font=('微软雅黑', 10), width=30).grid(row=0, column=1, padx=20, pady=10)

        tk.Label(basic_frame, text="公司名称：",
                font=('微软雅黑', 10),
                bg='white').grid(row=1, column=0, padx=20, pady=10, sticky='e')

        company_name = tk.StringVar(value=settings.get('company_name', '示例科技有限公司'))
        tk.Entry(basic_frame, textvariable=company_name,
                font=('微软雅黑', 10), width=30).grid(row=1, column=1, padx=20, pady=10)

        # 初始化密码设置
        tk.Label(basic_frame, text="初始化密码：",
                font=('微软雅黑', 10),
                bg='white').grid(row=2, column=0, padx=20, pady=10, sticky='e')

        init_password = tk.StringVar(value=self.get_init_password())
        password_entry = tk.Entry(basic_frame, textvariable=init_password,
                font=('微软雅黑', 10), width=30, show='*')
        password_entry.grid(row=2, column=1, padx=20, pady=10)

        # 保存按钮
        def save_settings():
            self.save_init_password(init_password.get())
            self.save_settings({
                'system_name': system_name.get(),
                'company_name': company_name.get()
            })
            messagebox.showinfo("成功", "设置已保存！")
        
        tk.Button(basic_frame, text="保存设置",
                 font=('微软雅黑', 11),
                 bg='#3498DB', fg='white',
                 pady=8, padx=30,
                 command=save_settings).grid(row=3, column=1, pady=30)

        # 用户管理
        user_frame = tk.Frame(notebook, bg='white')
        notebook.add(user_frame, text='用户管理')

        # 创建用户表格
        self.create_user_table(user_frame)
        
        # 权限设置
        permission_frame = tk.Frame(notebook, bg='white')
        notebook.add(permission_frame, text='权限设置')
        
        # 创建权限管理界面
        self.create_permission_settings(permission_frame)

        # 数据备份
        backup_frame = tk.Frame(notebook, bg='white')
        notebook.add(backup_frame, text='数据备份')

        tk.Label(backup_frame, text="数据备份与恢复",
                font=('微软雅黑', 14, 'bold'),
                bg='white').pack(pady=20)

        tk.Button(backup_frame, text="立即备份数据",
                 font=('微软雅黑', 11),
                 bg='#27AE60', fg='white',
                 pady=10, padx=30,
                 command=self.backup_database).pack(pady=20)

        tk.Button(backup_frame, text="恢复数据",
                 font=('微软雅黑', 11),
                 bg='#E67E22', fg='white',
                 pady=10, padx=30,
                 command=self.restore_database).pack(pady=10)

    def create_user_table(self, parent):
        """创建用户管理表格"""
        # 工具栏
        toolbar = tk.Frame(parent, bg='white')
        toolbar.pack(fill='x', padx=20, pady=10)

        tk.Button(toolbar, text="+ 添加用户",
                 font=('微软雅黑', 10),
                 bg='#2ECC71', fg='white',
                 pady=5, padx=15,
                 command=self.add_user).pack(side='left')
        
        tk.Button(toolbar, text="× 删除用户",
                 font=('微软雅黑', 10),
                 bg='#E74C3C', fg='white',
                 pady=5, padx=15,
                 command=self.delete_user).pack(side='left', padx=5)

        # 表格
        table_frame = tk.Frame(parent, bg='white')
        table_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        columns = ('用户名', '角色', '创建时间', '最后登录')
        self.user_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=10)

        for col in columns:
            self.user_tree.heading(col, text=col)
            self.user_tree.column(col, width=150)
        
        # 绑定双击事件
        self.user_tree.bind('<Double-Button-1>', self.edit_user)

        # 加载用户列表
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT username, role, created_time FROM users")
        users = cursor.fetchall()
        conn.close()
        
        for user in users:
            self.user_tree.insert('', 'end', values=(
                user[0],
                '管理员' if user[1] == 'admin' else '普通用户',
                user[2][:10] if user[2] else '-',
                '-'
            ))

        self.user_tree.pack(fill='both', expand=True)
    
    def load_user_table(self):
        """刷新用户列表"""
        if not hasattr(self, 'user_tree'):
            return
        
        # 清空现有数据
        for item in self.user_tree.get_children():
            self.user_tree.delete(item)
        
        # 加载用户列表
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT username, role, created_time FROM users")
        users = cursor.fetchall()
        conn.close()
        
        for user in users:
            self.user_tree.insert('', 'end', values=(
                user[0],
                '管理员' if user[1] == 'admin' else '普通用户',
                user[2][:10] if user[2] else '-',
                '-'
            ))
    
    def add_user(self):
        """添加用户"""
        add_win = tk.Toplevel(self.root)
        add_win.title("添加用户")
        add_win.geometry("400x280")
        add_win.transient(self.root)
        add_win.grab_set()
        add_win.resizable(False, False)
        
        # 居中
        add_win.update_idletasks()
        x = (add_win.winfo_screenwidth() // 2) - 200
        y = (add_win.winfo_screenheight() // 2) - 140
        add_win.geometry(f'400x320+{x}+{y}')
        
        # 先创建按钮栏（固定在底部）
        btn_frame = tk.Frame(add_win, bg='#ECF0F1', height=70)
        btn_frame.pack(fill='x', side='bottom')
        btn_frame.pack_propagate(False)
        
        # 内容区域（在按钮栏上方）
        content_frame = tk.Frame(add_win, bg='white')
        content_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # 用户名
        tk.Label(content_frame, text="用户名*：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        username_var = tk.StringVar()
        username_entry = tk.Entry(content_frame, textvariable=username_var,
                                 font=('微软雅黑', 11), width=30)
        username_entry.pack(fill='x', pady=(0, 15))
        username_entry.focus()
        
        # 密码
        tk.Label(content_frame, text="密码*：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        password_var = tk.StringVar()
        password_entry = tk.Entry(content_frame, textvariable=password_var,
                                 font=('微软雅黑', 11), width=30, show='*')
        password_entry.pack(fill='x', pady=(0, 15))
        
        # 确认密码
        tk.Label(content_frame, text="确认密码*：",
                font=('微软雅黑', 10),
                bg='white', anchor='w').pack(fill='x', pady=5)
        confirm_password_var = tk.StringVar()
        confirm_password_entry = tk.Entry(content_frame, textvariable=confirm_password_var,
                                         font=('微软雅黑', 11), width=30, show='*')
        confirm_password_entry.pack(fill='x', pady=(0, 15))
        
        # 按钮容器
        btn_container = tk.Frame(btn_frame, bg='#ECF0F1')
        btn_container.pack(expand=True, pady=15)
        
        def save_user():
            username = username_var.get().strip()
            password = password_var.get()
            confirm_password = confirm_password_var.get()
            
            if not username:
                messagebox.showerror("错误", "请输入用户名！")
                return
            
            if not password:
                messagebox.showerror("错误", "请输入密码！")
                return
            
            if password != confirm_password:
                messagebox.showerror("错误", "两次输入的密码不一致！")
                return
            
            if len(password) < 4:
                messagebox.showerror("错误", "密码长度不能少于4位！")
                return
            
            try:
                conn = sqlite3.connect(DB_FILE)
                cursor = conn.cursor()
                
                # 检查用户名是否已存在
                cursor.execute("SELECT id FROM users WHERE username=?", (username,))
                if cursor.fetchone():
                    messagebox.showerror("错误", "用户名已存在！")
                    conn.close()
                    return
                
                # 加密密码（使用新的PBKDF2哈希）
                password_hash = hash_password(password)
                
                # 插入新用户
                cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                             (username, password_hash, 'user'))
                conn.commit()
                conn.close()
                
                messagebox.showinfo("成功", f"用户 {username} 添加成功！")
                add_win.destroy()
                self.load_user_table()  # 刷新用户列表
                self.load_permission_users()  # 刷新权限设置页面的用户列表
                
            except Exception as e:
                messagebox.showerror("错误", f"添加用户失败：{str(e)}")
        
        tk.Button(btn_container, text="✓ 保存",
                 font=('微软雅黑', 11),
                 bg='#2ECC71', fg='white',
                 width=10, height=1,
                 command=save_user).pack(side='left', padx=5)
        
        tk.Button(btn_container, text="✗ 取消",
                 font=('微软雅黑', 11),
                 bg='#95A5A6', fg='white',
                 width=10, height=1,
                 command=add_win.destroy).pack(side='left', padx=5)
        
        # 绑定回车键
        username_entry.bind('<Return>', lambda e: password_entry.focus())
        password_entry.bind('<Return>', lambda e: confirm_password_entry.focus())
        confirm_password_entry.bind('<Return>', lambda e: save_user())
    
    def delete_user(self):
        """删除用户"""
        if not hasattr(self, 'user_tree'):
            messagebox.showwarning("提示", "请先打开用户管理页面！")
            return
            
        selection = self.user_tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先选择要删除的用户！")
            return
        
        item = self.user_tree.item(selection[0])
        username = item['values'][0]
        role = item['values'][1]
        
        if role == '管理员':
            messagebox.showerror("错误", "不能删除管理员用户！")
            return
        
        if username == self.user[1]:
            messagebox.showerror("错误", "不能删除当前登录用户！")
            return
        
        if not messagebox.askyesno("确认", f"确定要删除用户 {username} 吗？\n\n此操作不可恢复！"):
            return
        
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM users WHERE username=?", (username,))
            conn.commit()
            conn.close()
            
            messagebox.showinfo("成功", f"用户 {username} 已删除！")
            self.load_user_table()  # 刷新用户列表
            self.load_permission_users()  # 刷新权限设置页面的用户列表
            
        except Exception as e:
            messagebox.showerror("错误", f"删除用户失败：{str(e)}")
    
    def edit_user(self, event):
        """编辑用户（修改密码）"""
        if not hasattr(self, 'user_tree'):
            return
            
        selection = self.user_tree.selection()
        if not selection:
            return
        
        item = self.user_tree.item(selection[0])
        username = item['values'][0]
        
        edit_win = tk.Toplevel(self.root)
        edit_win.title(f"修改用户密码 - {username}")
        edit_win.geometry("400x250")
        edit_win.transient(self.root)
        edit_win.grab_set()
        edit_win.resizable(False, False)
        
        # 居中
        edit_win.update_idletasks()
        x = (edit_win.winfo_screenwidth() // 2) - 200
        y = (edit_win.winfo_screenheight() // 2) - 125
        edit_win.geometry(f'400x250+{x}+{y}')
        
        # 内容区域
        content_frame = tk.Frame(edit_win)
        content_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(content_frame, text=f"用户名：{username}",
                font=('微软雅黑', 11, 'bold')).pack(pady=(0, 20))
        
        # 新密码
        tk.Label(content_frame, text="新密码*：",
                font=('微软雅黑', 10),
                anchor='w').pack(fill='x', pady=5)
        password_var = tk.StringVar()
        password_entry = tk.Entry(content_frame, textvariable=password_var,
                                 font=('微软雅黑', 11), width=30, show='*')
        password_entry.pack(fill='x', pady=(0, 15))
        password_entry.focus()
        
        # 确认密码
        tk.Label(content_frame, text="确认密码*：",
                font=('微软雅黑', 10),
                anchor='w').pack(fill='x', pady=5)
        confirm_password_var = tk.StringVar()
        confirm_password_entry = tk.Entry(content_frame, textvariable=confirm_password_var,
                                         font=('微软雅黑', 11), width=30, show='*')
        confirm_password_entry.pack(fill='x', pady=(0, 15))
        
        # 按钮
        btn_frame = tk.Frame(edit_win, bg='#ECF0F1', height=65)
        btn_frame.pack(fill='x', side='bottom')
        btn_frame.pack_propagate(False)
        
        btn_container = tk.Frame(btn_frame, bg='#ECF0F1')
        btn_container.pack(expand=True, pady=12)
        
        def save_password():
            password = password_var.get()
            confirm_password = confirm_password_var.get()
            
            if not password:
                messagebox.showerror("错误", "请输入新密码！")
                return
            
            if password != confirm_password:
                messagebox.showerror("错误", "两次输入的密码不一致！")
                return
            
            if len(password) < 4:
                messagebox.showerror("错误", "密码长度不能少于4位！")
                return
            
            try:
                conn = sqlite3.connect(DB_FILE)
                cursor = conn.cursor()
                
                # 加密密码（使用新的PBKDF2哈希）
                password_hash = hash_password(password)
                
                # 更新密码
                cursor.execute("UPDATE users SET password=? WHERE username=?",
                             (password_hash, username))
                conn.commit()
                conn.close()
                
                messagebox.showinfo("成功", f"用户 {username} 的密码已修改！")
                edit_win.destroy()
                
            except Exception as e:
                messagebox.showerror("错误", f"修改密码失败：{str(e)}")
        
        tk.Button(btn_container, text="✓ 保存",
                 font=('微软雅黑', 11),
                 bg='#2ECC71', fg='white',
                 width=10, height=1,
                 command=save_password).pack(side='left', padx=5)
        
        tk.Button(btn_container, text="✗ 取消",
                 font=('微软雅黑', 11),
                 bg='#95A5A6', fg='white',
                 width=10, height=1,
                 command=edit_win.destroy).pack(side='left', padx=5)
        
        # 绑定回车键
        password_entry.bind('<Return>', lambda e: confirm_password_entry.focus())
        confirm_password_entry.bind('<Return>', lambda e: save_password())
    
    def load_settings(self):
        """加载系统设置"""
        try:
            if os.path.exists(SETTINGS_FILE):
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return {}
        except:
            return {}
    
    def save_settings(self, settings):
        """保存系统设置"""
        try:
            current_settings = self.load_settings()
            current_settings.update(settings)
            with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
                json.dump(current_settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("错误", f"保存设置失败：{str(e)}")
    
    def backup_database(self):
        """备份数据库"""
        try:
            import shutil
            from datetime import datetime
            
            # 确保备份目录存在
            os.makedirs(BACKUP_DIR, exist_ok=True)
            
            # 生成备份文件名（包含时间戳）
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = f'inventory_backup_{timestamp}.db'
            backup_path = os.path.join(BACKUP_DIR, backup_filename)
            
            # 复制数据库文件
            if os.path.exists(DB_FILE):
                shutil.copy2(DB_FILE, backup_path)
                messagebox.showinfo("成功", f"数据库备份成功！\n\n备份文件：{backup_filename}\n\n保存在 data/bak 目录下。")
            else:
                messagebox.showwarning("提示", "数据库文件不存在，无法备份。")
            
        except Exception as e:
            messagebox.showerror("错误", f"备份失败：{str(e)}")
    
    def restore_database(self):
        """恢复数据库"""
        try:
            import shutil
            from datetime import datetime
            
            # 确保备份目录存在
            os.makedirs(BACKUP_DIR, exist_ok=True)
            
            # 选择备份文件（默认从备份目录选择）
            backup_path = filedialog.askopenfilename(
                title="选择备份文件",
                filetypes=[("数据库文件", "*.db"), ("所有文件", "*.*")],
                initialdir=BACKUP_DIR
            )
            
            if not backup_path:
                return
            
            # 确认恢复
            if not messagebox.askyesno("确认", "确定要恢复数据库吗？\n\n当前数据将被覆盖，此操作不可恢复！"):
                return
            
            # 先备份当前数据库
            if os.path.exists(DB_FILE):
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                current_backup = os.path.join(BACKUP_DIR, f'inventory_before_restore_{timestamp}.db')
                shutil.copy2(DB_FILE, current_backup)
            
            # 恢复备份文件
            shutil.copy2(backup_path, DB_FILE)
            
            messagebox.showinfo("成功", "数据库恢复成功！\n\n请重新启动程序以使更改生效。")
            
        except Exception as e:
            messagebox.showerror("错误", f"恢复失败：{str(e)}")

    def create_permission_settings(self, parent):
        """创建权限设置界面"""
        # 标题
        tk.Label(parent, text="用户权限管理",
                font=('微软雅黑', 16, 'bold'),
                bg='white').pack(pady=20)
        
        # 说明
        tk.Label(parent, text="为不同用户设置可访问的功能模块",
                font=('微软雅黑', 10),
                bg='white', fg='#7F8C8D').pack(pady=(0, 20))
        
        # 左侧：用户列表
        list_frame = tk.Frame(parent, bg='white')
        list_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # 用户选择
        user_list_frame = tk.Frame(list_frame, bg='white')
        user_list_frame.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        tk.Label(user_list_frame, text="选择用户：",
                font=('微软雅黑', 11, 'bold'),
                bg='white').pack(anchor='w', pady=(0, 10))
        
        # 用户下拉框（保存为实例变量，以便后续刷新）
        self.permission_user_var = tk.StringVar()
        self.permission_user_combo = ttk.Combobox(user_list_frame, textvariable=self.permission_user_var,
                                                  font=('微软雅黑', 10), width=20, state='readonly')
        self.permission_user_combo.pack(fill='x', pady=(0, 10))
        
        # 加载用户列表
        self.load_permission_users()
        
        # 权限列表框架
        permission_frame = tk.Frame(list_frame, bg='white', relief='solid', bd=1)
        permission_frame.pack(side='left', fill='both', expand=True, padx=(10, 0))
        
        tk.Label(permission_frame, text="可访问的功能模块：",
                font=('微软雅黑', 11, 'bold'),
                bg='white').pack(anchor='w', padx=15, pady=15)
        
        # 权限选项（分组显示）
        permission_options = [
            ('dashboard', '📊 仪表盘'),
            ('inbound', '📥 入库管理'),
            ('outbound', '📤 出库管理'),
            ('inventory', '📦 库存查询'),
            ('reports', '📈 统计报表'),
            ('settings', '⚙️ 系统设置')
        ]
        
        # 基础管理子权限（分组显示）
        basic_management_options = [
            ('products', '📝 商品管理'),
            ('suppliers', '🏢 供应商管理'),
            ('customers', '👥 客户管理'),
            ('warehouses', '🏭 仓库管理')
        ]
        
        self.permission_vars = {}
        checkbox_frame = tk.Frame(permission_frame, bg='white')
        checkbox_frame.pack(fill='both', expand=True, padx=15, pady=(0, 15))
        
        # 显示主要权限选项
        for key, label in permission_options:
            var = tk.BooleanVar()
            self.permission_vars[key] = var
            cb = tk.Checkbutton(checkbox_frame, text=label,
                               font=('微软雅黑', 10),
                               bg='white', variable=var,
                               anchor='w')
            cb.pack(fill='x', pady=5)
        
        # 分隔线
        tk.Frame(checkbox_frame, bg='#BDC3C7', height=1).pack(fill='x', pady=10)
        
        # 基础管理标题
        tk.Label(checkbox_frame, text="基础管理（可多选）：",
                font=('微软雅黑', 10, 'bold'),
                bg='white', anchor='w').pack(fill='x', pady=(5, 10))
        
        # 显示基础管理子权限选项（缩进显示）
        for key, label in basic_management_options:
            var = tk.BooleanVar()
            self.permission_vars[key] = var
            cb = tk.Checkbutton(checkbox_frame, text=f"  └ {label}",
                               font=('微软雅黑', 10),
                               bg='white', variable=var,
                               anchor='w')
            cb.pack(fill='x', pady=3, padx=10)
        
        # 加载权限
        def load_permissions():
            username = self.permission_user_var.get()
            if username:
                permissions = self.get_user_permissions(username)
                # 更新复选框状态
                for key, var in self.permission_vars.items():
                    var.set(key in permissions)
        
        self.permission_user_combo.bind('<<ComboboxSelected>>', lambda e: load_permissions())
        if self.permission_user_var.get():
            load_permissions()
        
        # 保存按钮
        btn_frame = tk.Frame(parent, bg='white')
        btn_frame.pack(fill='x', padx=20, pady=20)
        
        def save_permissions():
            username = self.permission_user_var.get()
            if not username:
                messagebox.showwarning("提示", "请先选择用户！")
                return
            
            # 收集选中的权限
            permissions = [key for key, var in self.permission_vars.items() if var.get()]
            
            try:
                conn = sqlite3.connect(DB_FILE)
                cursor = conn.cursor()
                cursor.execute("UPDATE users SET permissions=? WHERE username=?",
                             (json.dumps(permissions), username))
                conn.commit()
                conn.close()
                
                messagebox.showinfo("成功", f"已成功保存用户 {username} 的权限设置！\n\n注意：用户需要重新登录才能生效。")
            except Exception as e:
                messagebox.showerror("错误", f"保存权限失败：{str(e)}")
        
        tk.Button(btn_frame, text="保存权限",
                 font=('微软雅黑', 11),
                 bg='#3498DB', fg='white',
                 pady=8, padx=30,
                 command=save_permissions).pack(side='left', padx=10)
        
        tk.Button(btn_frame, text="全选",
                 font=('微软雅黑', 11),
                 bg='#95A5A6', fg='white',
                 pady=8, padx=30,
                 command=lambda: [var.set(True) for var in self.permission_vars.values()]).pack(side='left', padx=10)
        
        tk.Button(btn_frame, text="全不选",
                 font=('微软雅黑', 11),
                 bg='#95A5A6', fg='white',
                 pady=8, padx=30,
                 command=lambda: [var.set(False) for var in self.permission_vars.values()]).pack(side='left', padx=10)
    
    def load_permission_users(self):
        """刷新权限设置页面的用户列表"""
        if not hasattr(self, 'permission_user_combo'):
            return
        
        try:
            # 加载用户列表
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("SELECT username, role FROM users WHERE role != 'admin' ORDER BY username")
            users = cursor.fetchall()
            conn.close()
            
            user_names = [user[0] for user in users]
            current_value = self.permission_user_var.get()
            
            # 更新下拉框的值
            self.permission_user_combo['values'] = user_names
            
            # 如果有用户，设置选中值
            if user_names:
                # 如果当前值还在列表中，保持选中；否则选择第一个
                if current_value in user_names:
                    self.permission_user_combo.set(current_value)
                    self.permission_user_var.set(current_value)
                else:
                    self.permission_user_combo.set(user_names[0])
                    self.permission_user_var.set(user_names[0])
            else:
                # 如果没有用户，清空选中值
                self.permission_user_var.set('')
        except Exception as e:
            print(f"加载权限用户列表时出错: {e}")

    def get_init_password(self):
        """获取初始化密码（默认123）"""
        try:
            if os.path.exists(SETTINGS_FILE):
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    return settings.get('init_password', '123')
            return '123'
        except:
            return '123'
    
    def save_init_password(self, password):
        """保存初始化密码"""
        try:
            settings = {}
            if os.path.exists(SETTINGS_FILE):
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
            settings['init_password'] = password
            with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("错误", f"保存密码失败：{str(e)}")
    
    def init_database_data(self):
        """初始化数据库数据（清空所有数据）"""
        # 创建密码输入对话框
        password_win = tk.Toplevel(self.root)
        password_win.title("初始化数据 - 密码验证")
        password_win.geometry("400x320")
        password_win.transient(self.root)
        password_win.grab_set()
        password_win.resizable(False, False)
        
        # 居中
        password_win.update_idletasks()
        x = (password_win.winfo_screenwidth() // 2) - 200
        y = (password_win.winfo_screenheight() // 2) - 160
        password_win.geometry(f'400x320+{x}+{y}')
        
        # 警告信息
        warning_frame = tk.Frame(password_win, bg='#FFF3CD', relief='solid', bd=1)
        warning_frame.pack(fill='x', padx=20, pady=(20, 10))
        
        tk.Label(warning_frame, text="⚠️ 警告：此操作将清空所有数据！",
                font=('微软雅黑', 12, 'bold'),
                bg='#FFF3CD', fg='#856404').pack(pady=15)
        
        tk.Label(warning_frame, text="包括：商品、入库单、出库单、库存等所有数据",
                font=('微软雅黑', 10),
                bg='#FFF3CD', fg='#856404').pack(pady=(0, 15))
        
        # 密码输入
        input_frame = tk.Frame(password_win, bg='white')
        input_frame.pack(fill='x', padx=20, pady=10)
        
        tk.Label(input_frame, text="请输入初始化密码：",
                font=('微软雅黑', 10),
                bg='white').pack(anchor='w', pady=(0, 5))
        
        password_var = tk.StringVar()
        password_entry = tk.Entry(input_frame, textvariable=password_var,
                                 font=('微软雅黑', 11), width=30, show='*')
        password_entry.pack(fill='x', pady=5)
        password_entry.focus()
        
        # 按钮栏（固定在底部）
        btn_frame = tk.Frame(password_win, bg='#ECF0F1', height=70)
        btn_frame.pack(fill='x', side='bottom')
        btn_frame.pack_propagate(False)
        
        btn_container = tk.Frame(btn_frame, bg='#ECF0F1')
        btn_container.pack(expand=True, pady=15)
        
        def verify_and_init():
            input_pwd = password_var.get()
            correct_pwd = self.get_init_password()
            
            if not input_pwd:
                messagebox.showerror("错误", "请输入初始化密码！")
                password_entry.focus()
                return
            
            if input_pwd != correct_pwd:
                messagebox.showerror("错误", "密码错误！")
                password_entry.delete(0, tk.END)
                password_entry.focus()
                return
            
            # 二次确认
            if not messagebox.askyesno("确认", "确定要清空所有数据吗？\n\n此操作不可恢复！"):
                return
            
            # 执行初始化
            try:
                conn = sqlite3.connect(DB_FILE)
                cursor = conn.cursor()
                
                # 清空所有表的数据（保留表结构）
                # 注意：先清空有关联关系的表，避免外键约束错误
                # 定义允许的表名白名单（防止SQL注入）
                ALLOWED_TABLES = [
                    'inventory',
                    'outbound_details',
                    'outbound_orders',
                    'inbound_details',
                    'inbound_orders',
                    'products',
                    'categories',
                    'warehouses',
                    'suppliers',
                    'customers'
                ]
                
                tables = [
                    'inventory',          # 库存表
                    'outbound_details',   # 出库明细
                    'outbound_orders',    # 出库单
                    'inbound_details',    # 入库明细
                    'inbound_orders',     # 入库单
                    'products',           # 商品表
                    'categories',         # 分类表
                    'warehouses',         # 仓库表
                    'suppliers',          # 供应商表
                    'customers'           # 客户表
                ]
                
                # 先禁用外键约束（SQLite默认是启用的，但为了保险起见）
                cursor.execute("PRAGMA foreign_keys = OFF")
                
                # 使用白名单验证表名，防止SQL注入
                for table in tables:
                    if table not in ALLOWED_TABLES:
                        raise ValueError(f"不允许操作表: {table}")
                    cursor.execute(f"DELETE FROM {table}")
                
                # 重置自增ID（可选，让ID从1开始）
                for table in tables:
                    if table not in ALLOWED_TABLES:
                        continue  # 跳过未验证的表名
                    try:
                        # 使用参数化查询（虽然表名不能参数化，但已通过白名单验证）
                        cursor.execute("DELETE FROM sqlite_sequence WHERE name=?", (table,))
                    except sqlite3.OperationalError:
                        # 如果表没有自增ID，忽略错误
                        pass
                    except Exception as e:
                        print(f"清理序列失败: {e}")
                
                # 重新启用外键约束
                cursor.execute("PRAGMA foreign_keys = ON")
                
                # 设置标记，防止重新插入示例数据
                import json
                settings = {}
                if os.path.exists(SETTINGS_FILE):
                    try:
                        with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                            settings = json.load(f)
                    except:
                        pass
                settings['skip_sample_data'] = True
                with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
                    json.dump(settings, f, ensure_ascii=False, indent=2)
                
                conn.commit()
                conn.close()
                
                messagebox.showinfo("成功", "数据初始化完成！\n\n所有数据已清空。")
                password_win.destroy()
                
                # 刷新库存列表
                if hasattr(self, 'inventory_tree'):
                    self.load_inventory_data()
                
                # 刷新仪表盘
                self.refresh_dashboard()
                
            except Exception as e:
                messagebox.showerror("错误", f"初始化失败：{str(e)}")
        
        # 提交按钮（确认初始化）
        tk.Button(btn_container, text="✓ 提交",
                 font=('微软雅黑', 11, 'bold'),
                 bg='#E74C3C', fg='white',
                 width=12, height=1,
                 pady=8, padx=20,
                 cursor='hand2',
                 command=verify_and_init).pack(side='left', padx=5)
        
        # 取消按钮
        tk.Button(btn_container, text="✗ 取消",
                 font=('微软雅黑', 11),
                 bg='#95A5A6', fg='white',
                 width=12, height=1,
                 pady=8, padx=20,
                 cursor='hand2',
                 command=password_win.destroy).pack(side='left', padx=5)
        
        # 绑定回车键
        password_entry.bind('<Return>', lambda e: verify_and_init())
    
    def update_time(self):
        """更新时间显示"""
        current_time = datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')
        self.time_label.config(text=current_time)
        self.root.after(1000, self.update_time)

    def logout(self):
        """退出系统"""
        # 询问是否备份数据
        backup_choice = messagebox.askyesnocancel(
            "退出确认",
            "退出前是否备份数据？\n\n"
            "是 - 备份数据后退出\n"
            "否 - 直接退出\n"
            "取消 - 返回"
        )
        
        if backup_choice is None:  # 取消
            return
        
        if backup_choice:  # 是 - 备份数据
            try:
                import shutil
                from datetime import datetime
                
                # 确保备份目录存在
                os.makedirs(BACKUP_DIR, exist_ok=True)
                
                if os.path.exists(DB_FILE):
                    # 生成备份文件名（包含时间戳）
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    backup_filename = f'inventory_backup_{timestamp}.db'
                    backup_path = os.path.join(BACKUP_DIR, backup_filename)
                    
                    # 复制数据库文件
                    shutil.copy2(DB_FILE, backup_path)
                    messagebox.showinfo("成功", f"数据备份成功！\n\n备份文件：{backup_filename}\n\n保存在 data/bak 目录下。")
                else:
                    messagebox.showwarning("提示", "数据库文件不存在，跳过备份。")
            except Exception as e:
                messagebox.showerror("错误", f"备份失败：{str(e)}\n\n是否仍要退出？")
                if not messagebox.askyesno("确认", "备份失败，是否仍要退出？"):
                    return
        
        # 确认退出
        if messagebox.askyesno("确认", "确定要退出系统吗？"):
            self.root.destroy()
            login = LoginWindow()
            login.run()

    def add_inbound_order(self):
        """新增入库单"""
        # 创建新窗口（增大窗口尺寸，确保所有内容可见）
        win = tk.Toplevel(self.root)
        win.title("新增入库单")
        win.geometry("900x700")
        win.minsize(800, 600)
        win.transient(self.root)
        win.grab_set()

        # 居中窗口
        win.update_idletasks()
        x = (win.winfo_screenwidth() // 2) - 450
        y = (win.winfo_screenheight() // 2) - 350
        win.geometry(f'900x700+{x}+{y}')

        # 基本信息区
        info_frame = tk.LabelFrame(win, text="基本信息", font=('微软雅黑', 12), padx=20, pady=20)
        info_frame.pack(fill='x', padx=20, pady=20)

        # 供应商
        tk.Label(info_frame, text="供应商：", font=('微软雅黑', 10)).grid(row=0, column=0, sticky='e', pady=5)
        supplier_var = tk.StringVar()
        supplier_frame = tk.Frame(info_frame)
        supplier_frame.grid(row=0, column=1, padx=10, pady=5, sticky='w')
        supplier_combo = ttk.Combobox(supplier_frame, textvariable=supplier_var, font=('微软雅黑', 10), width=25, state='normal')
        suppliers = self.get_suppliers_list()
        supplier_combo['values'] = suppliers
        supplier_combo.pack(side='left')
        # 默认选择第一个供应商（如果有）
        if suppliers:
            supplier_var.set(suppliers[0])
        # 快捷新增供应商按钮
        add_supplier_btn = tk.Button(supplier_frame, text="➕", font=('微软雅黑', 10),
                                     width=2, height=1, bg='#2ECC71', fg='white', relief='flat',
                                     cursor='hand2', 
                                     command=lambda: self.quick_add_supplier(supplier_combo, supplier_var, win))
        add_supplier_btn.pack(side='left', padx=(2, 0))

        # 仓库（复制供应商输入框的结构）
        tk.Label(info_frame, text="仓库：", font=('微软雅黑', 10)).grid(row=0, column=2, sticky='e', pady=5, padx=(20, 0))
        warehouse_var = tk.StringVar()
        warehouse_frame = tk.Frame(info_frame)
        warehouse_frame.grid(row=0, column=3, padx=10, pady=5, sticky='w')
        warehouse_combo = ttk.Combobox(warehouse_frame, textvariable=warehouse_var, font=('微软雅黑', 10), width=20, state='readonly')
        warehouses = self.get_warehouses_list()
        warehouse_combo['values'] = warehouses
        warehouse_combo.pack(side='left')
        # 默认选择第一个仓库（如果有）- 完全复制供应商的方式（只设置一次）
        if warehouses:
            warehouse_var.set(warehouses[0])
            # 延迟调用set()方法确保UI更新（调试代码中发现的关键：100ms后调用set()会触发UI刷新）
            def set_warehouse_display():
                warehouse_combo.set(warehouses[0])
            win.after(100, set_warehouse_display)
        # 添加快捷新增仓库按钮（占位，触发界面更新，与供应商按钮作用相同）
        add_warehouse_btn = tk.Frame(warehouse_frame, width=0, height=0)
        add_warehouse_btn.pack(side='left')

        # 入库日期
        tk.Label(info_frame, text="入库日期：", font=('微软雅黑', 10)).grid(row=1, column=0, sticky='e', pady=5)
        date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        date_frame = tk.Frame(info_frame)
        date_frame.grid(row=1, column=1, padx=10, pady=5, sticky='w')
        date_entry = tk.Entry(date_frame, textvariable=date_var, font=('微软雅黑', 10), width=28)
        date_entry.pack(side='left')
        date_entry.bind('<Button-1>', lambda e: self.show_date_picker(date_entry))
        # 日期选择按钮
        date_btn = tk.Button(date_frame, text="📅", font=('微软雅黑', 9),
                             width=2, height=1, bg='#ECF0F1', relief='flat',
                             cursor='hand2', command=lambda: self.show_date_picker(date_entry))
        date_btn.pack(side='left', padx=(2, 0))

        # 备注
        tk.Label(info_frame, text="备注：", font=('微软雅黑', 10)).grid(row=1, column=2, sticky='e', pady=5, padx=(20, 0))
        note_var = tk.StringVar()
        tk.Entry(info_frame, textvariable=note_var, font=('微软雅黑', 10), width=20).grid(row=1, column=3, padx=10, pady=5)

        # 商品明细区
        detail_frame = tk.LabelFrame(win, text="商品明细", font=('微软雅黑', 12), padx=20, pady=10)
        detail_frame.pack(fill='both', expand=True, padx=20, pady=10)

        # 工具栏（按钮区） - 放在表格上方
        toolbar_frame = tk.Frame(detail_frame)
        toolbar_frame.pack(fill='x', pady=(0, 10))

        tk.Button(toolbar_frame, text="+ 添加商品", font=('微软雅黑', 11, 'bold'),
                 command=lambda: self.add_product_to_inbound(tree, win), 
                 bg='#3498DB', fg='white', padx=20, pady=8, cursor='hand2').pack(side='left', padx=5)

        tk.Button(toolbar_frame, text="× 删除商品", font=('微软雅黑', 11, 'bold'),
                 command=lambda: self.delete_product_from_tree(tree, win), 
                 bg='#E74C3C', fg='white', padx=20, pady=8, cursor='hand2').pack(side='left', padx=5)

        # 提示文字
        hint_label = tk.Label(toolbar_frame, text="提示：点击「添加商品」按钮选择商品并输入数量和单价", 
                             font=('微软雅黑', 9), fg='#7F8C8D')
        hint_label.pack(side='left', padx=20)

        # 表格容器
        table_container = tk.Frame(detail_frame)
        table_container.pack(fill='both', expand=True)

        # 创建商品表格
        columns = ('商品编码', '商品名称', '规格型号', '数量', '单价', '金额', '批次号')
        tree = ttk.Treeview(table_container, columns=columns, show='headings', height=10)

        # 设置列
        column_widths = [100, 150, 150, 80, 80, 80, 100]
        for i, col in enumerate(columns):
            tree.heading(col, text=col)
            tree.column(col, width=column_widths[i])

        # 滚动条
        scrollbar = ttk.Scrollbar(table_container, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # 总金额（放在表格下方）
        total_frame = tk.Frame(detail_frame)
        total_frame.pack(fill='x', pady=(10, 0))
        
        total_label = tk.Label(total_frame, text="总金额：¥0.00", font=('微软雅黑', 14, 'bold'), fg='#E74C3C')
        total_label.pack(side='right')

        # 保存total_label到窗口对象，让子函数可以访问
        win.total_label = total_label

        # 底部按钮栏（固定在底部，始终可见）
        bottom_bar = tk.Frame(win, bg='#ECF0F1', height=70)
        bottom_bar.pack(fill='x', side='bottom')
        bottom_bar.pack_propagate(False)

        # 按钮容器
        btn_container = tk.Frame(bottom_bar, bg='#ECF0F1')
        btn_container.pack(side='right', padx=20, pady=15)

        tk.Button(btn_container, text="✓ 保存入库单", font=('微软雅黑', 12, 'bold'),
                 bg='#2ECC71', fg='white', width=18,
                 padx=20, pady=8, cursor='hand2',
                 command=lambda: self.save_inbound_order(tree, supplier_var.get(), warehouse_combo.get(),
                                                       date_var.get(), note_var.get(), total_label, win)).pack(side='left', padx=5)

        tk.Button(btn_container, text="✗ 取消", font=('微软雅黑', 12),
                 bg='#95A5A6', fg='white', width=18,
                 padx=20, pady=8, cursor='hand2',
                 command=win.destroy).pack(side='left', padx=5)

    def get_warehouses_list(self):
        """获取仓库列表"""
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM warehouses WHERE status=1")
        warehouses = [w[0] for w in cursor.fetchall()]
        conn.close()
        return warehouses

    def get_suppliers_list(self):
        """获取供应商列表"""
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM suppliers WHERE status=1 ORDER BY name")
        suppliers = [s[0] for s in cursor.fetchall()]
        conn.close()
        return suppliers

    def get_customers_list(self):
        """获取客户列表"""
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM customers WHERE status=1 ORDER BY name")
        customers = [c[0] for c in cursor.fetchall()]
        conn.close()
        return customers

    def add_product_to_inbound(self, tree, parent_win):
        """添加商品到入库单"""
        # 创建商品选择窗口（增大窗口尺寸）
        select_win = tk.Toplevel(self.root)
        select_win.title("选择商品 - 入库")
        select_win.geometry("750x550")
        select_win.transient(self.root)
        select_win.grab_set()

        # 居中窗口
        select_win.update_idletasks()
        x = (select_win.winfo_screenwidth() // 2) - 375
        y = (select_win.winfo_screenheight() // 2) - 275
        select_win.geometry(f'750x550+{x}+{y}')

        # 标题区域
        title_frame = tk.Frame(select_win, bg='#3498DB', height=40)
        title_frame.pack(fill='x')
        title_frame.pack_propagate(False)
        tk.Label(title_frame, text="请选择要入库的商品", font=('微软雅黑', 12, 'bold'),
                bg='#3498DB', fg='white').pack(expand=True)

        # 搜索框区域
        search_frame = tk.Frame(select_win, bg='white')
        search_frame.pack(fill='x', padx=20, pady=15)

        tk.Label(search_frame, text="搜索商品：", font=('微软雅黑', 11), bg='white').pack(side='left')
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var, font=('微软雅黑', 11), width=35)
        search_entry.pack(side='left', padx=10)
        search_entry.focus()

        tk.Label(search_frame, text="（输入商品编码或名称进行搜索）", 
                font=('微软雅黑', 9), bg='white', fg='#7F8C8D').pack(side='left')

        # 商品列表区域
        list_frame = tk.LabelFrame(select_win, text="商品列表", font=('微软雅黑', 11))
        list_frame.pack(fill='both', expand=True, padx=20, pady=(0, 10))

        columns = ('编码', '名称', '规格', '库存')
        product_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=18)

        # 设置列宽度
        product_tree.heading('编码', text='商品编码')
        product_tree.heading('名称', text='商品名称')
        product_tree.heading('规格', text='规格型号')
        product_tree.heading('库存', text='当前库存')
        
        product_tree.column('编码', width=120, anchor='center')
        product_tree.column('名称', width=200, anchor='w')
        product_tree.column('规格', width=200, anchor='w')
        product_tree.column('库存', width=100, anchor='center')

        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=product_tree.yview)
        product_tree.configure(yscrollcommand=scrollbar.set)

        product_tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        scrollbar.pack(side='right', fill='y', pady=10)

        # 加载商品数据
        def load_products():
            for item in product_tree.get_children():
                product_tree.delete(item)

            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()

            if search_var.get():
                cursor.execute('''
                    SELECT code, name, specification,
                           COALESCE(SUM(i.quantity), 0)
                    FROM products p
                    LEFT JOIN inventory i ON p.id = i.product_id
                    WHERE (p.code LIKE ? OR p.name LIKE ?) AND p.status = 1
                    GROUP BY p.id
                    ORDER BY p.code
                ''', (f'%{search_var.get()}%', f'%{search_var.get()}%'))
            else:
                cursor.execute('''
                    SELECT code, name, specification,
                           COALESCE(SUM(i.quantity), 0)
                    FROM products p
                    LEFT JOIN inventory i ON p.id = i.product_id
                    WHERE p.status = 1
                    GROUP BY p.id
                    ORDER BY p.code
                ''')

            for product in cursor.fetchall():
                product_tree.insert('', 'end', values=(
                    product[0], product[1], product[2] or '', product[3]
                ))

            conn.close()

        load_products()
        search_entry.bind('<KeyRelease>', lambda e: load_products())
        # 支持双击选择
        def on_double_click(event):
            confirm()

        product_tree.bind('<Double-1>', on_double_click)

        # 确认按钮区域 - 固定在底部，始终可见
        btn_frame_bottom = tk.Frame(select_win, bg='#ECF0F1', height=70)
        btn_frame_bottom.pack(fill='x', side='bottom')
        btn_frame_bottom.pack_propagate(False)

        # 提示文字
        hint_label = tk.Label(btn_frame_bottom, 
                             text="提示：双击商品行或选择商品后点击「确定选择」按钮",
                             font=('微软雅黑', 9), bg='#ECF0F1', fg='#7F8C8D')
        hint_label.pack(side='left', padx=20)

        # 按钮容器
        btn_container = tk.Frame(btn_frame_bottom, bg='#ECF0F1')
        btn_container.pack(side='right', padx=20, pady=15)

        # 确认按钮
        def confirm():
            selection = product_tree.selection()
            if not selection:
                messagebox.showwarning("提示", "请先选择商品！\n\n提示：点击商品行选中，或双击商品行快速选择")
                return
            
            item = product_tree.item(selection[0])
            product_data = item['values']

            # 弹出输入窗口（增大窗口尺寸，确保按钮完整显示）
            input_win = tk.Toplevel(select_win)
            input_win.title("输入数量和价格")
            input_win.geometry("420x320")
            input_win.transient(select_win)
            input_win.grab_set()
            input_win.resizable(False, False)

            # 居中
            input_win.update_idletasks()
            x = (input_win.winfo_screenwidth() // 2) - 210
            y = (input_win.winfo_screenheight() // 2) - 160
            input_win.geometry(f'420x320+{x}+{y}')

            # 内容区域
            content_frame = tk.Frame(input_win)
            content_frame.pack(fill='both', expand=True, padx=20, pady=15)

            # 商品信息
            info_frame = tk.Frame(content_frame, bg='#EBF5FB', relief='raised', bd=1)
            info_frame.pack(fill='x', pady=(0, 15))
            
            info_text = f"商品：{product_data[0]} - {product_data[1]}"
            tk.Label(info_frame, text=info_text, font=('微软雅黑', 11, 'bold'), 
                    fg='#2C3E50', bg='#EBF5FB').pack(pady=12)

            # 输入区域
            input_frame = tk.Frame(content_frame)
            input_frame.pack(fill='x', pady=5)

            # 数量
            qty_frame = tk.Frame(input_frame)
            qty_frame.pack(fill='x', pady=8)
            tk.Label(qty_frame, text="入库数量*：", font=('微软雅黑', 10), width=14, anchor='e').pack(side='left', padx=5)
            quantity_var = tk.StringVar()
            quantity_entry = tk.Entry(qty_frame, textvariable=quantity_var, font=('微软雅黑', 11), width=22)
            quantity_entry.pack(side='left', padx=5)
            quantity_entry.focus()

            # 单价
            price_frame = tk.Frame(input_frame)
            price_frame.pack(fill='x', pady=8)
            tk.Label(price_frame, text="单价*：", font=('微软雅黑', 10), width=14, anchor='e').pack(side='left', padx=5)
            price_var = tk.StringVar()
            tk.Entry(price_frame, textvariable=price_var, font=('微软雅黑', 11), width=22).pack(side='left', padx=5)

            # 批次号
            batch_frame = tk.Frame(input_frame)
            batch_frame.pack(fill='x', pady=8)
            tk.Label(batch_frame, text="批次号（可选）：", font=('微软雅黑', 10), width=14, anchor='e').pack(side='left', padx=5)
            batch_var = tk.StringVar()
            tk.Entry(batch_frame, textvariable=batch_var, font=('微软雅黑', 11), width=22).pack(side='left', padx=5)

            def add_to_tree():
                try:
                    if not quantity_var.get() or not price_var.get():
                        messagebox.showerror("错误", "请填写数量和单价")
                        return
                        
                    quantity = int(quantity_var.get())
                    if quantity <= 0:
                        raise ValueError("数量必须大于0")
                        
                    price = float(price_var.get())
                    if price < 0:
                        raise ValueError("单价不能为负数")
                        
                    amount = quantity * price

                    # 检查商品是否已添加
                    for child in tree.get_children():
                        item_data = tree.item(child)['values']
                        if item_data[0] == product_data[0]:
                            messagebox.showwarning("提示", "该商品已添加，请先删除再重新添加")
                            return

                    tree.insert('', 'end', values=(
                        product_data[0], product_data[1], product_data[2] or '',
                        quantity, f"{price:.2f}", f"{amount:.2f}", batch_var.get() or ''
                    ))

                    # 更新总金额 - 从父窗口获取total_label
                    if hasattr(parent_win, 'total_label'):
                        self.update_total_amount(tree, parent_win.total_label, is_inbound=True)

                    input_win.destroy()
                    select_win.destroy()
                except ValueError as e:
                    messagebox.showerror("错误", f"输入错误：{str(e)}")

            # 底部按钮栏（固定在底部，始终可见）
            bottom_bar = tk.Frame(input_win, bg='#ECF0F1', height=65)
            bottom_bar.pack(fill='x', side='bottom')
            bottom_bar.pack_propagate(False)

            btn_container = tk.Frame(bottom_bar, bg='#ECF0F1')
            btn_container.pack(expand=True, pady=12)

            tk.Button(btn_container, text="✓ 确定", command=add_to_tree,
                     bg='#2ECC71', fg='white', width=14, font=('微软雅黑', 11, 'bold'),
                     padx=20, pady=8, cursor='hand2').pack(side='left', padx=8)
            tk.Button(btn_container, text="✗ 取消", command=input_win.destroy,
                     bg='#95A5A6', fg='white', width=14, font=('微软雅黑', 11),
                     padx=20, pady=8, cursor='hand2').pack(side='left', padx=8)

        tk.Button(btn_container, text="✓ 确定选择", command=confirm,
                 bg='#2ECC71', fg='white', width=18, font=('微软雅黑', 11, 'bold'),
                 padx=20, pady=8, cursor='hand2').pack(side='left', padx=5)
        tk.Button(btn_container, text="✗ 取消", command=select_win.destroy,
                 bg='#95A5A6', fg='white', width=18, font=('微软雅黑', 11),
                 padx=20, pady=8, cursor='hand2').pack(side='left', padx=5)

    def delete_product_from_tree(self, tree, parent_win=None):
        """从表格中删除商品"""
        selection = tree.selection()
        if selection:
            tree.delete(selection[0])
            # 更新总金额
            if parent_win and hasattr(parent_win, 'total_label'):
                # 判断是入库还是出库（通过列数判断）
                columns = tree['columns']
                is_inbound = len(columns) == 7  # 入库单有7列（包含批次号）
                self.update_total_amount(tree, parent_win.total_label, is_inbound=is_inbound)
        else:
            messagebox.showwarning("提示", "请先选择要删除的商品")

    def update_total_amount(self, tree, total_label, is_inbound=True):
        """更新总金额
        Args:
            tree: 商品明细表格
            total_label: 总金额标签
            is_inbound: True为入库单（金额在第5列），False为出库单（金额在第6列）
        """
        total = 0.0
        for child in tree.get_children():
            item = tree.item(child)
            values = item['values']
            # 入库单：金额在第5列（索引5）；出库单：金额在第6列（索引6）
            amount_index = 5 if is_inbound else 6
            if len(values) > amount_index:
                amount_str = values[amount_index]
                try:
                    # 移除可能的千分位分隔符和¥符号
                    amount_str = amount_str.replace(',', '').replace('¥', '').strip()
                    total += float(amount_str)
                except (ValueError, IndexError):
                    pass
        total_label.config(text=f"总金额：¥{total:,.2f}")

    def save_inbound_order(self, tree, supplier, warehouse, date, note, total_label, win):
        """保存入库单"""
        # 检查是否有商品
        if not tree.get_children():
            messagebox.showerror("错误", "请添加商品")
            return

        # 生成入库单号
        order_no = f"IN{datetime.now().strftime('%Y%m%d%H%M%S')}"

        # 获取总金额
        total_text = total_label.cget("text")
        total_amount = float(total_text.replace("总金额：¥", "").replace(",", ""))

        # 保存到数据库
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        try:
            # 获取仓库ID
            cursor.execute("SELECT id FROM warehouses WHERE name=?", (warehouse,))
            warehouse_id = cursor.fetchone()[0]

            # 插入入库单主表
            cursor.execute('''
                INSERT INTO inbound_orders
                (order_no, warehouse_id, supplier, order_date, total_amount, operator, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (order_no, warehouse_id, supplier, date, total_amount, self.user[1], 1))

            order_id = cursor.lastrowid

            # 插入入库单明细
            for child in tree.get_children():
                item = tree.item(child)
                values = item['values']

                # 获取商品ID
                cursor.execute("SELECT id FROM products WHERE code=?", (values[0],))
                product_id = cursor.fetchone()[0]

                # 解析数量（确保转换为整数）
                try:
                    quantity = int(values[3]) if isinstance(values[3], (int, float)) else int(float(str(values[3])))
                except (ValueError, TypeError):
                    messagebox.showerror("错误", f"商品 {values[1]} 的数量格式错误：{values[3]}")
                    conn.rollback()
                    conn.close()
                    return

                # 解析单价和金额
                try:
                    price = float(values[4])
                    amount = float(str(values[5]).replace(',', '').replace('¥', ''))
                except (ValueError, TypeError) as e:
                    messagebox.showerror("错误", f"商品 {values[1]} 的单价或金额格式错误：{e}")
                    conn.rollback()
                    conn.close()
                    return

                # 插入明细
                cursor.execute('''
                    INSERT INTO inbound_details
                    (order_id, product_id, quantity, price, amount, batch_no)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (order_id, product_id, quantity, price, amount, values[6] if values[6] else None))

                # 更新库存（入库增加库存）
                batch_no = values[6] if values[6] else None
                
                # 检查是否已存在该商品的库存记录（使用更精确的匹配条件）
                if batch_no:
                    # 有批次号的情况
                    cursor.execute('''
                        SELECT id, quantity, available_quantity FROM inventory
                        WHERE product_id=? AND warehouse_id=? AND batch_no=?
                    ''', (product_id, warehouse_id, batch_no))
                else:
                    # 无批次号的情况，只匹配batch_no为NULL的记录
                    cursor.execute('''
                        SELECT id, quantity, available_quantity FROM inventory
                        WHERE product_id=? AND warehouse_id=? AND batch_no IS NULL
                    ''', (product_id, warehouse_id))
                
                existing = cursor.fetchone()
                if existing:
                    # 更新现有库存（使用id精确匹配，避免更新多条记录）
                    cursor.execute('''
                        UPDATE inventory
                        SET quantity = quantity + ?,
                            available_quantity = available_quantity + ?,
                            updated_time = CURRENT_TIMESTAMP
                        WHERE id=?
                    ''', (quantity, quantity, existing[0]))
                else:
                    # 插入新库存记录
                    cursor.execute('''
                        INSERT INTO inventory
                        (product_id, warehouse_id, quantity, available_quantity, batch_no, updated_time)
                        VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    ''', (product_id, warehouse_id, quantity, quantity, batch_no))

            conn.commit()
            messagebox.showinfo("成功", f"入库单保存成功！\n单号：{order_no}")
            win.destroy()
            self.load_inbound_data()  # 刷新列表

        except Exception as e:
            conn.rollback()
            messagebox.showerror("错误", f"保存失败：{str(e)}")
        finally:
            conn.close()

    def add_outbound_order(self):
        """新增出库单"""
        # 创建新窗口（增大窗口尺寸，确保所有内容可见）
        win = tk.Toplevel(self.root)
        win.title("新增出库单")
        win.geometry("900x700")
        win.minsize(800, 600)
        win.transient(self.root)
        win.grab_set()

        # 居中窗口
        win.update_idletasks()
        x = (win.winfo_screenwidth() // 2) - 450
        y = (win.winfo_screenheight() // 2) - 350
        win.geometry(f'900x700+{x}+{y}')

        # 基本信息区
        info_frame = tk.LabelFrame(win, text="基本信息", font=('微软雅黑', 12), padx=20, pady=20)
        info_frame.pack(fill='x', padx=20, pady=20)

        # 客户
        tk.Label(info_frame, text="客户：", font=('微软雅黑', 10)).grid(row=0, column=0, sticky='e', pady=5)
        customer_var = tk.StringVar()
        customer_frame = tk.Frame(info_frame)
        customer_frame.grid(row=0, column=1, padx=10, pady=5, sticky='w')
        customer_combo = ttk.Combobox(customer_frame, textvariable=customer_var, font=('微软雅黑', 10), width=25, state='normal')
        customers = self.get_customers_list()
        customer_combo['values'] = customers
        customer_combo.pack(side='left')
        # 默认选择第一个客户（如果有）
        if customers:
            customer_var.set(customers[0])
        # 快捷新增客户按钮
        add_customer_btn = tk.Button(customer_frame, text="➕", font=('微软雅黑', 10),
                                     width=2, height=1, bg='#2ECC71', fg='white', relief='flat',
                                     cursor='hand2',
                                     command=lambda: self.quick_add_customer(customer_combo, customer_var, win))
        add_customer_btn.pack(side='left', padx=(2, 0))

        # 仓库（复制供应商输入框的结构）
        tk.Label(info_frame, text="仓库：", font=('微软雅黑', 10)).grid(row=0, column=2, sticky='e', pady=5, padx=(20, 0))
        warehouse_var = tk.StringVar()
        warehouse_frame = tk.Frame(info_frame)
        warehouse_frame.grid(row=0, column=3, padx=10, pady=5, sticky='w')
        warehouse_combo = ttk.Combobox(warehouse_frame, textvariable=warehouse_var, font=('微软雅黑', 10), width=20, state='readonly')
        warehouses = self.get_warehouses_list()
        warehouse_combo['values'] = warehouses
        warehouse_combo.pack(side='left')
        # 默认选择第一个仓库（如果有）- 完全复制供应商的方式（只设置一次）
        if warehouses:
            warehouse_var.set(warehouses[0])
            # 延迟调用set()方法确保UI更新（调试代码中发现的关键：100ms后调用set()会触发UI刷新）
            def set_warehouse_display():
                warehouse_combo.set(warehouses[0])
            win.after(100, set_warehouse_display)
        # 添加快捷新增仓库按钮（占位，触发界面更新，与供应商按钮作用相同）
        add_warehouse_btn = tk.Frame(warehouse_frame, width=0, height=0)
        add_warehouse_btn.pack(side='left')

        # 出库日期
        tk.Label(info_frame, text="出库日期：", font=('微软雅黑', 10)).grid(row=1, column=0, sticky='e', pady=5)
        date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        date_frame = tk.Frame(info_frame)
        date_frame.grid(row=1, column=1, padx=10, pady=5, sticky='w')
        date_entry = tk.Entry(date_frame, textvariable=date_var, font=('微软雅黑', 10), width=28)
        date_entry.pack(side='left')
        date_entry.bind('<Button-1>', lambda e: self.show_date_picker(date_entry))
        # 日期选择按钮
        date_btn = tk.Button(date_frame, text="📅", font=('微软雅黑', 9),
                             width=2, height=1, bg='#ECF0F1', relief='flat',
                             cursor='hand2', command=lambda: self.show_date_picker(date_entry))
        date_btn.pack(side='left', padx=(2, 0))

        # 备注
        tk.Label(info_frame, text="备注：", font=('微软雅黑', 10)).grid(row=1, column=2, sticky='e', pady=5, padx=(20, 0))
        note_var = tk.StringVar()
        tk.Entry(info_frame, textvariable=note_var, font=('微软雅黑', 10), width=20).grid(row=1, column=3, padx=10, pady=5)

        # 中间内容区域容器
        content_frame = tk.Frame(win)
        content_frame.pack(fill='both', expand=True)
        
        # 商品明细区
        detail_frame = tk.LabelFrame(content_frame, text="商品明细", font=('微软雅黑', 12), padx=20, pady=10)
        detail_frame.pack(fill='both', expand=True, padx=20, pady=10)

        # 工具栏（按钮区） - 放在表格上方
        toolbar_frame = tk.Frame(detail_frame)
        toolbar_frame.pack(fill='x', pady=(0, 10))

        tk.Button(toolbar_frame, text="+ 添加商品", font=('微软雅黑', 11, 'bold'),
                 command=lambda: self.add_product_to_outbound(tree, warehouse_combo.get(), win),
                 bg='#3498DB', fg='white', padx=20, pady=8, cursor='hand2').pack(side='left', padx=5)

        tk.Button(toolbar_frame, text="× 删除商品", font=('微软雅黑', 11, 'bold'),
                 command=lambda: self.delete_product_from_tree(tree, win), 
                 bg='#E74C3C', fg='white', padx=20, pady=8, cursor='hand2').pack(side='left', padx=5)

        # 提示文字
        hint_label = tk.Label(toolbar_frame, text="提示：点击「添加商品」按钮选择商品并输入出库数量和单价", 
                             font=('微软雅黑', 9), fg='#7F8C8D')
        hint_label.pack(side='left', padx=20)

        # 表格容器
        table_container = tk.Frame(detail_frame)
        table_container.pack(fill='both', expand=True)

        # 创建商品表格
        columns = ('商品编码', '商品名称', '规格型号', '库存', '出库数量', '单价', '金额')
        tree = ttk.Treeview(table_container, columns=columns, show='headings', height=10)

        # 设置列
        column_widths = [100, 150, 150, 80, 80, 80, 80]
        for i, col in enumerate(columns):
            tree.heading(col, text=col)
            tree.column(col, width=column_widths[i])

        # 滚动条
        scrollbar = ttk.Scrollbar(table_container, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # 总金额（放在表格下方）
        total_frame = tk.Frame(detail_frame)
        total_frame.pack(fill='x', pady=(10, 0))
        
        total_label = tk.Label(total_frame, text="总金额：¥0.00", font=('微软雅黑', 14, 'bold'), fg='#E74C3C')
        total_label.pack(side='right')

        # 保存total_label到窗口对象
        win.total_label = total_label

        # 底部按钮栏（固定在底部，始终可见）
        bottom_bar = tk.Frame(win, bg='#ECF0F1', height=70)
        bottom_bar.pack(fill='x', side='bottom')
        bottom_bar.pack_propagate(False)

        # 按钮容器
        btn_container = tk.Frame(bottom_bar, bg='#ECF0F1')
        btn_container.pack(side='right', padx=20, pady=15)

        tk.Button(btn_container, text="✓ 保存出库单", font=('微软雅黑', 12, 'bold'),
                 bg='#E74C3C', fg='white', width=18,
                 padx=20, pady=8, cursor='hand2',
                 command=lambda: self.save_outbound_order(tree, customer_var.get(), warehouse_combo.get(),
                                                       date_var.get(), note_var.get(), total_label, win)).pack(side='left', padx=5)

        tk.Button(btn_container, text="✗ 取消", font=('微软雅黑', 12),
                 bg='#95A5A6', fg='white', width=18,
                 padx=20, pady=8, cursor='hand2',
                 command=win.destroy).pack(side='left', padx=5)

    def add_product_to_outbound(self, tree, warehouse, parent_win):
        """添加商品到出库单"""
        # 创建商品选择窗口（增大窗口尺寸）
        select_win = tk.Toplevel(self.root)
        select_win.title(f"选择商品 - 出库（仓库：{warehouse}）")
        select_win.geometry("750x550")
        select_win.transient(self.root)
        select_win.grab_set()

        # 居中窗口
        select_win.update_idletasks()
        x = (select_win.winfo_screenwidth() // 2) - 375
        y = (select_win.winfo_screenheight() // 2) - 275
        select_win.geometry(f'750x550+{x}+{y}')

        # 标题区域
        title_frame = tk.Frame(select_win, bg='#E74C3C', height=40)
        title_frame.pack(fill='x')
        title_frame.pack_propagate(False)
        tk.Label(title_frame, text=f"请选择要出库的商品（仓库：{warehouse}）", font=('微软雅黑', 12, 'bold'),
                bg='#E74C3C', fg='white').pack(expand=True)

        # 搜索框区域
        search_frame = tk.Frame(select_win, bg='white')
        search_frame.pack(fill='x', padx=20, pady=15)

        tk.Label(search_frame, text="搜索商品：", font=('微软雅黑', 11), bg='white').pack(side='left')
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var, font=('微软雅黑', 11), width=35)
        search_entry.pack(side='left', padx=10)
        search_entry.focus()

        tk.Label(search_frame, text="（输入商品编码或名称进行搜索，仅显示有库存的商品）", 
                font=('微软雅黑', 9), bg='white', fg='#7F8C8D').pack(side='left')

        # 商品列表区域
        list_frame = tk.LabelFrame(select_win, text="商品列表（仅显示有库存的商品）", font=('微软雅黑', 11))
        list_frame.pack(fill='both', expand=True, padx=20, pady=(0, 10))

        columns = ('编码', '名称', '规格', '库存')
        product_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=18)

        # 设置列宽度
        product_tree.heading('编码', text='商品编码')
        product_tree.heading('名称', text='商品名称')
        product_tree.heading('规格', text='规格型号')
        product_tree.heading('库存', text='当前库存')
        
        product_tree.column('编码', width=120, anchor='center')
        product_tree.column('名称', width=200, anchor='w')
        product_tree.column('规格', width=200, anchor='w')
        product_tree.column('库存', width=100, anchor='center')

        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=product_tree.yview)
        product_tree.configure(yscrollcommand=scrollbar.set)

        product_tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        scrollbar.pack(side='right', fill='y', pady=10)

        # 加载商品数据
        def load_products():
            for item in product_tree.get_children():
                product_tree.delete(item)

            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()

            # 先获取仓库ID
            cursor.execute("SELECT id FROM warehouses WHERE name=?", (warehouse,))
            warehouse_result = cursor.fetchone()
            if not warehouse_result:
                conn.close()
                messagebox.showerror("错误", f"仓库 '{warehouse}' 不存在")
                select_win.destroy()
                return
            warehouse_id = warehouse_result[0]

            # 查询指定仓库的商品库存
            if search_var.get():
                cursor.execute('''
                    SELECT p.code, p.name, p.specification, COALESCE(i.quantity, 0) as stock
                    FROM products p
                    LEFT JOIN inventory i ON p.id = i.product_id AND i.warehouse_id = ?
                    WHERE p.status = 1 AND (p.code LIKE ? OR p.name LIKE ?)
                    ORDER BY p.code
                ''', (warehouse_id, f'%{search_var.get()}%', f'%{search_var.get()}%'))
            else:
                cursor.execute('''
                    SELECT p.code, p.name, p.specification, COALESCE(i.quantity, 0) as stock
                    FROM products p
                    LEFT JOIN inventory i ON p.id = i.product_id AND i.warehouse_id = ?
                    WHERE p.status = 1
                    ORDER BY p.code
                ''', (warehouse_id,))

            for product in cursor.fetchall():
                if product[3] > 0:  # 只显示有库存的商品
                    product_tree.insert('', 'end', values=product)

            conn.close()

        load_products()
        search_entry.bind('<KeyRelease>', lambda e: load_products())
        # 支持双击选择
        def on_double_click(event):
            confirm()

        product_tree.bind('<Double-1>', on_double_click)

        # 确认按钮区域 - 固定在底部，始终可见
        btn_frame_bottom = tk.Frame(select_win, bg='#ECF0F1', height=70)
        btn_frame_bottom.pack(fill='x', side='bottom')
        btn_frame_bottom.pack_propagate(False)

        # 提示文字
        hint_label = tk.Label(btn_frame_bottom, 
                             text="提示：双击商品行或选择商品后点击「确定选择」按钮",
                             font=('微软雅黑', 9), bg='#ECF0F1', fg='#7F8C8D')
        hint_label.pack(side='left', padx=20)

        # 按钮容器
        btn_container = tk.Frame(btn_frame_bottom, bg='#ECF0F1')
        btn_container.pack(side='right', padx=20, pady=15)

        # 确认按钮
        def confirm():
            selection = product_tree.selection()
            if not selection:
                messagebox.showwarning("提示", "请先选择商品！\n\n提示：点击商品行选中，或双击商品行快速选择")
                return
            
            item = product_tree.item(selection[0])
            product_data = item['values']

            # 弹出输入窗口（增大窗口尺寸，确保按钮完整显示）
            input_win = tk.Toplevel(select_win)
            input_win.title("输入数量和价格")
            input_win.geometry("420x290")
            input_win.transient(select_win)
            input_win.grab_set()
            input_win.resizable(False, False)

            # 居中
            input_win.update_idletasks()
            x = (input_win.winfo_screenwidth() // 2) - 210
            y = (input_win.winfo_screenheight() // 2) - 145
            input_win.geometry(f'420x290+{x}+{y}')

            # 内容区域
            content_frame = tk.Frame(input_win)
            content_frame.pack(fill='both', expand=True, padx=20, pady=15)

            # 商品信息
            info_frame = tk.Frame(content_frame, bg='#FADBD8', relief='raised', bd=1)
            info_frame.pack(fill='x', pady=(0, 15))
            
            info_text = f"商品：{product_data[0]} - {product_data[1]}\n当前库存：{product_data[3]}"
            tk.Label(info_frame, text=info_text, font=('微软雅黑', 11, 'bold'), 
                    fg='#2C3E50', bg='#FADBD8').pack(pady=12)

            # 输入区域
            input_frame = tk.Frame(content_frame)
            input_frame.pack(fill='x', pady=5)

            # 数量
            qty_frame = tk.Frame(input_frame)
            qty_frame.pack(fill='x', pady=8)
            tk.Label(qty_frame, text="出库数量*：", font=('微软雅黑', 10), width=14, anchor='e').pack(side='left', padx=5)
            quantity_var = tk.StringVar()
            quantity_entry = tk.Entry(qty_frame, textvariable=quantity_var, font=('微软雅黑', 11), width=22)
            quantity_entry.pack(side='left', padx=5)
            quantity_entry.focus()

            # 单价
            price_frame = tk.Frame(input_frame)
            price_frame.pack(fill='x', pady=8)
            tk.Label(price_frame, text="单价*：", font=('微软雅黑', 10), width=14, anchor='e').pack(side='left', padx=5)
            price_var = tk.StringVar()
            tk.Entry(price_frame, textvariable=price_var, font=('微软雅黑', 11), width=22).pack(side='left', padx=5)

            def add_to_tree():
                try:
                    if not quantity_var.get() or not price_var.get():
                        messagebox.showerror("错误", "请填写数量和单价")
                        return
                        
                    quantity = int(quantity_var.get())
                    if quantity <= 0:
                        raise ValueError("数量必须大于0")
                    if quantity > product_data[3]:
                        messagebox.showerror("错误", f"出库数量不能大于库存数量（当前库存：{product_data[3]}）")
                        return

                    price = float(price_var.get())
                    if price < 0:
                        raise ValueError("单价不能为负数")
                        
                    amount = quantity * price

                    # 检查商品是否已添加
                    for child in tree.get_children():
                        item_data = tree.item(child)['values']
                        if item_data[0] == product_data[0]:
                            messagebox.showwarning("提示", "该商品已添加，请先删除再重新添加")
                            return

                    tree.insert('', 'end', values=(
                        product_data[0], product_data[1], product_data[2] or '',
                        product_data[3], quantity, f"{price:.2f}", f"{amount:.2f}"
                    ))

                    # 更新总金额 - 从父窗口获取total_label
                    if hasattr(parent_win, 'total_label'):
                        self.update_total_amount(tree, parent_win.total_label, is_inbound=False)

                    input_win.destroy()
                    select_win.destroy()
                except ValueError as e:
                    messagebox.showerror("错误", f"输入错误：{str(e)}")

            # 底部按钮栏（固定在底部，始终可见）
            bottom_bar = tk.Frame(input_win, bg='#ECF0F1', height=65)
            bottom_bar.pack(fill='x', side='bottom')
            bottom_bar.pack_propagate(False)

            btn_container = tk.Frame(bottom_bar, bg='#ECF0F1')
            btn_container.pack(expand=True, pady=12)

            tk.Button(btn_container, text="✓ 确定", command=add_to_tree,
                     bg='#2ECC71', fg='white', width=14, font=('微软雅黑', 11, 'bold'),
                     padx=20, pady=8, cursor='hand2').pack(side='left', padx=8)
            tk.Button(btn_container, text="✗ 取消", command=input_win.destroy,
                     bg='#95A5A6', fg='white', width=14, font=('微软雅黑', 11),
                     padx=20, pady=8, cursor='hand2').pack(side='left', padx=8)

        tk.Button(btn_container, text="✓ 确定选择", command=confirm,
                 bg='#2ECC71', fg='white', width=18, font=('微软雅黑', 11, 'bold'),
                 padx=20, pady=8, cursor='hand2').pack(side='left', padx=5)
        tk.Button(btn_container, text="✗ 取消", command=select_win.destroy,
                 bg='#95A5A6', fg='white', width=18, font=('微软雅黑', 11),
                 padx=20, pady=8, cursor='hand2').pack(side='left', padx=5)

    def save_outbound_order(self, tree, customer, warehouse, date, note, total_label, win):
        """保存出库单"""
        # 检查是否有商品
        if not tree.get_children():
            messagebox.showerror("错误", "请添加商品")
            return

        # 生成出库单号
        order_no = f"OUT{datetime.now().strftime('%Y%m%d%H%M%S')}"

        # 获取总金额
        total_text = total_label.cget("text")
        total_amount = float(total_text.replace("总金额：¥", "").replace(",", ""))

        # 保存到数据库
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        try:
            # 获取仓库ID
            cursor.execute("SELECT id FROM warehouses WHERE name=?", (warehouse,))
            warehouse_id = cursor.fetchone()[0]

            # 插入出库单主表
            cursor.execute('''
                INSERT INTO outbound_orders
                (order_no, warehouse_id, customer, order_date, total_amount, operator, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (order_no, warehouse_id, customer, date, total_amount, self.user[1], 1))

            order_id = cursor.lastrowid

            # 插入出库单明细
            for child in tree.get_children():
                item = tree.item(child)
                values = item['values']

                # 获取商品ID
                cursor.execute("SELECT id FROM products WHERE code=?", (values[0],))
                product_id = cursor.fetchone()[0]

                # 插入明细
                out_quantity = int(values[4])
                cursor.execute('''
                    INSERT INTO outbound_details
                    (order_id, product_id, quantity, price, amount)
                    VALUES (?, ?, ?, ?, ?)
                ''', (order_id, product_id, out_quantity, float(values[5]), float(values[6])))

                # 检查库存是否足够
                cursor.execute('''
                    SELECT quantity, available_quantity FROM inventory
                    WHERE product_id = ? AND warehouse_id = ?
                ''', (product_id, warehouse_id))
                
                stock_result = cursor.fetchone()
                if not stock_result:
                    raise Exception(f"商品 {values[0]} 在仓库 {warehouse} 中没有库存")
                
                current_stock, available_stock = stock_result
                if available_stock < out_quantity:
                    raise Exception(f"商品 {values[0]} 可用库存不足（当前可用：{available_stock}，需要：{out_quantity}）")

                # 更新库存（出库减少库存）
                cursor.execute('''
                    UPDATE inventory
                    SET quantity = quantity - ?,
                        available_quantity = available_quantity - ?,
                        updated_time = CURRENT_TIMESTAMP
                    WHERE product_id = ? AND warehouse_id = ?
                ''', (out_quantity, out_quantity, product_id, warehouse_id))

            conn.commit()
            messagebox.showinfo("成功", f"出库单保存成功！\n单号：{order_no}")
            win.destroy()
            self.load_outbound_data()  # 刷新列表

        except Exception as e:
            conn.rollback()
            messagebox.showerror("错误", f"保存失败：{str(e)}")
        finally:
            conn.close()

    def export_inventory(self):
        """导出库存数据"""
        from tkinter import filedialog
        import csv

        # 选择保存位置
        filename = filedialog.asksaveasfilename(
            title="导出库存数据",
            defaultextension=".csv",
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )

        if filename:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()

            cursor.execute('''
                SELECT p.code, p.name, p.specification, w.name,
                       i.quantity, i.available_quantity, p.unit, p.min_stock
                FROM inventory i
                JOIN products p ON i.product_id = p.id
                JOIN warehouses w ON i.warehouse_id = w.id
                WHERE p.status = 1
                ORDER BY w.name, p.code
            ''')

            with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['商品编码', '商品名称', '规格型号', '仓库', '库存数量',
                               '可用库存', '单位', '最小库存'])
                writer.writerows(cursor.fetchall())

            conn.close()
            messagebox.showinfo("成功", f"库存数据已导出到：\n{filename}")

    def add_product(self):
        """新增商品"""
        win = tk.Toplevel(self.root)
        win.title("新增商品")
        win.geometry("550x650")
        win.minsize(500, 600)
        win.transient(self.root)
        win.grab_set()

        # 居中
        win.update_idletasks()
        x = (win.winfo_screenwidth() // 2) - 275
        y = (win.winfo_screenheight() // 2) - 325
        win.geometry(f'550x650+{x}+{y}')

        # 内容区域
        content_frame = tk.Frame(win)
        content_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # 表单
        form_frame = tk.LabelFrame(content_frame, text="商品信息", font=('微软雅黑', 11), padx=15, pady=15)
        form_frame.pack(fill='both', expand=True)

        # 商品编码
        tk.Label(form_frame, text="商品编码*", font=('微软雅黑', 10)).grid(row=0, column=0, sticky='e', pady=5)
        code_var = tk.StringVar()
        tk.Entry(form_frame, textvariable=code_var, font=('微软雅黑', 10), width=30).grid(row=0, column=1, padx=10, pady=5)

        # 商品名称
        tk.Label(form_frame, text="商品名称*", font=('微软雅黑', 10)).grid(row=1, column=0, sticky='e', pady=5)
        name_var = tk.StringVar()
        tk.Entry(form_frame, textvariable=name_var, font=('微软雅黑', 10), width=30).grid(row=1, column=1, padx=10, pady=5)

        # 规格型号
        tk.Label(form_frame, text="规格型号", font=('微软雅黑', 10)).grid(row=2, column=0, sticky='e', pady=5)
        spec_var = tk.StringVar()
        tk.Entry(form_frame, textvariable=spec_var, font=('微软雅黑', 10), width=30).grid(row=2, column=1, padx=10, pady=5)

        # 单位
        tk.Label(form_frame, text="单位", font=('微软雅黑', 10)).grid(row=3, column=0, sticky='e', pady=5)
        unit_var = tk.StringVar(value="个")
        unit_combo = ttk.Combobox(form_frame, textvariable=unit_var, font=('微软雅黑', 10), width=28)
        unit_combo['values'] = ('个', '件', '箱', '套', '台', '张', '包', '瓶', '斤', '公斤', '米', '升')
        unit_combo.grid(row=3, column=1, padx=10, pady=5)

        # 采购价
        tk.Label(form_frame, text="采购价", font=('微软雅黑', 10)).grid(row=4, column=0, sticky='e', pady=5)
        purchase_var = tk.StringVar()
        tk.Entry(form_frame, textvariable=purchase_var, font=('微软雅黑', 10), width=30).grid(row=4, column=1, padx=10, pady=5)

        # 销售价
        tk.Label(form_frame, text="销售价", font=('微软雅黑', 10)).grid(row=5, column=0, sticky='e', pady=5)
        sale_var = tk.StringVar()
        tk.Entry(form_frame, textvariable=sale_var, font=('微软雅黑', 10), width=30).grid(row=5, column=1, padx=10, pady=5)

        # 最小库存
        tk.Label(form_frame, text="最小库存", font=('微软雅黑', 10)).grid(row=6, column=0, sticky='e', pady=5)
        min_stock_var = tk.StringVar(value="0")
        tk.Entry(form_frame, textvariable=min_stock_var, font=('微软雅黑', 10), width=30).grid(row=6, column=1, padx=10, pady=5)

        # 最大库存
        tk.Label(form_frame, text="最大库存", font=('微软雅黑', 10)).grid(row=7, column=0, sticky='e', pady=5)
        max_stock_var = tk.StringVar(value="999999")
        tk.Entry(form_frame, textvariable=max_stock_var, font=('微软雅黑', 10), width=30).grid(row=7, column=1, padx=10, pady=5)

        def save():
            if not code_var.get() or not name_var.get():
                messagebox.showerror("错误", "请填写必填项（商品编码和商品名称）")
                return

            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()

            try:
                cursor.execute('''
                    INSERT INTO products
                    (code, name, specification, unit, purchase_price, sale_price, min_stock, max_stock)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (code_var.get(), name_var.get(), spec_var.get(), unit_var.get(),
                      float(purchase_var.get() or 0), float(sale_var.get() or 0),
                      int(min_stock_var.get() or 0), int(max_stock_var.get() or 999999)))

                conn.commit()
                messagebox.showinfo("成功", "商品保存成功")
                win.destroy()
                self.load_products_data()  # 刷新列表

            except sqlite3.IntegrityError:
                messagebox.showerror("错误", "商品编码已存在，请使用其他编码")
            except Exception as e:
                messagebox.showerror("错误", f"保存失败：{str(e)}")
            finally:
                conn.close()

        # 底部按钮栏（固定在底部，始终可见）
        bottom_bar = tk.Frame(win, bg='#ECF0F1', height=70)
        bottom_bar.pack(fill='x', side='bottom')
        bottom_bar.pack_propagate(False)

        btn_container = tk.Frame(bottom_bar, bg='#ECF0F1')
        btn_container.pack(side='right', padx=20, pady=15)

        tk.Button(btn_container, text="✓ 保存", font=('微软雅黑', 12, 'bold'),
                 bg='#2ECC71', fg='white', width=18,
                 padx=20, pady=8, cursor='hand2',
                 command=save).pack(side='left', padx=5)

        tk.Button(btn_container, text="✗ 取消", font=('微软雅黑', 12),
                 bg='#95A5A6', fg='white', width=18,
                 padx=20, pady=8, cursor='hand2',
                 command=win.destroy).pack(side='left', padx=5)

    def search_inbound(self):
        """搜索入库单（展开明细，每个商品一行）"""
        start_date = self.start_date.get()
        end_date = self.end_date.get()
        order_no = self.order_no.get()

        # 清空现有数据
        for item in self.inbound_tree.get_children():
            self.inbound_tree.delete(item)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        # 构建查询条件
        conditions = []
        params = []

        if start_date:
            conditions.append("io.order_date >= ?")
            params.append(start_date)

        if end_date:
            conditions.append("io.order_date <= ?")
            params.append(end_date)

        if order_no:
            conditions.append("io.order_no LIKE ?")
            params.append(f"%{order_no}%")

        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""

        # 查询入库单明细，每个商品一行（与load_inbound_data保持一致）
        cursor.execute(f'''
            SELECT io.order_no, io.order_date, io.supplier, w.name,
                   p.code, p.name, p.specification,
                   id.quantity, id.price, id.amount, id.batch_no,
                   io.status, io.operator
            FROM inbound_orders io
            LEFT JOIN warehouses w ON io.warehouse_id = w.id
            LEFT JOIN inbound_details id ON io.id = id.order_id
            LEFT JOIN products p ON id.product_id = p.id
            {where_clause}
            ORDER BY io.created_time DESC, io.id, id.id
        ''', params)

        records = cursor.fetchall()
        conn.close()

        for record in records:
            status_text = '已入库' if record[11] == 1 else '待入库'
            self.inbound_tree.insert('', 'end', values=(
                record[0],  # 单号
                record[1],  # 日期
                record[2] or '',  # 供应商
                record[3] or '',  # 仓库
                record[4] or '',  # 商品编码
                record[5] or '',  # 商品名称
                record[6] or '',  # 规格型号
                record[7] or 0,  # 数量
                f"{float(record[8] or 0):.2f}",  # 单价
                f"¥{float(record[9] or 0):.2f}",  # 金额
                record[10] or '',  # 批次号
                status_text,  # 状态
                record[12] or ''  # 操作人
            ))

    def search_outbound(self):
        """搜索出库单（展开明细，每个商品一行）"""
        start_date = self.out_start_date.get()
        end_date = self.out_end_date.get()
        customer = self.customer.get()

        # 清空现有数据
        for item in self.outbound_tree.get_children():
            self.outbound_tree.delete(item)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        # 构建查询条件
        conditions = []
        params = []

        if start_date:
            conditions.append("oo.order_date >= ?")
            params.append(start_date)

        if end_date:
            conditions.append("oo.order_date <= ?")
            params.append(end_date)

        if customer:
            conditions.append("oo.customer LIKE ?")
            params.append(f"%{customer}%")

        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""

        # 查询出库单明细，每个商品一行（与load_outbound_data保持一致）
        cursor.execute(f'''
            SELECT oo.order_no, oo.order_date, oo.customer, w.name,
                   p.code, p.name, p.specification,
                   od.quantity, od.price, od.amount,
                   oo.status, oo.operator
            FROM outbound_orders oo
            LEFT JOIN warehouses w ON oo.warehouse_id = w.id
            LEFT JOIN outbound_details od ON oo.id = od.order_id
            LEFT JOIN products p ON od.product_id = p.id
            {where_clause}
            ORDER BY oo.created_time DESC, oo.id, od.id
        ''', params)

        records = cursor.fetchall()
        conn.close()

        for record in records:
            status_text = '已出库' if record[10] == 1 else '待出库'
            self.outbound_tree.insert('', 'end', values=(
                record[0],  # 单号
                record[1],  # 日期
                record[2] or '',  # 客户
                record[3] or '',  # 仓库
                record[4] or '',  # 商品编码
                record[5] or '',  # 商品名称
                record[6] or '',  # 规格型号
                record[7] or 0,  # 出库数量
                f"{float(record[8] or 0):.2f}",  # 单价
                f"¥{float(record[9] or 0):.2f}",  # 金额
                status_text,  # 状态
                record[11] or ''  # 操作人
            ))

    def search_inventory(self):
        """搜索库存"""
        product = self.product_search.get()
        spec = self.spec_search.get()
        warehouse = self.warehouse_var.get()
        status = self.stock_status.get()

        # 清空现有数据
        for item in self.inventory_tree.get_children():
            self.inventory_tree.delete(item)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        # 构建查询条件
        conditions = ["p.status = 1"]
        params = []

        if product:
            conditions.append("(p.code LIKE ? OR p.name LIKE ?)")
            params.append(f"%{product}%")
            params.append(f"%{product}%")

        if spec:
            conditions.append("p.specification LIKE ?")
            params.append(f"%{spec}%")

        if warehouse != '全部仓库':
            conditions.append("w.name = ?")
            params.append(warehouse)

        where_clause = "WHERE " + " AND ".join(conditions)

        cursor.execute(f'''
            SELECT i.id, p.code, p.name, p.specification, w.name,
                   i.quantity, i.available_quantity, p.unit, p.min_stock
            FROM inventory i
            JOIN products p ON i.product_id = p.id
            JOIN warehouses w ON i.warehouse_id = w.id
            {where_clause}
            ORDER BY i.updated_time DESC
        ''', params)

        records = cursor.fetchall()
        
        for record in records:
            # 判断库存状态（quantity在索引5，min_stock在索引8）
            quantity = record[5]
            min_stock = record[8]
            
            if quantity == 0:
                status = '无库存'
                status_color = '#95A5A6'
            elif min_stock and quantity < min_stock:
                status = '库存不足'
                status_color = '#E74C3C'
            elif min_stock and quantity > min_stock * 2:
                status = '积压'
                status_color = '#F39C12'
            else:
                status = '正常'
                status_color = '#27AE60'
            
            # record[0]是i.id，record[1-4]是code,name,specification,warehouse
            # record[5-6]是quantity,available_quantity，record[7]是unit，record[8]是min_stock
            self.inventory_tree.insert('', 'end', values=(
                record[1], record[2], record[3] or '', record[4],
                record[5], record[6], record[7], record[8] or 0, status
            ), tags=(status_color, f"inv_id_{record[0]}"))
        
        # 设置行颜色
        self.inventory_tree.tag_configure('#E74C3C', background='#FADBD8')
        self.inventory_tree.tag_configure('#F39C12', background='#FCF3CF')
        self.inventory_tree.tag_configure('#27AE60', background='#D5F4E6')
        self.inventory_tree.tag_configure('#95A5A6', background='#ECF0F1')
        
        # 根据状态筛选
        for record in records:
            # record[0]是i.id，record[1-4]是code,name,specification,warehouse
            # record[5-6]是quantity,available_quantity，record[7]是unit，record[8]是min_stock
            quantity = record[5]
            min_stock = record[8]
            
            if quantity == 0:
                status_text = '无库存'
                status_color = '#95A5A6'
            elif min_stock and quantity < min_stock:
                status_text = '库存不足'
                status_color = '#E74C3C'
            elif min_stock and quantity > min_stock * 2:
                status_text = '积压'
                status_color = '#F39C12'
            else:
                status_text = '正常'
                status_color = '#27AE60'

            # 应用状态筛选
            if status != '全部':
                # 状态映射
                status_map = {
                    '正常': '正常',
                    '库存不足': '库存不足',
                    '积压': '积压',
                    '无库存': '无库存'
                }
                if status_text not in status_map or status_map[status_text] != status:
                    continue

            self.inventory_tree.insert('', 'end', values=(
                record[1], record[2], record[3] or '', record[4],
                record[5], record[6], record[7], record[8] or 0, status_text
            ), tags=(status_color, f"inv_id_{record[0]}"))
        
        # 设置行颜色
        self.inventory_tree.tag_configure('#E74C3C', background='#FADBD8')
        self.inventory_tree.tag_configure('#F39C12', background='#FCF3CF')
        self.inventory_tree.tag_configure('#27AE60', background='#D5F4E6')
        self.inventory_tree.tag_configure('#95A5A6', background='#ECF0F1')
        
        # 更新统计
        self.update_inventory_stats()
        
        conn.close()

    def generate_report(self):
        """生成报表"""
        report_type = self.report_type.get()
        start_date = self.report_start.get()
        end_date = self.report_end.get()

        if not start_date or not end_date:
            messagebox.showwarning("提示", "请选择日期范围")
            return

        # 清空现有数据
        for item in self.report_tree.get_children():
            self.report_tree.delete(item)

        # 根据报表类型重新设置列
        if report_type == "销售报表":
            columns = ('日期', '单号', '客户', '商品编码', '商品名称', '数量', '单价', '金额')
        elif report_type == "采购报表":
            columns = ('日期', '单号', '供应商', '商品编码', '商品名称', '数量', '单价', '金额')
        elif report_type == "库存报表":
            columns = ('商品编码', '商品名称', '规格型号', '仓库', '库存数量', '可用库存', '单位', '最小库存', '状态')
        elif report_type == "利润报表":
            columns = ('日期', '商品编码', '商品名称', '数量', '采购价', '销售价', '利润')
        else:
            columns = ('日期', '单号', '客户/供应商', '商品编码', '商品名称', '数量', '单价', '金额')

        # 重新配置列
        self.report_tree['columns'] = columns
        for col in self.report_tree['columns']:
            self.report_tree.heading(col, text=col)
            self.report_tree.column(col, width=100)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        try:
            if report_type == "销售报表":
                cursor.execute('''
                    SELECT oo.order_date, oo.order_no, oo.customer, p.code, p.name, 
                           od.quantity, od.price, od.amount
                    FROM outbound_orders oo
                    JOIN outbound_details od ON oo.id = od.order_id
                    JOIN products p ON od.product_id = p.id
                    WHERE oo.order_date >= ? AND oo.order_date <= ? AND oo.status = 1
                    ORDER BY oo.order_date DESC, oo.order_no
                ''', (start_date, end_date))

                for record in cursor.fetchall():
                    self.report_tree.insert('', 'end', values=(
                        record[0], record[1], record[2], record[3], record[4],
                        record[5], f"¥{float(record[6]):.2f}", f"¥{float(record[7]):.2f}"
                    ))

            elif report_type == "采购报表":
                cursor.execute('''
                    SELECT io.order_date, io.order_no, io.supplier, p.code, p.name,
                           id.quantity, id.price, id.amount
                    FROM inbound_orders io
                    JOIN inbound_details id ON io.id = id.order_id
                    JOIN products p ON id.product_id = p.id
                    WHERE io.order_date >= ? AND io.order_date <= ? AND io.status = 1
                    ORDER BY io.order_date DESC, io.order_no
                ''', (start_date, end_date))

                for record in cursor.fetchall():
                    self.report_tree.insert('', 'end', values=(
                        record[0], record[1], record[2], record[3], record[4],
                        record[5], f"¥{float(record[6]):.2f}", f"¥{float(record[7]):.2f}"
                    ))

            elif report_type == "库存报表":
                cursor.execute('''
                    SELECT p.code, p.name, p.specification, w.name,
                           i.quantity, i.available_quantity, p.unit, p.min_stock
                    FROM inventory i
                    JOIN products p ON i.product_id = p.id
                    JOIN warehouses w ON i.warehouse_id = w.id
                    WHERE p.status = 1
                    ORDER BY w.name, p.code
                ''')

                for record in cursor.fetchall():
                    status = '不足' if record[4] < record[7] else '正常'
                    self.report_tree.insert('', 'end', values=(
                        record[0], record[1], record[2] or '', record[3],
                        record[4], record[5], record[6], record[7], status
                    ))

            elif report_type == "利润报表":
                cursor.execute('''
                    SELECT oo.order_date, p.code, p.name, od.quantity,
                           COALESCE((SELECT id.price FROM inbound_details id 
                                    WHERE id.product_id = p.id ORDER BY id.id DESC LIMIT 1), 0) as purchase_price,
                           od.price as sale_price,
                           (od.price - COALESCE((SELECT id.price FROM inbound_details id 
                                               WHERE id.product_id = p.id ORDER BY id.id DESC LIMIT 1), 0)) * od.quantity as profit
                    FROM outbound_orders oo
                    JOIN outbound_details od ON oo.id = od.order_id
                    JOIN products p ON od.product_id = p.id
                    WHERE oo.order_date >= ? AND oo.order_date <= ? AND oo.status = 1
                    ORDER BY oo.order_date DESC
                ''', (start_date, end_date))

                for record in cursor.fetchall():
                    purchase = float(record[4] or 0)
                    sale = float(record[5] or 0)
                    profit = float(record[6] or 0)
                    self.report_tree.insert('', 'end', values=(
                        record[0], record[1], record[2], record[3],
                        f"¥{purchase:.2f}", f"¥{sale:.2f}", f"¥{profit:.2f}"
                    ))

            # 保存当前报表类型和日期范围，用于导出
            self.current_report_type = report_type
            self.current_start_date = start_date
            self.current_end_date = end_date
            self.current_report_columns = columns

            messagebox.showinfo("成功", f"{report_type}生成成功！")

        except Exception as e:
            messagebox.showerror("错误", f"生成报表失败：{str(e)}")
        finally:
            conn.close()

    def export_report_to_excel(self):
        """导出报表到Excel"""
        if not hasattr(self, 'current_report_type') or not self.current_report_type:
            messagebox.showwarning("提示", "请先生成报表")
            return

        if not HAS_OPENPYXL:
            messagebox.showerror("错误", "未安装openpyxl库，无法导出Excel\n\n请运行：pip install openpyxl")
            return

        filename = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not filename:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = self.current_report_type

            # 获取报表数据
            headers = list(self.current_report_columns) if hasattr(self, 'current_report_columns') else []
            data_rows = []

            for item in self.report_tree.get_children():
                values = self.report_tree.item(item)['values']
                data_rows.append(values)

            # 写入表头
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 写入数据
            for row_idx, row_data in enumerate(data_rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 调整列宽
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

            wb.save(filename)
            messagebox.showinfo("成功", f"报表已导出到：\n{filename}")

        except Exception as e:
            messagebox.showerror("错误", f"导出Excel失败：{str(e)}")

    def run(self):
        """运行主窗口"""
        self.root.mainloop()

if __name__ == "__main__":
    login = LoginWindow()
    login.run()